from django.shortcuts import render, redirect
from main.models import Produit, Mark, Category, Supplier, Stockin, Itemsbysupplier, Client, Represent, Order, Orderitem, Clientprices, Bonlivraison, Facture, Outfacture, Livraisonitem, PaymentClientbl, PaymentClientfc,  PaymentSupplier, Bonsregle, Returnedsupplier, Avoirclient, Returned, Avoirsupplier, Orderitem, Carlogos, Ordersnotif, Connectedusers, Promotion, UserSession, Refstats, Notavailable, Cart, Wich, Devi, Notification, Modifierstock, Command, Notesrepresentant, Achathistory, Excelecheances, Bonsortie, Devisupplier, Commandsupplier, Avanceclient, Avancesupplier, Factureachat, Outfactureachat, Sortieitem, Caisse, Bank, DeviItem, CommandItem
from django.contrib.auth import logout
from django.http import JsonResponse, HttpResponse
import openpyxl
import qrcode
# import Count
from django.contrib.auth.models import User
from django.db.models import Count, F, Sum, Q, BooleanField, Case, When, Value
from django.db.models.functions import Cast
from django.contrib.sessions.models import Session
from functools import wraps
from django.contrib.auth.decorators import user_passes_test, login_required
import json
from django.contrib.auth.models import Group
from django.db import transaction
from datetime import datetime, date, timedelta
from django.utils import timezone
import pandas as pd
from itertools import chain, groupby
from django.core.serializers import serialize
import re
from django.contrib.auth import authenticate, login, logout
import requests as req
from collections import defaultdict
import calendar
from django.db.models.functions import TruncDay
import uuid
import ast

today = timezone.now().date()
thisyear=timezone.now().year



def isadmin(user):
    if not user.groups.filter(name='admin').exists():
        redirect('main:logoutuser')

# call isadmin here so that it executes before executing all the views

def getproductsbycategory(request):
    # category = Category.objects.get(pk=request.POST.get('category'))
    # products = category.product.filter(category=category)[:10]
    # get ten products from the category
    target=request.GET.get('target')
    products = Produit.objects.filter(category__pk=request.GET.get('category'))
    # get marks of the products filtered
    marks = Mark.objects.filter(produit__in=products).distinct().annotate(count=Count('produit'))
    ctx={
        'products':products,
        'home':False,
        'target':target,
        'marks':marks,
        'isfarah':target=='f',
        'target':target
    }
    return JsonResponse({
        'data':render(request, 'stocktrs.html', ctx).content.decode('utf-8'),
        # 'stocktotal':products.aggregate(Sum('stocktotal'))['stocktotal__sum'] or 0,
        # 'stockfacture':products.aggregate(Sum('stockfacture'))['stockfacture__sum'] or 0,
    })

def adminaddproductpage(request):
    categories=Category.objects.all()
    marques=Mark.objects.all()
    ctx={'categories':categories,
         'marques':marques,
         'commercials':Represent.objects.all(),
         'carlogos':Carlogos.objects.all(),
        }
    return render(request, 'addproduct.html', ctx)



def categoriespage(request):
    ctx={
        'categories':Category.objects.all().order_by('code'),
        #'commercials':Represent.objects.all(),
        'title':'Categories'
    }
    return render(request, 'categories.html', ctx)

def createcategory(request):
    name=request.POST.get('categoryname')
    code=request.POST.get('categorycode')
    affichage=request.POST.get('categoryaffichage')
    hideclient=request.POST.get('hideclient')=='True'
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)

    # get image file
    image=request.FILES.get('categoryimage')
    # create category
    category=Category.objects.create(name=name, image=image, code=code, affichage=affichage, masqueclients=hideclient)
    if len(commercialexcluded) > 0:
        category.excludedrep.set(reps)
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'title':'Categories'
    }
    print({

        'name':name,
        'code':code,
        'affichage':affichage,
        'hideclient':hideclient,
        'commercialexcluded':commercialexcluded,
        # get image file
        'image':category.image.url.replace('/media/', '') if category.image else ''
    })
    # req.get('http://serverip/products/createcategory', {

    #     'name':name,
    #     'code':code,
    #     'affichage':affichage,
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     # get image file
    #     'image':category.image.url.replace('/media/', '') if category.image else ''
    # })
    return JsonResponse({
        'html':render(request, 'categories.html', ctx).content.decode('utf-8')
    })

def updatecategory(request):
    print(request.POST.get('updatecategoryaffichage'))
    image=request.FILES.get('updatecategoryimage') or None
    id=request.POST.get('id')
    hideclient=request.POST.get('hideclient')=='True'
    print('>>>>>>>in updtae',request.POST.get('hideclient'))
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    category=Category.objects.get(pk=id)
    category.masqueclients=hideclient
    category.excludedrep.clear()
    category.excludedrep.set(reps)
    category.name=request.POST.get('updatecategoryname')
    category.code=request.POST.get('updatecategorycode')
    category.affichage=request.POST.get('updatecategoryaffichage')
    if image:
        category.image=image

    category.save()
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'title':'Categories'
    }
    # req.get('http://serverip/products/updatecategory', {
    #     'id':id,
    #     'image':category.image.url.replace('/media/', '') if category.image else '',
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     'name':request.POST.get('updatecategoryname'),
    #     'code':request.POST.get('updatecategorycode'),
    #     'affichage':request.POST.get('updatecategoryaffichage'),
    # })
    return JsonResponse({
        'html':render(request, 'categories.html', ctx).content.decode('utf-8')
    })


def marquespage(request):
    ctx={
        'marques':Mark.objects.all(),
        'title':'List des marques',
    }
    return render(request, 'marques.html', ctx)

def createmarque(request):
    name=request.POST.get('marquename')
    # get image file
    image=request.FILES.get('marqueimage')
    # create category
    hideclient=request.POST.get('hideclientmrk')=='True'
    commercialexcluded=request.POST.getlist('commercialexcludedmrk')
    reps=Represent.objects.filter(pk__in=commercialexcluded)

    mrq=Mark.objects.create(name=name, image=image, masqueclients=hideclient)
    if len(commercialexcluded) > 0:
        mrq.excludedrep.set(reps)
    # req.get('http://serverip/products/createmarque', {
    #     'name':name,
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     # get image file
    #     'image':mrq.image.url.replace('/media/', '') if mrq.image else ''
    # })
    return JsonResponse({
        'success':True
    })

def updatemarque(request):
    image=request.FILES.get('image') or None
    id=request.POST.get('id')
    hideclient=request.POST.get('hideclientmrk')=='True'
    commercialexcluded=request.POST.getlist('commercialexcludedmrk')
    reps=Represent.objects.filter(pk__in=commercialexcluded)

    mark=Mark.objects.get(pk=id)
    mark.name=request.POST.get('name')
    mark.masqueclients=hideclient
    mark.excludedrep.set(reps)
    if image:
        mark.image=image
    mark.save()
    # req.get('http://serverip/products/updatemarque', {
    #     'id':id,
    #     'name':request.POST.get('name'),
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     # get image file
    #     'image':mark.image.url.replace('/media/', '') if mark.image else ''
    # })
    ctx={
        'marques':Mark.objects.all(),
        'title':'List des marques'
    }
    return JsonResponse({
        'html':render(request, 'marques.html', ctx).content.decode('utf-8')
    })

def checkref(request):
    ref=request.POST.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref)
    notavailable=Notavailable.objects.filter(ref=ref).first()
    print(ref, notavailable)
    if product:
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False,
        'name':notavailable.name if notavailable else '',
        'equiv':notavailable.equiv if notavailable else '',
    })

    # print(ref, category, product)
    # if product:
    # else:
    #     return JsonResponse({
    #         'exist':False
    #     })

def supplierspage(request):
    target=request.GET.get('target')
    lastid=Supplier.objects.last()
    print(lastid)
    if lastid:
        lastid=lastid.id
    else:
        lastid=0
    code=f'FOR{lastid+1}'
    print(code)
    ctx={
        'suppliers':Supplier.objects.all(),
        'target':target,
        'title':'List des fournisseurs'
    }
    return render(request, 'suppliers.html', ctx)

def addsupplier(request):
    name=request.POST.get('suppnameinp')
    
    if Supplier.objects.filter(name=name).exists():
        return JsonResponse({
            'success':False
        })

    personalname=request.POST.get('supppersonalnameinp')
    # dont di this error again
    #image=request.POST.get('suppimageinp')
    image = request.FILES.get('suppimageinp')
    print('image>>', image)
    phone=request.POST.get('suppphone')
    phone2=request.POST.get('suppphone2')
    address=request.POST.get('supplieraddress')
    city=request.POST.get('suppliercity')
    email=request.POST.get('supplieremail')
    note=request.POST.get('note')
    sold=request.POST.get('sold') or 0
    ice=request.POST.get('suppice')
    suppif=request.POST.get('supplierif')
    rc=request.POST.get('supplierrc')
    plafon=request.POST.get('supplierplafon') or 0
    modereglement=request.POST.get('modereglement')
    print('>> name', name, 'phone', phone, 'address', address, 'image', image, 'personalname', personalname, 'phone2', phone2, 'rest', sold, 'note', note, 'city', city, 'email', email, 'ice', ice, 'suppif', suppif, 'rc', rc, 'modereglement', modereglement, 'plafon', plafon)
    lastid=Supplier.objects.last()
    print(lastid)
    if lastid:
        lastid=lastid.id
    else:
        lastid=0
    code=f'FOR{lastid+1}'
    Supplier.objects.create(
        code=code,
        name=name,
        phone=phone,
        address=address,
        image=image,
        personalname=personalname,
        phone2=phone2,
        rest=sold,
        note=note,
        city=city,
        email=email,
        ice=ice,
        suppif=suppif,
        rc=rc,
        modereglement=modereglement,
        plafon=plafon,

    )
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return JsonResponse({
        'success':True,
        #'html':render(request, 'suppliers.html', ctx).content.decode('utf-8')
    })

def getsupplierdata(request):
    id=request.POST.get('id')
    target=request.POST.get('target')
    supplier=Supplier.objects.get(pk=id)
    return render(request, 'editsupplier.html', {'supplier':supplier, 'target':target})
    # return JsonResponse({
    #     'name':supplier.name,
    #     'phone':supplier.phone,
    #     'address':supplier.address,
    #     'id':supplier.id
    # })


def updatesupplier(request):
    supplierid=request.POST.get('updatesuppid')
    image=request.FILES.get('updatesuppimage')
    name=request.POST.get('updatesuppname')
    personalname=request.POST.get('updatesupppersonalname')
    phone=request.POST.get('updatesuppphone')
    phone2=request.POST.get('updatesuppphone2')
    plafon=request.POST.get('updatesuppplafon') or 0
    ice=request.POST.get('updatesuppice')
    rc=request.POST.get('updatesupprc')
    suppif=request.POST.get('updatesuppname')
    city=request.POST.get('updatesuppcity')
    email=request.POST.get('updatesuppemail')
    address=request.POST.get('updatesuppaddress')
    modereglement=request.POST.get('updatesuppmodereglement')
    note=request.POST.get('updatesuppnote')
    print('>> suppid, name, phone, address, image, personalname, phone2, plafon, ice, rc, suppif, city, email, modereglement, note', supplierid, name, phone, address, image, personalname, phone2, plafon, ice, rc, suppif, city, email, modereglement, note, image)
    supplier=Supplier.objects.get(pk=supplierid)
    supplier.name=name
    supplier.phone=phone
    supplier.address=address
    supplier.personalname=personalname
    supplier.phone2=phone2
    supplier.plafon=plafon
    supplier.ice=ice
    supplier.rc=rc
    supplier.suppif=suppif
    supplier.city=city
    supplier.email=email
    supplier.modereglement=modereglement
    supplier.note=note
    if image:
        print('>> image', image)
        supplier.image=image
    supplier.save()
    return JsonResponse({
        'success':True
    })

def addoneproduct(request):
    target=request.POST.get('target')
    isfarah=target=='f'
    try:
        ref=request.POST.get('refinadd').lower().strip()
        name=request.POST.get('nameinadd').strip()
        category=request.POST.get('categoryinadd')
        unite=request.POST.get('unite')
        commercialsprix=request.POST.get('commercialsprix') or "[]"
        mark=request.POST.get('marqueinadd') or None
        logo=request.POST.get('logoinadd', None)
        image=request.FILES.get('imageinadd', None)
        sellprice=request.POST.get('sellpriceinadd') or 0
        qtyjeu=request.POST.get('qtyjeuinadd') or 0
        supplier=request.POST.get('supplier') or None
        minstock=request.POST.get('minstockinadd') or 0
        remise=request.POST.get('remiseinadd') or 0
        diametre=request.POST.get('diametreinadd') or ''
        representprice=request.POST.get('repprice') or None
        code=request.POST.get('codeinadd') or ''
        block=request.POST.get('blockinadd') or ''
        equivalent=request.POST.get('equivinadd') or ''
        cars=request.POST.get('carsinadd') or ''
        #netprice=round(float(sellprice)-(float(sellprice)*float(remise)/100), 2)
        # create product
        product=Produit.objects.create(
            ref=ref,
            name=name,
            qtyjeu=qtyjeu,
            buyprice=0,
            remise1=0,
            frbuyprice=0,
            frremise1=0,
            diametre=diametre,
            frsellprice=sellprice,
            frremisesell=remise,
            sellprice=sellprice,
            unite=unite,
            remisesell=remise,
            stocktotalfarah=0,
            stocktotalorgh=0,
            stockfacturefarah=0,
            stockfactureorgh=0,
            representprice=representprice,
            minstock=minstock,
            equivalent=equivalent,
            cars=cars,
            category_id=category,
            supplier_id=supplier,
            mark_id=mark,
            image=image,
            code=code,
            repsprice=commercialsprix,
            block=block,
            carlogos_id=logo,
            isactive=False,
            farahref='fr-'+ref
        )
        if isfarah:
            product.frsellprice=sellprice
            product.frremisesell=remise
        else:
            product.sellprice=sellprice
            product.remisesell=remise
        product.save()
        # req.get('http://serverip/products/addoneproduct', {
        #     'ref':ref,
        #     'name':name,
        #     'buyprice':buyprice,
        #     'diametre':diametre,
        #     'sellprice':sellprice,
        #     'remise':remise,
        #     'prixnet':netprice,
        #     'representprice':representprice,
        #     'minstock':minstock,
        #     'equivalent':equivalent,
        #     'cars':cars,
        #     'category':category,
        #     'supplier':supplier,
        #     'mark':mark,
        #     'image':product.image.url if product.image else '',
        #     'code':code,
        #     'repsprice':commercialsprix,
        #     'block':block,
        #     'carlogos_id':logo,
        #     'stocktotal':0,
        #     'stockfacture':0
        # })

        return JsonResponse({
            'success':True,

        })
    except Exception as e:
        print(e)
        return JsonResponse({
            'error':e
        })


def viewoneproduct(request, id):
    target=request.GET.get('target')
    isfarah=target=='f'
    product=Produit.objects.get(pk=id)
    stockin=Stockin.objects.filter(product=product, isfarah=isfarah, isavoir=False)
    # if isfarah:
    #     if product.frstockinitial:
    #         stockin+=product.frstockinitial
    # else:
    #     if product.stockinitial:
    #         stockin+=product.stockinitial
    outbl=Livraisonitem.objects.filter(product=product, isfacture=False, isfarah=isfarah).aggregate(Sum('qty'))['qty__sum'] or 0
    #outfacture=Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True).aggregate(Sum('qty'))['qty__sum'] or 0
    if target=='f':
        revbl=Livraisonitem.objects.filter(product=product, isfarah=True, isfacture=False).aggregate(Sum('total'))['total__sum'] or 0
        revfacture=Outfacture.objects.filter(product=product, isfarah=True).aggregate(Sum('total'))['total__sum'] or 0
    elif target=='o':
        revbl=Livraisonitem.objects.filter(product=product, isorgh=True, isfacture=False).aggregate(Sum('total'))['total__sum'] or 0
        revfacture=Outfacture.objects.filter(product=product, isorgh=True).aggregate(Sum('total'))['total__sum'] or 0
    else:
        revbl=Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('total'))['total__sum'] or 0
        revfacture=Outfacture.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
    totalout=outbl
    totalrev=round(revbl+revfacture, 2)
    if target=='f':
        stockout=Livraisonitem.objects.filter(product=product, isfarah=True, isfacture=False).order_by('-id')
        stockoutfc=Outfacture.objects.filter(product=product, isfarah=True).exclude(facture__bon__isnull=True).order_by('-id')
    elif target=='o':
        stockout=Livraisonitem.objects.filter(product=product, isorgh=True, isfacture=False).order_by('-id')
        stockoutfc=Outfacture.objects.filter(product=product, isorgh=True).exclude(facture__bon__isnull=True).order_by('-id')
    else:
        stockout=Sortieitem.objects.filter(product=product, isfacture=False).order_by('-id')
        stockoutfc=Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True).order_by('-id')
    #stockout=Livraisonitem.objects.filter(product=product, isfacture=False).order_by('-id')
    # stockoutfc=Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True).order_by('-id')
    avoirs=Stockin.objects.filter(product=product, isavoir=True, isfarah=target=='f', isorgh=target=='o')
    qtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    qtyavoir=avoirs.aggregate(Sum('quantity'))['quantity__sum'] or 0
    releve = chain(*[
    ((outbl, 'outbl') for outbl in stockout),
    ((outfc, 'outfc') for outfc in stockoutfc),
    ])
    thisproductreliquat=Orderitem.objects.filter(order__note__icontains='Reliquat', product=product, islivraison= False)


    # Sort the items by date
    outs = sorted(releve, key=lambda item: item[0].date)
    ctx={
        'isfarah':isfarah,
        'thisproductreliquat':thisproductreliquat,
        'outs':outs,
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'carlogos':Carlogos.objects.all(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':stockin,
        'sorties':stockout,
        'totalqtyin':qtyin+qtyavoir,
        'totalcout':stockin.aggregate(Sum('total'))['total__sum'] or 0,
        'totalqtyout':totalout,
        'totalcoutout':totalrev,
        'avoirs':avoirs,
        'reps':Represent.objects.all(),
        'today':timezone.now().date(),
        'target':target
    }
    return render(request, 'viewoneproduct.html', ctx)

def updateproduct(request):
    ref=request.POST.get('ref').lower().strip()
    target=request.POST.get('target')
    isfarah=target=='f'
    productid=request.POST.get('productid')
    product=Produit.objects.filter(ref=ref).exclude(pk=productid).first()
    if product:
        return JsonResponse({
            'success':False,
            'error':'Ref exist deja'
            })
    image=request.FILES.get('image') or None
    remise=request.POST.get('remise') or 0
    sellprice=request.POST.get('sellprice') or 0
    #netprice=round(float(sellprice)-(float(sellprice)*float(remise)/100), 2)
    product=Produit.objects.get(pk=productid)
    # if float(sellprice) != float(product.sellprice):
    #     print('price changed')
    #     reliquas=wishlist.objects.filter(product=product)
    #     for i in reliquas:
    #         i.total=round(float(netprice)*float(i.qty), 2)
    #         i.save()
    #     cartitems=Cartitems.objects.filter(product=product)
    #     for i in cartitems:
    #         newtotal=round(float(netprice)*float(i.qty), 2)
    #         newcarttotal=i.cart.total-i.total+newtotal
    #         i.total=newtotal
    #         i.save()
    #         i.cart.total=newtotal
    #         i.cart.save()
    # equivalent=' '.join(i for i in request.POST.get('equivalent').split())
    equivalent=request.POST.get('equivalent')
    cars=request.POST.get('cars')
    minstock=request.POST.get('minstock') or 0
    qtyjeu=request.POST.get('qtyjeu') or 0
    product.minstock=minstock
    product.qtyjeu=qtyjeu
    product.equivalent=equivalent
    product.cars=cars
    # product.refeq1=request.POST.get('refeq1').strip()
    # product.refeq2=request.POST.get('refeq2').strip()
    # product.refeq3=request.POST.get('refeq3').strip()
    # product.refeq4=request.POST.get('refeq4').strip()
    # product.coderef=request.POST.get('updatecoderef')
    # product.representprice=request.POST.get('updaterepprice') or 0
    # product.representremise=request.POST.get('updaterepremise') or 0
    if isfarah:
        product.frsellprice=sellprice
        product.frremisesell=remise
    else:
        product.sellprice=sellprice
        product.remisesell=remise
    
    product.farahref='fr-'+ref
    #product.prixnet=netprice
    product.name=request.POST.get('name')
    # product.cars=json.dumps(request.POST.getlist('cars'))
    product.ref=ref
    product.category_id=request.POST.get('category')
    product.mark_id=request.POST.get('marque')
    product.diametre=request.POST.get('diametre')
    product.block=request.POST.get('block')

    if image:
        product.image=image
    product.save()
    # data={
    #     #'image':product.image.url.replace('/media/', '') if product.image else '/media/default.png',
    #     'new':True if request.POST.get('switch')=='on' else False,
    #     'logo':request.POST.get('updatepdctlogo'),
    #     'productid':request.POST.get('productid'),
    #     'remise':request.POST.get('remise'),
    #     'sellprice':request.POST.get('sellprice'),
    #     'netprice':round(float(sellprice)-(float(sellprice)*float(remise)/100), 2),
    #     'equivalent':equivalent,
    #     'near':near,
    #     'code':request.POST.get('updatecode'),
    #     'refeq1':request.POST.get('refeq1'),
    #     'refeq2':request.POST.get('refeq2'),
    #     'refeq3':request.POST.get('refeq3'),
    #     'refeq4':request.POST.get('refeq4'),
    #     'repprice':request.POST.get('updaterepprice') or 0,
    #     # 'coderef':request.POST.get('updatecoderef'),
    #     'name':request.POST.get('name'),
    #     'cars':json.dumps(request.POST.getlist('cars')),
    #     'ref':request.POST.get('ref').lower().strip(),
    #     'category_id':request.POST.get('category'),
    #     'mark_id':request.POST.get('marque'),
    #     'diametre':request.POST.get('diametre'),
    #     'stock':product.stocktotal
    # }
    # if image:
    #     print('sending new image')
    #     data['image']=product.image.url.replace('/media/', '') if product.image else '/media/default.png',
    # print('>>end ',product)
    # print('>>>>>>>>>>>>>> equivalent>',equivalent)

    # res=req.get('http://serverip/products/updateproduct', data)
    # print('>>>>>>', res)
    # if res.status_code == 400:
    #         print('Error message:', res.text)
    # print('>>>>>>', request.POST.getlist('cars'))
    # req.get('http://serverip/products/updatepdctdata', {
    #     'password':'gadwad123',
    #     'id':request.POST.get('productid'),
    #     'ref':request.POST.get('ref').lower().strip(),
    #     'stocktotal':product.stocktotal,
    #     'cars':json.dumps(request.POST.getlist('cars')),
    # })
    return JsonResponse({
        'success':True
    })

def alertstock(request):
    target=request.GET.get('target')
    if target=='f':
        targets = Category.objects.filter(produit__stocktotalfarah__lte=F('produit__minstock')).annotate(
    total_products=Count('produit')
    )
    else:
        targets = Category.objects.filter(produit__stocktotalorgh__lte=F('produit__minstock')).annotate(
    total_products=Count('produit')
    )
    return render(request, 'alertstock.html', {'title':'Alert Stock', 'categories':targets,
    'suppliers':Supplier.objects.all(),
    'target':target})

def getlowbycategory(request):
    category=request.GET.get('category')
    supplierid=request.GET.get('supplierid')
    target=request.GET.get('target')
    print('supplierid >>>>>>>>>><', supplierid)
    #products = Product.objects.filter(category_id=category_id, stock__lte=F('minstock'), originsupp_id=supplierid)
    isfarah=target=='f'
    if isfarah:
        products = Produit.objects.filter(category_id=category, stocktotalfarah__lte=F('minstock')).annotate(
        is_preferred_supplier=Case(
            When(froriginsupp_id=supplierid, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        )).order_by('-is_preferred_supplier')
        marks=set([i.mark for i in products])
        marks=[{"name":i.name if i else '', 'id':i.id if i else ''} for i in marks]
        ctx={
            'products':products,
            'marks':marks
        }
        return JsonResponse({
        'data':render(request, 'fralertstocktrs.html', ctx).content.decode('utf-8')
    })
    else:
        products = Produit.objects.filter(category_id=category, stocktotalorgh__lte=F('minstock')).annotate(
        is_preferred_supplier=Case(
            When(originsupp_id=supplierid, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        )).order_by('-is_preferred_supplier')
    #suppliers=Supplier.objects.all()
    marks=set([i.mark for i in products])
    marks=[{"name":i.name if i else '', 'id':i.id if i else ''} for i in marks]
    ctx={
        'products':products,
        'marks':marks
    }
    return JsonResponse({
        'data':render(request, 'alertstocktrs.html', ctx).content.decode('utf-8')
    })

def commandsupplier(request):
    productid=request.POST.get('productid')
    supplierid=request.POST.get('supplierid')
    qty=request.POST.get('qty')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=qty
    product.suppliercommand_id=supplierid
    product.iscommanded=True
    product.save()
    return JsonResponse({
        'success':True
    })

def cacelcommand(request):
    productid=request.POST.get('productid')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=0
    product.suppliercommand_id=None
    product.iscommanded=False
    product.save()
    return JsonResponse({
        'success':True
    })

def recevoir(request):
    from random import randint
    target=request.GET.get('target')
    isfarah=target=='f'
    lastid=Itemsbysupplier.objects.last()
    print(lastid)
    if lastid:
        lastid=lastid.id
    else:
        lastid=0
    if isfarah:
        bonno=f'FR-BA00{lastid+1}'


    else:
        bonno=f'BA00{lastid+1}'

    print('>>>>>>', bonno)
    return render(request, 'recevoir.html', {'title':"Bon d'achat", 'suppliers':Supplier.objects.all(), 'today':timezone.now().date(), "target":target, 'bonno':bonno})

def bonlivraison(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it
    target=request.GET.get('target')
    year = timezone.now().strftime("%y")
    latest_receipt = Bonlivraison.objects.filter(
        bon_no__startswith=f'BL{year}'
    ).order_by("-bon_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.bon_no[-5:])
        receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"BL{year}00001"
    print('>>>>>>', receipt_no)
    return render(request, 'bonlivraison.html', {
        'title':'Bon de livraison',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        # 'commercials':Represent.objects.all(),
        'target':target
    })

def facture(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it

    return render(request, 'facture.html', {
        'title':'Facture',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        # 'commercials':Represent.objects.all(),
        'target':request.GET.get('target')
        #'order_no':receipt_no
    })


def suppliercommanproducts(request):
    supplierid=request.POST.get('supplierid')
    products=Produit.objects.filter(suppliercommand_id=supplierid)
    return JsonResponse({
        'data':render(request, 'suppliercommandproducts.html', {'products':products}).content.decode('utf-8')
    })


def searchref(request):
    ref=request.POST.get('ref')
    products=Produit.objects.filter(ref__istartswith=ref)
    return JsonResponse({
        'data':render(request, 'productsbon.html', {'products':products}).content.decode('utf-8')
    })

def addsupply(request):
    user=request.user
    supplierid=request.POST.get('supplierid')
    mantant=json.loads(request.POST.get('mantant'))
    moderegl=json.loads(request.POST.get('moderegl'))
    npiece=json.loads(request.POST.get('npiece'))
    bank=json.loads(request.POST.get('bank'))
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    # finish this
    products=request.POST.get('products')
    target=request.POST.get('target')
    devid=request.POST.get('devid')
    cmndid=request.POST.get('cmndid')
    nbon=request.POST.get('nbon')
    datebon=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    datefacture=datetime.strptime(request.POST.get('datefacture'), '%Y-%m-%d')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    isfarah=target=='f'
    print('>>>>>>', 'target', target, 'devid', devid, 'cmndid', cmndid, 'nbon', nbon, 'datebon', datebon, 'datefacture', datefacture, 'isfacture', isfacture, 'totalbon', totalbon, 'supplierid', supplierid, 'products', products)
    

    supplier=Supplier.objects.get(pk=supplierid)
    # supplier.rest=float(supplier.rest)+float(totalbon)
    # supplier.save()
    tva=0
    # if isfacture:
    #     print('>> creating facture')
    #     facture=Factureachat.objects.create(
    #         user=user,
    #         facture_no=nbon,
    #         supplier_id=supplierid,
    #         isfarah=isfarah,
    #         date=datefacture,
    #         total=totalbon
    #     )
    #     tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    # else:
    #     
    # 
    bon=Itemsbysupplier.objects.create(
        #user=user,
        isfarah=True if target=='f' else False,
        isorgh=True if target=='o' else False,
        supplier_id=supplierid,
        total=totalbon,
        date=datefacture,
        nbon=nbon,
    )
    if request.user.is_anonymous:
        print('>> user is anonym')
        bon.user=None
    else:
        bon.user=request.user
    bon.save()
    if not devid == "":
        devi=Devisupplier.objects.get(pk=devid)
        devi.generatedbl=True
        if isfacture:
            devi.facture=facture
            devi.save()
        else:
            devi.bl=bon
            devi.save()
    if not cmndid == "":
        cmnd=Commandsupplier.objects.get(pk=cmndid)
        cmnd.generatedbl=True
        if isfacture:
            cmnd.facture=facture
            cmnd.save()
        else:
            cmnd.bl=bon
            cmnd.save()
    for i in json.loads(products):
        product=Produit.objects.get(pk=i['productid'])
        remise1=0 if i['remise1']=='' else float(i['remise1'])
        remise2=0 if i['remise2']=='' else float(i['remise2'])
        remise3=0 if i['remise3']=='' else float(i['remise3'])
        remise4=0 if i['remise4']=='' else float(i['remise4'])

        buyprice=float(i['price'])
        print('>>> buyprice', buyprice, buyprice-(buyprice*(remise1/100)))
        # netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
        netwithremise1=round(buyprice-(buyprice*(remise1/100)), 2)
        netprice=round(float(i['total'])/float(i['qty']), 2)
        print('>>> net price', netprice)
        #product.isnew=True
        
        # if isfacture:
        #      Outfactureachat.objects.create(
        #         facture=facture,
        #         remise1=remise1,
        #         remise2=remise2,
        #         remise3=remise3,
        #         remise4=remise4,
        #         price=i['price'],
        #         ref=i['ref'],
        #         name=i['name'],
        #         product=product,
        #         qty=i['qty'],
        #         total=i['total'],
        #         supplier_id=supplierid,
        #         date=datebon,
        #         isfarah=target=='f',
        #         isorgh=target=='o',
        #     )
        # else:
        Stockin.objects.create(
            date=datefacture,
            product=product,
            quantity=i['qty'],
            price=i['price'],
            ref=i['ref'],
            name=i['name'],
            remise1=remise1,
            remise2=remise2,
            remise3=remise3,
            remise4=remise4,
            net=netwithremise1,
            # remise=remise,
            qtyofprice=i['qty'],
            total=i['total'],
            supplier_id=supplierid,
            isfacture=isfacture,
            isfarah=target=='f',
            isorgh=target=='o',
            nbon=bon
        )
        if target=='f':
            # if product.isnegativeinfr:
            #     print('>> product is negative')
            #     qtynegative=json.loads(product.frnegative)
            #     sortitems=json.loads(product.frsorties)
            #     print('>> ', qtynegative, sortitems)
            #     thisqty=int(i['qty'])
            #     #for , ss in zip(qtynegative, sortitems):
            #     for index, (qq, ss) in enumerate(zip(qtynegative, sortitems)):
            #         print('>> thisqty, qq', thisqty, qq)
            #         if not int(qq)==0:
            #             if not thisqty<=0:
            #                 # find sortie and add qty and price
            #                 sorti=Sortieitem.objects.get(pk=ss)
            #                 prices=json.loads(sorti.pricesofout)
            #                 qtyofout=json.loads(sorti.qtyofout)
            #                 if thisqty>=qq:
            #                     print(">>thisqty>=qq", thisqty>=qq, qq)
            #                     qtyofout.append(qq)
            #                     prices.append(st.id)
            #                     diff=int(thisqty)-int(qq)
            #                     thisqty=diff
            #                     st.qtyofprice=diff
            #                     st.save()
            #                     qtyofout[index]=0
            #                     product.save()
            #                     # qtynegative.remove(qq)
            #                     # sortitems.remove(ss)
            #                 else:
            #                     diff=int(qq)-int(thisqty)
            #                     qtyofout.append(diff)
            #                     prices.append(st.id)
            #                     st.qtyofprice=0
            #                     st.save()
            #                     qtyofout[index]=diff
            #                 sorti.pricesofout=prices
            #                 sorti.qtyofout=qtyofout
            #                 sorti.save()
            #                     # finish
            #         if sum(qtynegative)==0:
            #             print('>> sum', sum(qtynegative), qtynegative)
            #             product.isnegativeinfr=False
            #         product.save()            

                    
            # calcul pondiré, stock needs to be more than 0
            # if product.stocktotalfarah>0:
            #     print('>> has stock')
            #     totalqtys=int(product.stocktotalfarah)+int(i['qty'])
            #     actualtotal=product.stocktotalfarah*product.frnetbuyprice
            #     print('totalqty, actualtotal', totalqtys, actualtotal)
            #     # remainingstock=Stockin.objects.filter(qtyofprice__gt=0, product=product, isfarah=True, isavoir=False)
            #     # for b in remainingstock:
            #     #     actualtotal+=float(b.price)*float(b.qtyofprice)
            #     thistotal=int(i['qty'])*netprice
            #     print('>>>>>> thistotal', thistotal)
            #     totalprices=round(float(i['total'])+actualtotal, 2)
            #     pondire=round(totalprices/totalqtys, 2)
            #     product.frcoutmoyen=pondire
            #     product.save()
            #     print('>> coout m', pondire)
            # else:
            #print('>> cooutm', netprice)
            #product.frcoutmoyen=netprice
            product.frremise1=remise1
            product.frremise2=remise2
            product.frremise3=remise3
            product.frremise4=remise4
            product.frbuyprice=buyprice
            product.froriginsupp_id=supplierid
            product.frnetbuyprice=netprice
            print('>> addin qty')
            product.stocktotalfarah=float(product.stocktotalfarah)+float(i['qty'])
            # product.frsellprice=buyprice
            # product.frremisesell=remise1
            # if isfacture:
            #     product.stockfacturefarah=int(product.stockfacturefarah)+int(i['qty'])
        else:
            # if product.isnegative:
            #     qtynegative=json.loads(product.negative)
            #     sortitems=json.loads(product.sorties)
            #     thisqty=int(i['qty'])
                
            #     #for , ss in zip(qtynegative, sortitems):
            #     for index, (qq, ss) in enumerate(zip(qtynegative, sortitems)):
            #         if not int(qq)==0:
            #             if not thisqty<=0:
            #                 # find sortie and add qty and price
            #                 sorti=Sortieitem.objects.get(pk=ss)
            #                 prices=json.loads(sorti.pricesofout)
            #                 qtyofout=json.loads(sorti.qtyofout)
            #                 if thisqty>=qq:
            #                     qtyofout.append(qq)
            #                     prices.append(st.id)
            #                     diff=int(thisqty)-int(qq)
            #                     thisqty=diff
            #                     st.qtyofprice=diff
            #                     st.save()
            #                     qtyofout[index]=0
            #                     product.save()
            #                     # qtynegative.remove(qq)
            #                     # sortitems.remove(ss)
            #                 else:
            #                     diff=int(qq)-int(thisqty)
            #                     qtyofout.append(diff)
            #                     prices.append(st.id)
            #                     st.qtyofprice=0
            #                     st.save()
            #                     qtyofout[index]=diff
            #                 sorti.pricesofout=prices
            #                 sorti.qtyofout=qtyofout
            #                 sorti.save()
            #                     # finish
            #         if sum(qtynegative)==0:
            #             product.isnegative=False
            #         product.save()
            
            # if product.stocktotalorgh>0:
            #     totalqtys=int(product.stocktotalorgh)+int(i['qty'])
            #     actualtotal=product.stocktotalorgh*product.netbuyprice
            #     # remainingstock=Stockin.objects.filter(qtyofprice__gt=0, product=product, isfarah=False)
            #     # for i in remainingstock:
            #     #     actualtotal+=float(i.price)*float(i.qtyofprice)
            #     thistotal=int(i['qty'])*buyprice
            #     totalprices=round(float(i['total'])+actualtotal, 2)
            #     pondire=round(totalprices/totalqtys, 2)
            #     product.coutmoyen=pondire
            #     product.save()
            #     print('>> coout m', pondire)
            # else:
            #     print('>> cooutm', netprice)
            #product.coutmoyen=netprice
            product.remise1=remise1
            product.remise2=remise2
            product.remise3=remise3
            product.remise4=remise4
            product.buyprice=buyprice
            product.netbuyprice=netprice
            product.originsupp_id=supplierid
            product.stocktotalorgh=float(product.stocktotalorgh)+float(i['qty'])
            # product.sellprice=buyprice
            # product.remisesell=remise1
        
            # if isfacture:
            #     product.stockfactureorgh=int(product.stockfactureorgh)+int(i['qty'])
        # recodrd remise 1, 2, 3, 4
        product.save()
    # # update cout moyen, it will be calculated by deviding total prices by total qty

        # totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
        # totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
        #product.coutmoyen=round(totalprices/totalqty, 2)
        product.qtycommande=0
        product.save()
        ### this was reglement
        # totalamount=sum(x if x is not None else 0 for x in mantant)
        # if float(totalamount)==float(totalbon):
        #     if isfacture:
        #         facture.ispaid=True
        #         facture.save()
        #     else: 
        #         bon.ispaid=True
        #         bon.save()

        # else:
        #     diff=float(totalbon)-float(totalamount)
        #     if isfacture:
        #         facture.rest=diff 
        #         facture.save()
        #     else: 
        #         bon.rest=diff
        #         bon.save()

        # if totalamount>0:
        #     for m, mod, np, ech, bk in zip(mantant, moderegl, npiece, echeance, bank):
        #         if m is not None:
        #             p=PaymentSupplier.objects.create(
        #                 supplier_id=supplierid,
        #                 amount=m,
        #                 date=timezone.now(),
        #                 echeance=ech,
        #                 mode=mod,
        #                 npiece=np,
        #                 bank=bk,
        #                 isfarah=isfarah,
        #             )
        #             if isfacture:
        #                 p.set.facture([facture])
        #             else:
        #                 p.set.bons([bon])

    return JsonResponse({
        'success':True
        #'html': render(request, 'recevoir.html', {'title':'Recevoir Les produits', 'suppliers':Supplier.objects.all(), 'products':Produit.objects.all()}).content.decode('utf-8')
    })

@login_required(login_url='main:home')
def addbonlivraison(request):
    # mantant=request.POST.get('mantant')
    # mode=request.POST.get('mode')
    # npiece=request.POST.get('npiece')
    # echeance=request.POST.get('echeance')
    
    #current_time = datetime.now().strftime('%H:%M:%S')
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    devid=request.POST.get('devid')
    comndid=request.POST.get('cmndid')
    # orderno
    transport=request.POST.get('transport')

    note=request.POST.get('note')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')
    print('>>', "clientid", clientid, "repid", repid, "products", products, "totalbon", totalbon, "devid", devid, "comndid", comndid, "transport", transport, "note", note, "datebon", datebon, "target", target)
    client=Client.objects.get(pk=clientid)
    print('cleint', client, 'soldtotal', client.soldtotal, 'totalbon', totalbon)
    

    # get the last bon no
    year = timezone.now().strftime("%y")
    isfarah=target=='f'
    isorgh=target=='o'
    print('isfarah, target', isfarah, target)
    if isfarah:
        latest_receipt = Bonlivraison.objects.filter(
            bon_no__startswith=f'FR-BL{year}'
        ).last()
        # latest_receipt = Bonsortie.objects.filter(
        #     bon_no__startswith=f'FR-BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.bon_no[-9:])
            receipt_no = f"FR-BL{year}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"FR-BL{year}000000001"
    else:
        latest_receipt = Bonlivraison.objects.filter(
            bon_no__startswith=f'BL{year}'
        ).last()
        # latest_receipt = Bonsortie.objects.filter(
        #     bon_no__startswith=f'BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.bon_no[-9:])
            receipt_no = f"BL{year}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"BL{year}000000001"
    order=Bonlivraison.objects.create(
        command_id=comndid,
        devi_id=devid,
        client_id=clientid,
        salseman_id=repid,
        total=totalbon,
        date=datebon,
        modlvrsn=transport,
        bon_no=receipt_no,
        notebon=note,
        isfarah=isfarah,
        isorgh=isorgh,
    )
    # if user is anonym
    if request.user.is_anonymous:
        print('>> user is anonym')
        order.user=None
    else:
        order.user=request.user
    order.save()
    if not comndid == "":
        cmnd=Command.objects.get(pk=comndid)
        cmnd.generatedbl=True
        cmnd.bl=order
        cmnd.save()
    if not devid == "":
        cmnd=Devi.objects.get(pk=devid)
        cmnd.generatedbl=True
        cmnd.bl=order
        cmnd.save()
    print('>>>>>>', len(json.loads(products))>0)
    with transaction.atomic():
        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            pricesofout=[]
            qtyofout=[]
            try:
                coutm = i['coutmoyen']
            except:
                coutm = 0
            # achatids=i['achatids'].split(',')
            # remainqties=i['remainqties'].split(',')
            # oldqties=i['oldqties'].split(',')
            if isfarah:
                print('>>> we are in farah')
                product.stocktotalfarah=float(product.stocktotalfarah)-float(i['qty'])
            else:
                thisqty=float(i['qty'])
                product.stocktotalorgh=float(product.stocktotalorgh)-float(i['qty'])      
            product.save()
            Livraisonitem.objects.create(
                qtyofout=qtyofout,
                pricesofout=pricesofout,
                bon=order,
                remise=i['remise'],
                name=i['name'],
                ref=i['ref'],
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
                client_id=clientid,
                date=datebon,
                isfarah=isfarah,
                # achatids=achatids,
                # remainqties=remainqties,
                # oldqties=oldqties,
                coutmoyen=coutm,
            )
    #order.pricesofout=pricesofout
    order.save()
    # client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    # client.soldbl=round(float(client.soldbl)+float(totalbon), 2)
    # client.save()
    # if mantant=="":
    #     return JsonResponse({
    #         "success":True
    #     })
    # if float(mantant)==0:
    #     return JsonResponse({
    #         "success":True
    #     })
    # if echeance=="":
    #     echeance=None
    # else:
    #     echeance=datetime.strptime(echeance, '%Y-%m-%d')
    # if float(mantant)>float(totalbon):
    #     diff=float(mantant)-float(totalbon)
    #     # for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
    #     regl=PaymentClientbl.objects.create(
    #         client_id=clientid,
    #         amount=mantant,
    #         date=datebon,
    #         echance=echeance,
    #         mode=mode,
    #         npiece=npiece,
    #         isfarah=isfarah,
    #         isorgh=isorgh
    #     )
    #     regl.bon=order
    #     order.ispaid=True
    #     Avanceclient.objects.create(
    #         client_id=clientid,
    #         amount=diff,
    #         date=datebon,
    #         isfarah=isfarah,
    #         isorgh=isorgh,
    #         bon=order
    #     )
    # elif float(mantant)==float(totalbon):
    #     # for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
    #     regl=PaymentClientbl.objects.create(
    #         client_id=clientid,
    #         amount=mantant,
    #         date=datebon,
    #         echance=echeance,
    #         mode=mode,
    #         npiece=npiece,
    #         isfarah=isfarah,
    #         isorgh=isorgh
    #     )
    #     regl.bon=order
    #     order.ispaid=True
    # else:
    #     diff=float(totalbon)-float(mantant)
    #     # for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
    #     regl=PaymentClientbl.objects.create(
    #         client_id=clientid,
    #         amount=mantant,
    #         date=datebon,
    #         echance=echeance,
    #         mode=mode,
    #         npiece=npiece,
    #         isfarah=isfarah,
    #         isorgh=isorgh
    #     )
    #     regl.bon=order
    #     order.rest=diff
    # #diff=float(mantant)-float(totalbon)
    # increment it
    order.save()
    return JsonResponse({
        "success":True
    })

# add facture not generer
def addfacture(request):
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    transport=request.POST.get('transport', '')
    note=request.POST.get('note', '')
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    isfarah=False
    isorgh=False
    if target=='f':
        isfarah=True
        latest_receipt = Facture.objects.filter(
            facture_no__startswith=f'FR-FC{year}'
        ).last()
        # latest_receipt = Bonsortie.objects.filter(
        #     facture_no__startswith=f'FR-BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.facture_no[-9:])
            receipt_no = f"FR-FC{year}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"FR-FC{year}000000001"
    else:
        isorgh=True
        latest_receipt = Facture.objects.filter(
            facture_no__startswith=f'FC{year}'
        ).last()
        # latest_receipt = Bonsortie.objects.filter(
        #     facture_no__startswith=f'FR-BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.facture_no[-9:])
            receipt_no = f"FC{year}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"FC{year}000000001"
    client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    client.soldfacture=round(float(client.soldfacture)+float(totalbon), 2)
    client.save()
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    year = timezone.now().strftime("%y")

    print('>>>>>>>',latest_receipt)
    facture=Facture.objects.create(
        facture_no=receipt_no,
        total=totalbon,
        tva=tva,
        date=datebon,
        client_id=clientid,
        transport=transport,
        note=note
    )
    if len(json.loads(products))>0:
        with transaction.atomic():
            for i in json.loads(products):
                product=Produit.objects.get(pk=i['productid'])
                if isfarah:
                    product.stockfacture=int(product.stockfacture)-int(i['qty'])
                product.save()
                Outfacture.objects.create(
                    facture=facture,
                    remise=i['remise'],
                    name=i['name'],
                    ref=i['ref'],
                    product=product,
                    qty=i['qty'],
                    price=i['price'],
                    total=i['total'],
                    client_id=clientid,
                    date=datebon,
                )

    # year = timezone.now().strftime("%y")
    # latest_receipt = Facture.objects.filter(
    #     facture_no__startswith=f'FC{year}'
    # ).order_by("-facture_no").first()
    # latest_receipt_no = int(latest_receipt.facture_no[-5:])
    # receipt_no = f"FC{year}{latest_receipt_no + 1:05}"

    # increment it
    return JsonResponse({
        'success':True
    })



def supplierinfo(request, id):
    supplier=Supplier.objects.get(pk=id)
    target=request.GET.get('target')
    isfarah=target=='f'
    avances=Avancesupplier.objects.filter(supplier=supplier, isfarah=isfarah)
    print('>> ssssss')
    ctx={
        'target':target,
        'title':f'Info fournisseur {supplier.name.upper}',
        'supplier':supplier,
        'totalavoirs':Avoirsupplier.objects.filter(supplier=supplier, isfarah=isfarah, ispaid=False).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentSupplier.objects.filter(supplier=supplier, isfarah=isfarah, isavoir=False).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totaltr':Itemsbysupplier.objects.filter(supplier=supplier, isfarah=isfarah).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':Itemsbysupplier.objects.filter(supplier=supplier, isfarah=isfarah),
        'payments':PaymentSupplier.objects.filter(supplier=supplier, isfarah=isfarah, isavoir=False),
        'factures':Factureachat.objects.filter(supplier=supplier, isfarah=isfarah),
        'avances':avances,
        'totalavances':avances.aggregate(Sum('amount'))['amount__sum'] or 0,
        'devis':Devisupplier.objects.filter(supplier=supplier, isfarah=isfarah),
        'avoirs':Avoirsupplier.objects.filter(supplier=supplier, isfarah=isfarah),
        'commandes':Commandsupplier.objects.filter(supplier=supplier, isfarah=isfarah),
    }
    return render(request, 'supplierinfo.html', ctx)

def clientinfo(request, id):
    target=request.GET.get('target')
    isfarah=target=='f'
    client=Client.objects.get(pk=id)
    if target=='s':
        bons=Bonsortie.objects.filter(client=client, total__gt=0)
    else:
        bons=Bonlivraison.objects.filter(client=client, total__gt=0)
    ctx={
        'target':target,
        'client':client,
        'totalavoirs':Avoirclient.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentClientbl.objects.filter(client=client).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totalbons':Bonlivraison.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':bons,
        'payments':PaymentClientbl.objects.filter(client=client),
        'avoirs':Avoirclient.objects.filter(client=client),
        'avances':Avanceclient.objects.filter(client=client),
        'factures':Facture.objects.filter(client=client),
        'devis':Devi.objects.filter(client=client),
        'commandes':Command.objects.filter(client=client),
        'nbrbons':bons.count(),
        'title':f'Compte client: {client.name}'
    }
    return render(request, 'clientinfo.html', ctx)

def addpaymentssupp(request):
    supplierid=request.POST.get('supplierid')
    pass


def dashboard(request):

    ctx={
        'title':'Dashboard',
        'orders':Order.objects.filter(date__date=date.today()).count(),
        'products':Produit.objects.all().count(),
        'productthismonth':Orderitem.objects.filter(order__date__month=date.today().month).order_by('-qty')[:20],
        'alerts':Produit.objects.filter(stocktotal__lte=F('minstock')).count(),
        'blnotpaid':Bonlivraison.objects.filter(ispaid=False).count(),
        'boncommand':Order.objects.filter(isdelivered=False).count(),
        'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),



    }
    return render(request, 'pdashboard.html', ctx)

def clientspage(request):
    # sortie=request.GET.get('sortie')=='1'
    # farah=request.GET.get('farah')=='1'
    # orgh=request.GET.get('orgh')=='1'
    # if sortie:
    #     clients=Client.objects.filter(clientsortie=True).order_by('-soldtotal')[:50]
    # if farah:
    #     clients=Client.objects.filter(clientfarah=True).order_by('-soldtotal')[:50]
    # if orgh:
    #     clients=Client.objects.filter(clientorgh=True).order_by('-soldtotal')[:50]

    target=request.GET.get('target')
    print('>> terget', target)
    print('faracl', Client.objects.filter(clientfarah=True).count())
    if target=='s':
        try:
            lastcode = Client.objects.filter(code__startswith='CP-').last()
            print('lastcode', lastcode.code)
            if lastcode:
                codecl = f"CP-{int(lastcode.code.split('-')[1]) + 1}"
            else:
                codecl = f"CP-1"
        except:
            codecl="CP-1"
        clients=Client.objects.filter(clientsortie=True).order_by('-soldtotal')[:50]
    elif target=='f':
        try:
            lastcode = Client.objects.filter(code__startswith='CF').last()
            print('lastcode', lastcode.code)
            if lastcode:

                codecl = f"CF-{int(lastcode.code.split('-')[1]) + 1}"
            else:
                codecl = f"CF-1"
        except:
            codecl="CF-1"
        clients=Client.objects.filter(clientfarah=True).order_by('-soldtotal')[:50]
    else:
        try:
            lastcode = Client.objects.filter(code__startswith='CO').last()
            print('lastcode', lastcode.code)
            if lastcode:

                codecl = f"CO-{int(lastcode.code.split('-')[1]) + 1}"
            else:
                codecl = f"CO-1"
        except:
            codecl="CO-1"
        clients=Client.objects.filter(clientorgh=True).order_by('-soldtotal')[:50]
    ctx={
        'clients':clients,
        'title':'List des clients',
        'lastcode':codecl,
        'target':target
        # 'sortiesection':sortie,
        # 'farahsection':farah,
        # 'orghsection':orgh,
        # 'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),
        # 'soldbl':round(Client.objects.aggregate(Sum('soldbl'))['soldbl__sum'] or 0, 2),
        # 'soldfacture':round(Client.objects.aggregate(Sum('soldfacture'))['soldfacture__sum'] or 0, 2),
    }
    return render(request, 'clients.html', ctx)



def checkusername(request):
    username=request.POST.get('username')
    if User.objects.filter(username=username).exists():
        return JsonResponse({
            'exist':True
        })
    else:
        return JsonResponse({
            'exist':False
        })

def commercialspage(request):
    ctx={
        'commercials':Represent.objects.all(),
        'title':'List des Commeciaux'
    }
    return render(request, 'commerciaux.html', ctx)

def addcommercial(request):
    repusername=request.POST.get('repusername')
    reppassword=request.POST.get('reppassword')
    repname=request.POST.get('repname')
    repphone=request.POST.get('repphone')
    repregion=request.POST.get('repregion')
    repinfo=request.POST.get('repinfo')
    try:
        # request=req.get('http://serverip/products/addcommercial',{
        #     'repusername':repusername,
        #     'reppassword':reppassword,
        #     'repname':repname,
        #     'repphone':repphone,
        #     'repregion':repregion,
        #     'repinfo':repinfo
        # })
        # request.raise_for_status()
        user=User.objects.create_user(username=repusername, password=reppassword)
        # Get or create the group
        group, created = Group.objects.get_or_create(name="salsemen")

        # Add the user to the group
        user.groups.add(group)

        # Save the user
        user.save()
        Represent.objects.create(
            user=user,
            name=repname,
            phone=repphone,
            region=repregion,
            info=repinfo
        )
        # old code 04/07/2024
        # ctx={
        #     'commercials':Represent.objects.all(),
        #     'title':'List des Commeciaux'
        # }
        # return JsonResponse({
        #     'html':render(request, 'commerciaux.html', ctx).content.decode('utf-8')
        # })
        return JsonResponse({
            'success':True
        })
    except:
        return JsonResponse({
            'success':False,
            'message': 'ERROR CONEXION'
        })

def checkcodeclient(request):
    code=request.GET.get('code')
    name=request.GET.get('name')
    print(Client.objects.filter(Q(name=name) | Q(code=code)).exists())
    if Client.objects.filter(Q(name=name) | Q(code=code)).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })
#this to add clients that are divers
def addclientdivers(request):
    target=request.GET.get('target')
    name=request.GET.get('name')
    code=request.GET.get('code')
    city=request.GET.get('ville')
    city=request.GET.get('ville')
    client=Client.objects.create(
        city=city,
        code=code,
        name=name,
        soldtotal=0.00,
        soldfacture=0.00,
        soldbl=0.00,
        diver=True
    )

    if target=='f':
        client.clientfarah=True
    elif target=='s':
        client.clientsortie=True
    else:
        client.clientorgh=True
    return JsonResponse({
        'succes':True
    })



def addclient(request):
    name=request.POST.get('clientnameinp')
    target=request.POST.get('target')
    phone=request.POST.get('clientphone')
    phone2=request.POST.get('clientphone2')
    address=request.POST.get('clientaddress')
    code=request.POST.get('clientcode')
    city=request.POST.get('clientcity')
    clientemail=request.POST.get('clientemail')
    ice=request.POST.get('clientice')
    clientif=request.POST.get('clientif')
    clientnamepersinp=request.POST.get('clientnamepersinp')
    clientrc=request.POST.get('clientrc')
    modereglement=request.POST.get('modereglement')
    note=request.POST.get('note')
    clientsold=request.POST.get('clientsold') or 0
    #region=request.POST.get('clientregion').lower().strip()
    plafon=request.POST.get('clientplafon') or 0
    isfarah=target=='f'
    if Client.objects.filter(Q(name=name) | Q(code=code), clientfarah=isfarah).exists():
        print('>>> client exists', Client.objects.filter(Q(name=name) | Q(code=code), clientfarah=isfarah))
        return JsonResponse({
            'success':False,
            'error':'Code ou Nom exist deja'
        })
    try:
        # response=req.get('http://serverip/products/addclient', {
        #     'city':city,
        #     'ice':ice,
        #     'region':region,
        #     'represent_id':representant,
        #     'code':code,
        #     'name':name,
        #     'phone':phone,
        #     'address':address,
        # })
        # response.raise_for_status()
        client=Client.objects.create(
            city=city,
            ice=ice,
            #region=region,
            plafon=plafon,
            code=code,
            name=name,
            phone=phone,
            phone2=phone2,
            address=address,
            diver=False,
            soldbl=clientsold,
            note=note,
            modereglement=modereglement,
            clientif=clientif,
            clientrc=clientrc,
            email=clientemail,
            clientname=clientnamepersinp,
            soldfacture=0.00,
        )
        if target=="s":
            client.clientsortie=True
        elif target=="f":
            client.clientfarah=True
        else:
            client.clientorgh=True
        client.save()

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print('>>', e)
        return JsonResponse({
            'success':False,
            'error':e
        })

def getclientdata(request):
    id=request.GET.get('id')
    print('>>> id cilent data ', id)
    target=request.GET.get('target')
    print('>> target getdata client', target)
    client=Client.objects.get(pk=id)
    return JsonResponse({
        'html':render(request, 'clientmodaldata.html', {'client':client, 'target':target}).content.decode('utf-8'),
        'sold':client.sold().sold
    })

    # return JsonResponse({
    #     'personalname':client.clientname,
    #     'name':client.name,
    #     'phone':client.phone,
    #     'phone2':client.phone2,
    #     'address':client.address,
    #     'id':client.id,
    #     'code':client.code,
    #     'city':client.city,
    #     #'location':client.location,
    #     'region':client.region,
    #     'ice':client.ice,
    #     'plafon':client.plafon,
    #     #'rep':client.represent_id,
    # })


def updateclient(request):
    id=request.POST.get('updateclientid')
    code=request.POST.get('updateclientcode')
    name=request.POST.get('updateclientname')
    client=Client.objects.get(pk=id)
    if Client.objects.filter(name=name).exclude(pk=id).exists():
         return JsonResponse({
             'success':False,
             'error':'Code ou Nom exist deja'
         })
    client.name=request.POST.get('updateclientname')
    client.phone=request.POST.get('updateclientphone')
    client.city=request.POST.get('updateclientcity')
    client.clientname=request.POST.get('updateclientpersonalname')
    client.phone2=request.POST.get('updateclientphone2')
    client.ice=request.POST.get('updateclientice')
    client.plafon=request.POST.get('updateclientplafon') or 0
    client.clientrc=request.POST.get('updateclientrc')
    client.clientif=request.POST.get('updateclientif')
    client.city=request.POST.get('updateclientcity')
    client.modereglement=request.POST.get('modereglement')
    client.address=request.POST.get('updateclientaddress')
    client.save()
    # req.get('http://serverip/products/updateclient', {
    #     'clientcode':oldcode,
    #     'name':request.POST.get('updateclientname'),
    #     'phone':request.POST.get('updateclientphone'),
    #     'address':request.POST.get('updateclientaddress'),
    #     'ice':request.POST.get('updateclientice'),
    #     'code':request.POST.get('updateclientcode'),
    #     'city':request.POST.get('updateclientcity'),
    #     'address':request.POST.get('updateclientaddress'),
    #     'region':request.POST.get('updateclientregion'),
    #     'rep':request.POST.get('updateclientrep'),
    # })
    return JsonResponse({
        'success':True
    })


def getscommercialdata(request):
    id=request.POST.get('id')
    rep=Represent.objects.get(pk=id)
    return JsonResponse({
        'name':rep.name,
        'phone':rep.phone,
        'phone2':rep.phone2,
        'region':rep.region,
        'info':rep.info,
        'id':rep.id
    })

def updatecommercial(request):
    id=request.POST.get('updaterepid')
    name=request.POST.get('updaterepname')
    phone=request.POST.get('updaterepphone')
    phone2=request.POST.get('updaterepphone2')
    region=request.POST.get('updaterepregion')
    region=request.POST.get('updaterepregion')
    rep=Represent.objects.get(pk=id)
    rep.name=name
    rep.phone=phone
    rep.phone2=phone2
    rep.region=region
    rep.save()
    ctx={
        'commercials':Represent.objects.all(),
        'title':'List des Commeciaux'
    }
    return JsonResponse({
        'html':render(request, 'commerciaux.html', ctx).content.decode('utf-8')
    })


def onereppage(request, id):
    rep=Represent.objects.get(pk=id)
    note=Notesrepresentant.objects.filter(represent=rep).first()
    print('>>>>>>>>', note)
    ctx={
        'title':f'Page de {rep.name}',
        'rep':rep,
        'notes':note.note if note else '',
        'today':timezone.now().date()
    }
    return render(request, 'onereppage.html', ctx)

def adminpage(request):
    return render(request, 'adminpage.html')

def bonlivraisondetails(request, id):
    target=request.GET.get('target')
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__name')
    print('orderitems', orderitems)
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    facture=order.facture
    reglementsfc=PaymentClientbl.objects.filter(factures__in=[facture])
    avances=Avanceclient.objects.filter(bonofavance=order.bon_no)
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+34] for i in range(0, len(orderitems), 34)]
    ctx={
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'avances':avances,
        'orderitems':orderitems,
        'reglements':reglements,
        'reglementsfc':reglementsfc,
        'reps':Represent.objects.all(),
        'target':target
    }
    return render(request, 'bonlivraisondetails.html', ctx)


def facturedetails(request, id):
    target=request.GET.get('target')
    order=Facture.objects.get(pk=id)
    
    
    bons=order.bons.all()
    total=bons.aggregate(Sum('total')).get('total__sum')
    print('>> total', total)
    order.total=total
    order.save()
    orderitems=Livraisonitem.objects.filter(bon__in=bons)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]
    reglements=PaymentClientbl.objects.filter(factures__in=[order])
    ctx={
        'reglements':reglements,
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'target':target,
        'ht':round(order.total-order.tva, 2),
        'reps':Represent.objects.all()
    }
    return render(request, 'facturedetails.html', ctx)

def supplierfacturedetails(request, id):
    target=request.GET.get('target')
    order=Factureachat.objects.get(pk=id)
    bons=order.bons.all()
    print('>> bons', bons)
    if bons:
        total=bons.aggregate(Sum('total')).get('total__sum')
        print('>> total', total)
        order.total=total
        order.save()
    # orderitems=Outfactureachat.objects.filter(facture=order).order_by('product__name')
    orderitems=Stockin.objects.filter(nbon__in=bons)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]
    reglements=PaymentSupplier.objects.filter(factures__in=[order])
    print('>>bons', bons)
    ctx={
        'reglements':reglements,
        'title':f'Facture achat {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        # 'tva':order.tva,
        # 'ttc':order.total,
        # 'ht':round(order.total-order.tva, 2),
        'target':target,
    }
    return render(request, 'supplierfacturedetails.html', ctx)


def avoirdetails(request, id):
    target=request.GET.get('target')
    order=Avoirclient.objects.get(pk=id)
    orderitems=Stockin.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'avoir {order.no}',
        'avoir':order,
        'target':target,
        'reglement':PaymentClientbl.objects.filter(avoirs__id=order.id),
        'orderitems':orderitems

    }
    return render(request, 'avoirdetails.html', ctx)

def avoirsuppdetails(request, id):
    target=request.GET.get('target')
    order=Avoirsupplier.objects.get(pk=id)
    orderitems=Returnedsupplier.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'avoir {order.no}',
        'target':target,
        'isfarah':target=='f',
        'avoir':order,
        'orderitems':orderitems,
        'reglement':PaymentSupplier.objects.filter(avoirs__id=order.id)
    }
    return render(request, 'avoirsuppdetails.html', ctx)



def getrepswithprice(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    repsprices=[]
    if product.repsprice:
        repsprices=json.loads(product.repsprice)
        print('get reps with price',json.loads(product.repsprice))
    return JsonResponse({
        'repswithprice': repsprices,
        'representremise':product.representremise
    })


def getclientprice(request):
    pdctid=request.POST.get('id')
    clientid=request.POST.get('clientid')
    mode=request.POST.get('mode')
    inavoir=request.POST.get('inavoir')=='true'
    target=request.POST.get('target')
    isfarah=target=='f'
    product=Produit.objects.get(pk=pdctid)
    client=Client.objects.get(pk=clientid)
    print('>>>>',request.POST.get('target'), client.name, product.ref, isfarah, mode, inavoir)
    price=0
    remise=0

    # try:
    if target=='s':
        producthistory=Sortieitem.objects.filter(bon__client_id=clientid, product_id=pdctid)
    # elif target=='f':
    #     producthistory=Livraisonitem.objects.filter(client_id=clientid, product_id=pdctid, isfarah=True)
    else:
        print('>> farah clientprice', clientid, pdctid, isfarah)
        producthistory=Livraisonitem.objects.filter(bon__client_id=clientid, product_id=pdctid, isfarah=isfarah)
    print('producthistory', producthistory)
    clientprice=producthistory.last()
    if clientprice:
        price=clientprice.price
        remise=clientprice.remise
    # if mode=='bl':
    # else:
    #     producthistory=Outfacture.objects.filter(client_id=clientid, product_id=pdctid)
    #     clientprice=producthistory.last()
    #     if clientprice:
    #         price=clientprice.price
    #         remise=clientprice.remise


    return JsonResponse({
        'price':price,
        'remise':remise,
        'table':render(request, 'prodctprices.html', {'producthistory':producthistory.order_by('-id'), 'isfarah':isfarah}).content.decode('utf-8')
    })
    # except Exception as e:
    #     print('error', e)
    #     return JsonResponse({
    #         'price':0,
    #         'remise':0,
    #     })
    #if target=='bl':
    #    try:
    #        clientprice=Livraisonitem.objects.filter(client_id=clientid, product_id=id).last()
    #        price=clientprice.price
    #        remise=clientprice.remise
    #        return JsonResponse({
    #            'price':price,
    #            'remise':remise
    #        })
    #    except:
    #        return JsonResponse({
    #            'price':0
    #        })
    #else:
    #    try:
    #        clientprice=Outfacture.objects.filter(client_id=clientid, product_id=id).last()
    #        price=clientprice.price
    #        remise=clientprice.remise
    #        return JsonResponse({
    #            'price':price,
    #            'remise':remise
    #        })
    #    except:
    #        return JsonResponse({
    #            'price':0
    #        })

def listbonlivraison(request):
    target=request.GET.get('target')
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    #depasser = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0).count()
    # get only the last 100 orders of the current year
    # only check one target as bon livraison is only for farah or orgh, pos has bonsortie
    if target=='f':
        bons= Bonlivraison.objects.filter(isfarah=True, isvalid=False, iscanceled=False).order_by('-bon_no')[:50]
        total=Bonlivraison.objects.filter(isfarah=True, isvalid=False, iscanceled=False).aggregate(Sum('total')).get('total__sum')
    else:
        bons= Bonlivraison.objects.filter(isfarah=False, isvalid=False, iscanceled=False).order_by('-bon_no')[:50]
        total=Bonlivraison.objects.filter(isfarah=False, isvalid=False, iscanceled=False).aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Bons de livraison',
        'bons':bons,
        'total':total,
        'all':1,
        #'boncommand':Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat').count(),
        #'depasser':depasser,
        #'reps':Represent.objects.all(),
        'today':timezone.now().date(),
        'target':target
    }
    return render(request, 'listbonlivraison.html', ctx)

def exportbl(request):
    rep=request.GET.get('rep')
    datefrom=request.GET.get('startdate')
    dateto=request.GET.get('enddate')
    region=request.GET.get('region').lower().strip()
    print('>>>>>>', rep, datefrom, dateto)
    if rep and region:
        print('rep and region')
        bons=bons=Bonlivraison.objects.filter(salseman_id=rep,client__region=region, date__range=[datefrom, dateto])
    if rep and not region:
        print('rep and not region')
        bons=bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    if not rep and region:
        print('not rep and region')
        bons=bons=Bonlivraison.objects.filter(client__region=region, date__range=[datefrom, dateto])
    if not region and not rep:
        print('nothing')
        bons=bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])

    # if rep and datefrom and dateto:
    #     print('date and rep')
    #     # convert date to datetime
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    # if rep and not datefrom and not dateto:
    #     print('only rep')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__year=timezone.now().year)
    # if not rep and datefrom and dateto:
    #     print('only date')
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])
    # if not rep and not datefrom and not dateto:
    #     print('nothing')
    #     bons=Bonlivraison.objects.filter(date__year=timezone.now().year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° Bon', 'Date', 'Client', 'Code cl.', 'total', 'region', 'ville', 'soldbl', 'rep.', 'status', 'facture', 'transport'])

    # Write product data
    for product in bons:
        ws.append([
            product.bon_no, product.date.strftime("%d/%m/%Y"), product.client.name, product.client.code, product.total, product.client.region, product.client.city, product.client.soldbl, product.salseman.name, 'R0' if product.ispaid else 'N1', 'OUI' if product.isfacture else 'NON', product.modlvrsn])

    response['Content-Disposition'] = f'attachment; filename="bonlivraison.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response


def exportfc(request):
    target=request.GET.get('target')
    year=request.GET.get('year')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    isfarah=target=='f'
    # today in d/m/Y
    today=timezone.now().strftime("%d-%m-%Y")
    society='FARAH' if target=='f' else 'ORGH'
    print(">> is farah", target=='f')
    bons=Facture.objects.filter(isfarah=isfarah, date__range=[startdate, enddate])
    
    print('bons', bons.count())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° facture', 'Date', 'Client', 'TOTAL TTC', 'HT', 'TVA', 'Mode reglement', 'N° piece', 'Banque'])

    # Write product data
    for bon in bons:
        ws.append([
            bon.facture_no,
            bon.date.strftime("%d/%m/%Y"),
            bon.client.name if bon.client else '--',
            
            bon.total,
            bon.ht(),
            bon.thistva(),
            bon.reglements()[0].mode if len(bon.reglements())>0 else '--',
            bon.reglements()[0].npiece if len(bon.reglements())>0 else '--',
            bon.reglements()[0].bank if len(bon.reglements())>0 else '--'
        ])

    response['Content-Disposition'] = f'attachment; filename="facture{society}{today}.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response


def exportfcachat(request):
    target=request.GET.get('target')
    year=request.GET.get('year')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    isfarah=target=='f'
    # today in d/m/Y
    today=timezone.now().strftime("%d-%m-%Y")
    society='FARAH' if target=='f' else 'ORGH'
    print(">> is farah", target=='f')
    bons=Factureachat.objects.filter(isfarah=isfarah, date__range=[startdate, enddate])
    
    print('bons', bons.count())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° facture', 'Date', 'fournisseur', 'TOTAL TTC', 'HT', 'TVA', 'Mode reglement', 'N° piece', 'Banque'])

    # Write product data
    for bon in bons:
        ws.append([
            bon.facture_no,
            bon.date.strftime("%d/%m/%Y"),
            bon.supplier.name if bon.supplier else '--',
            
            bon.total,
            bon.ht(),
            bon.thistva(),
            bon.reglements()[0].mode if len(bon.reglements())>0 else '--',
            bon.reglements()[0].npiece if len(bon.reglements())>0 else '--',
            bon.reglements()[0].bank if len(bon.reglements())>0 else '--'
        ])

    response['Content-Disposition'] = f'attachment; filename="facture{society}{today}.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response





def listavoirclient(request):
    target=request.GET.get('target')
    isfarah=target=='f'
    if target=='s':
        bons= Avoirclient.objects.filter(date__year=thisyear, issortie=True).order_by('-id')
    elif target=='o':
        bons= Avoirclient.objects.filter(date__year=thisyear, isorgh=True).order_by('-id')
    else:
        bons= Avoirclient.objects.filter(date__year=thisyear, isfarah=isfarah).order_by('-id')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Avoir Client',
        'bons':bons,
        'total':total,
        'target':target
    }
    return render(request, 'listavoirclient.html', ctx)


def listavoirsupplier(request):
    print('>>>>>>',)
    target=request.GET.get('target')
    bons= Avoirsupplier.objects.filter(isfarah=target=="f").order_by('-id')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Avoir Achat',
        'bons':bons,
        'total':total,
        'target':target
    }
    return render(request, 'listavoirsupplier.html', ctx)

def listfactures(request):
    target=request.GET.get('target')
    isfarah=target=='f'
    three_months_ago = timezone.now() - timedelta(days=90)
    depasser = Facture.objects.filter(date__lt=three_months_ago, ispaid=False).count()
    year=timezone.now().strftime("%y")
    # get only the last 100 orders of the current year
    if target=='f':
        bons= Facture.objects.filter(isvalid=False, isfarah=True).order_by('-facture_no')[:50]
        total=Facture.objects.filter(isvalid=False, isfarah=True).aggregate(Sum('total')).get('total__sum') or 0
        lastdatefacture=Facture.objects.filter(isvalid=False, isfarah=True).last().date if bons else timezone.now().date()
        latest_receipt = Facture.objects.filter(
            facture_no__startswith=f'FR-FC').last()
        # latest_receipt = Bonsortie.objects.filter(
        #     facture_no__startswith=f'FR-BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.facture_no[-9:])
            receipt_no = f"{latest_receipt.facture_no[:7]}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"FR-FC{year}000000001"
    else:
        bons= Facture.objects.filter(isvalid=False, isfarah=False).order_by('-facture_no')[:50]
        total=Facture.objects.filter(isvalid=False, isfarah=False).aggregate(Sum('total')).get('total__sum') or 0
        lastdatefacture=Facture.objects.filter(isvalid=False, isfarah=False).last().date if bons else timezone.now().date()
        latest_receipt = Facture.objects.filter(facture_no__startswith=f'FC').order_by("facture_no").last()
        print("==== last receit", latest_receipt.facture_no)
        # latest_receipt = Bonsortie.objects.filter(
        #     facture_no__startswith=f'FR-BL{year}'
        # ).order_by("-bon_no").first()
        if latest_receipt:
            latest_receipt_no = int(latest_receipt.facture_no[-9:])
            receipt_no = f"{latest_receipt.facture_no[:4]}{latest_receipt_no + 1:09}"
        else:
            receipt_no = f"FC{year}000000001"
    ctx={
        'total':total,
        'title':'List des factures',
        'bons':bons,
        'reps':Represent.objects.all(),
        'depasserfc':depasser,
        'today':timezone.now().date(),
        'target':target,
        'receipt_no':receipt_no,
        'lastdatefacture':lastdatefacture,
        'term':0
    }
    # if bons:
    #     ctx['total']=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     ctx['totaltva']=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    return render(request, 'listfactures.html', ctx)

def activerproduct(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    product.isactive=True
    product.save()
    ctx={
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':Stockin.objects.filter(product=product, qtyofprice__gt=0),
        'sorties':Orderitem.objects.filter(product=product),
    }
    # req.get('http://serverip/products/activerproduct', {
    #     'id':request.POST.get('id')
    # })
    return JsonResponse({
        'html':render(request, 'viewoneproduct.html', ctx).content.decode('utf-8')
    })

def desactiverproduct(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    product.isactive=False
    product.save()
    ctx={
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':Stockin.objects.filter(product=product),
        'sorties':Orderitem.objects.filter(product=product),
    }
    # req.get('http://serverip/products/desactiverproduct', {
    #     'id':request.POST.get('id')
    # })
    return JsonResponse({
        'html':render(request, 'viewoneproduct.html', ctx).content.decode('utf-8')
    })


def generate_qrcode(request):
    id=request.GET.get('id')
    target=request.GET.get('target')
    product=Produit.objects.get(pk=id)
    qr_text = (
        123
    )

    # Generate the QR Code
    qr = qrcode.QRCode(
        version=1,  # Controls the size of the QR code
        error_correction=qrcode.constants.ERROR_CORRECT_H,  # High error correction
        box_size=10,
        border=4,
    )
    qr.add_data(qr_text)
    qr.make(fit=True)

    # Save QR Code to a BytesIO buffer
    buffer = BytesIO()
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(buffer, format="PNG")

    # Convert image to Base64
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    buffer.close()

    # Determine date
    lastachat = StockIn.objects.filter(product=product).last()
    if achat:
        date = datetime.today().strftime('%m%y')
    else:
        date = lastachat.reciept.date.strftime('%m%y') if lastachat and lastachat.reciept else '-'

    # Dynamic text for display
    text = (
        f"{product.category.name.upper()} {product.ref.split()[0].upper()} "
        f"{product.mark.name.upper()} {price}/{date} {product.car}"
    )

    # Render the template with the QR code and additional data
    return render(request, 'products/barcode.html', {
        'barcodes': qr_code_base64,
        'text': text,
        'product': product,
        'date': date,
        'price': price,
    })

def generatefacture(request, id):
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison)
    lastdate=Facture.objects.last().date
    year = timezone.now().strftime("%y")
    latest_receipt = Facture.objects.filter(
        facture_no__startswith=f'FC{year}'
    ).order_by("-facture_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.facture_no[-5:])
        receipt_no = f"FC{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"FC{year}00001"
    ctx={
        'today':timezone.now().date(),
        'livraison':livraison,
        'items':items,
        'receipt_no':receipt_no,
        'lastdate':lastdate
    }
    return render(request, 'generatefacture.html', ctx)

# genereate facture from bl
def createfacture(request):
    bon=request.POST.get('bon')
    total=request.POST.get('totalbon')
    datefacture=request.POST.get('datefacture')
    datefacture=datetime.strptime(datefacture, '%Y-%m-%d')
    orderno=request.POST.get('orderno')
    products=json.loads(request.POST.get('products'))
    livraison=Bonlivraison.objects.get(pk=bon)
    livraison.isfacture=True
    livraison.statusfc='f1'

    # watch out for negative total
    livraison.total=round(livraison.total-float(total), 2)
    livraison.save()
    thisclient=Client.objects.get(pk=livraison.client_id)
    # we substract sold bl because we generate from bin livraisso
    thisclient.soldbl=round(thisclient.soldbl-float(total), 2)
    thisclient.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    thisclient.save()
    # list of ids in oredritems
    tva=round(float(total)-(float(total)/1.2), 2)
    facture=Facture.objects.create(
        bon_id=bon,
        facture_no=orderno,
        total=total,
        tva=tva,
        date=datefacture,
        client=livraison.client,
        salseman=livraison.salseman,
        printed=False
    )


    product_ids_to_remove = [i['productid'] for i in products]

    # Delete the matching Orderitem objects in a single transaction
    Livraisonitem.objects.filter(bon=livraison, product_id__in=product_ids_to_remove).update(isfacture=True)
    with transaction.atomic():

        for i in products:
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Outfacture.objects.create(
                facture=facture,
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
                remise=i['remise'],
                ref=i['ref'],
                name=i['name'],
                client=livraison.client,
                date=datefacture
            )
    # if livraison.ispaid:
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         # give reglement bl to regl fact
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()

    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             amountofeachbon=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #     else:
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             if i.amount<=round(float(total)-reglfacture, 2):
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=i.amount,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=i.amount
    #                 i.amount=0
    #                 i.usedinfacture=True
    #                 i.save()
    #             else:
    #                 wanted=round(float(total)-reglfacture, 2)
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=wanted,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=wanted
    #                 i.amount=round(float(i.amount)-wanted)
    #                 i.usedinfacture=True
    #                 i.save()

    #     # Facturesregle.objects.create(
    #     #     payment=regfac,
    #     #     bon=facture,
    #     #     amount=float(total)

    #     # )
    #     # finish substraction form reglement bon
    #     facture.ispaid=True
    #     facture.save()
    # elif livraison.rest > 0:
    #     if float(total)==float(livraison.total):
    #         reglements=PaymentClientbl.objects.filter
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()
    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             echance=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #         facture.ispaid=True
    #         facture.save()
    #         # new bon rest
    #         #total reglements

    #     else:
    #         # if onre of the reglements is not gte total of facture
    #         # we need to iterate over regl and sum up the regl
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             reglfacture+=i.amount
    #             i.amount=0
    #             i.usedinfacture=True
    #             i.save()
    #     totalreglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison).aggregate(Sum('amount'))['amount__sum']
    #     livraison.rest=round(livraison.total-totalreglements, 2)
    #     livraison.save()
    #     livraison.client.soldbl=round(livraison.client.soldbl-float(totalreglements), 2)
    # else:
    #     livraison.client.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    #     livraison.client.save()
    return JsonResponse({
        'success':True
    })


def degenerer(request):
    bonid=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=bonid)
    Livraisonitem.objects.filter(bon=livraison).update(isfacture=False)
    # delete facture and outfacture
    facture=Facture.objects.get(bon=livraison)
    outfactures=Outfacture.objects.filter(facture=facture)
    print(livraison, facture, outfactures)

    # products=Produit.objects.get(pk=i.product_id) for i in outfactures

    for i in outfactures:

        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()
    livraison.isfacture=False
    livraison.statusfc='b1'
    livraison.total=round(livraison.total+float(facture.total), 2)
    livraison.save()
    facture.delete()
    livraison.client.soldbl=round(livraison.client.soldbl+float(facture.total), 2)
    livraison.client.soldfacture=round(livraison.client.soldfacture-float(facture.total), 2)
    livraison.client.save()
    return JsonResponse({
        'html':render(request, 'bonlivraisonbody.html', {'order':livraison}).content.decode('utf-8')
    })

def modifierlivraison(request, id):
    target=request.GET.get('target')
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison, isfacture=False)
    ctx={
        'target':target,
        'title':'Modifier '+livraison.bon_no,
        'livraison':livraison,
        'items':items,
        'commercials':Represent.objects.all(),
        # 'products':Produit.objects.all(),
        # 'clients':Client.objects.all(),
    }
    return render(request, 'modifierlivraison.html', ctx)

def modifieravoir(request, id):
    target=request.GET.get('target')
    avoir=Avoirclient.objects.get(pk=id)
    items=Stockin.objects.filter(avoir=avoir)
    ctx={
        'title':'Mofier avoir '+ avoir.no,
        'target':target,
        'avoir':avoir,
        'items':items,
        'commercials':Represent.objects.all(),
    }
    return render(request, 'modifieravoir.html', ctx)


def modifieravoirsupp(request):
    target=request.GET.get('target')
    id=request.GET.get('id')
    avoir=Avoirsupplier.objects.get(pk=id)
    items=Returnedsupplier.objects.filter(avoir=avoir)
    ctx={
        'title':'Mofier avoir Fournisseur'+ avoir.no,
        'target':target,
        'avoir':avoir,
        'items':items,
        'commercials':Represent.objects.all(),
    }
    return render(request, 'modifieravoirsupp.html', ctx)


def modifierfacture(request, id):
    facture=Facture.objects.get(pk=id)
    items=Outfacture.objects.filter(facture=facture)
    try:
        lastcode = Client.objects.order_by('code').last()
        print('lastcode', lastcode.code)
        if lastcode:

            codecl = f"{int(lastcode.code) + 1:06}"
        else:
            codecl = f"000001"
    except:
        codecl="000001"
    ctx={
        'lastcode':codecl,
        'facture':facture,
        'items':items,
        'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
        'clients':Client.objects.all(),
    }
    return render(request, 'modifierfacture.html', ctx)


def updatebonlivraison(request):
    id=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=id)

    client=Client.objects.get(pk=request.POST.get('clientid'))
    totalbon=request.POST.get('totalbon')
    target=request.POST.get('target')
    transport=request.POST.get('transport')
    note=request.POST.get('note')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')
    thisclient=livraison.client
    #facture of bon
    facture=Facture.objects.filter(bons__in=[livraison]).first()
    if facture:
        print('>>> has facture', facture.total, 'updating total', round(float(facture.total)-float(livraison.total)+float(totalbon), 2))
        newtotal=round(float(facture.total)-float(livraison.total)+float(totalbon), 2)
        facture.total=newtotal
        facture.save()
    # if livraison.client==client:
    #     print('same client', float(thisclient.soldtotal), float(livraison.total), float(totalbon))
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total)+float(totalbon), 2)
    #     thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total)+float(totalbon), 2)
    #     thisclient.save()
    # else:
    #     print('not same')
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total), 2)
    #     thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total), 2)
    #     thisclient.save()
    #     print('old', thisclient.soldtotal)
    #     client.soldtotal+=float(totalbon)
    #     client.soldbl+=float(totalbon)
    #     client.save()
    #     print('new', client.soldtotal)
    
    livraison.modlvrsn=transport
    livraison.client=client
    livraison.notebon=note
    # livraison.salseman_id=request.POST.get('repid')
    livraison.total=totalbon
    livraison.date=datebon
    # livraison.bon_no=request.POST.get('orderno')
    livraison.save()
    items=Livraisonitem.objects.filter(bon=livraison)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        if target=='f':
            print('>>> deleting old bon items in farah')
            product.stocktotalfarah=float(product.stocktotalfarah)+float(i.qty)
            
            
        else:
            print('>>> deleting old bon items in orgh')
            product.stocktotalorgh=float(product.stocktotalorgh)+float(i.qty)
            # st=Stockin.objects.filter(isfarah=False, product=product).last()
            # st.qtyofprice=st.qtyofprice+int(i.qty)
            # st.save()
        #st=Stockin.objects.filter(pk__in=json.loads(i.pricesofout)).last()
        #st.qtyofprice=st.qtyofprice+int(i.qty)
        #st.save()
        product.save()
        i.delete()

    print('client:', livraison.client.id)
    for i in json.loads(request.POST.get('products')):
        # ABORTER FOR NOW
        # clientpricehistory=Clientprices.objects.filter(client_id=livraison.client.id, product_id=i['productid']) or None
        # if clientpricehistory:
        #     print('clientpricehistory exist')
        #     clientpricehistory.update(price=i['price'])
        # else:
        #     print('clientpricehistory not exist')
        #     Clientprices.objects.create(client_id=livraison.client.id, product_id=i['productid'], price=i['price'])


        qty=float(i['qty'])
        product=Produit.objects.get(pk=i['productid'])
        pricesofout=[]
        qtyofout=[]
        # achatids=ast.literal_eval(i['achatids'])
        # remainqties=ast.literal_eval(i['remainqties'])
        # oldqties=ast.literal_eval(i['oldqties'])
        if target=='f':
            product.stocktotalfarah=float(product.stocktotalfarah)-qty
            
        else:
            product.stocktotalorgh=float(product.stocktotalorgh)-qty
            
        product.save()
        # if i.get('bonitemid', None):
        #     st=Livraisonitem.objects.filter(pk=i['bonitemid']).first()
        # # if not itemid means it's a new line
        #     print('>>> already, updating old date')
        #     st.remise=i['remise'],
        #     st.name=i['name'],
        #     st.ref=i['ref'],
        #     st.qty=qty,
        #     st.price=i['price'],
        #     st.total=i['total'],
        #     st.date=datebon,
        #     st.save()
        # else:
        #     print('>> new line')
        # create new livraison items
        Livraisonitem.objects.create(
            # qtyofout=qtyofout,
            # pricesofout=pricesofout,
            bon=livraison,
            remise=i['remise'],
            name=i['name'],
            ref=i['ref'],
            product=product,
            qty=qty,
            price=i['price'],
            total=i['total'],
            date=datebon,
            isfarah=target=='f',
            client=client,
            # achatids=achatids,
            # remainqties=remainqties,
            # oldqties=oldqties,
            # coutmoyen=i['coutmoyen'],
        )
        # if achatids:
        #     stockins=Stockin.objects.filter(pk__in=achatids)
        #     for s, r in zip(stockins, remainqties):
        #         s.qtyofprice=r
        #         s.save()

    return JsonResponse({
        'success':True
    })


def updatebonfacture(request):
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=request.POST.get('clientid'))
    facture=Facture.objects.get(pk=request.POST.get('factureid'))
    totalbon=request.POST.get('totalbon')
    thisclient=facture.client
    if facture.client==client:
        print('same client', float(thisclient.soldtotal), float(facture.total), float(totalbon))
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total)+float(totalbon), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total)+float(totalbon), 2)
        thisclient.save()
    else:
        print('not same')
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total), 2)
        thisclient.save()
        print('old', thisclient.soldtotal)
        client.soldtotal+=float(totalbon)
        client.soldfacture+=float(totalbon)
        client.save()
        print('new', client.soldtotal)
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    facture.tva=tva
    facture.client=client
    facture.salseman_id=request.POST.get('repid')
    facture.total=totalbon
    facture.facture_no=request.POST.get('orderno')
    facture.note=request.POST.get('note')
    facture.date=datebon
    facture.save()
    items=Outfacture.objects.filter(facture=facture)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()

    print('client:', facture.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            # update price in facture
            # clientpricehistory=Clientprices.objects.filter(client_id=facture.client.id, product_id=i['productid']) or None
            # if clientpricehistory:
            #     print('clientpricehistory exist')
            #     clientpricehistory.update(price=i['price'])
            # else:
            #     print('clientpricehistory not exist')
            #     Clientprices.objects.create(client_id=facture.client.id, product_id=i['productid'], price=i['price'])


            qty=int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-qty

            product.save()

            # create new livraison items
            Outfacture.objects.create(
                facture=facture,
                remise=i['remise'],
                name=i['name'],
                ref=i['ref'],
                product=product,
                qty=qty,
                price=i['price'],
                total=i['total'],
                client=client,
                date=datebon,
            )

    return JsonResponse({
        'success':True
    })


def listreglementbl(request):
    target=request.GET.get('target')
    if target=="f":
        reglements=PaymentClientbl.objects.filter(isfarah=True).order_by('-id')[:50]
    elif target=="o":
        reglements=PaymentClientbl.objects.filter(isorgh=True).order_by('-id')[:50]
    else:
        reglements=PaymentClientbl.objects.filter(issortie=True).order_by('-id')[:50]
    print('lenreg', len(reglements))
    ctx={
        'title':'List des reglements CL BL' +target,
        'reglements':reglements,
        'today':timezone.now().date(),
        'target':target,
        'banks':Bank.objects.all()
    }
    # if reglements:
    #     ctx['total']=round(PaymentClientbl.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)

    return render(request, 'listreglementbl.html', ctx)

def listavances(request):
    target=request.GET.get('target')
    if target=="f":
        avances=Avanceclient.objects.filter(isfarah=True).order_by('-id')
    elif target=="o":
        avances=Avanceclient.objects.filter(isorgh=True).order_by('-id')
    else:
        avances=Avanceclient.objects.filter(issortie=True).order_by('-id')
    print('lenreg', len(avances))
    ctx={
        'title':'List des avances CL BL',
        'avances':avances,
        'today':timezone.now().date(),
        'target':target,
    }
    # if avances:
    #     ctx['total']=round(Avanceclient.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)

    return render(request, 'listavances.html', ctx)

def addavanceclient(request):
    clientid=request.GET.get('clientid')
    bonofavance=request.GET.get('bonofavance')
    client=Client.objects.get(pk=clientid)
    
    target=request.GET.get('target')
    mantant=json.loads(request.GET.get('mantant'))
    client.soldtotal-=sum(mantant)
    client.soldbl-=sum(mantant)
    client.save()
    bank=json.loads(request.GET.get('bank'))
    mode=json.loads(request.GET.get('mode'))
    npiece=json.loads(request.GET.get('npiece'))
    note=json.loads(request.GET.get('note'))
    # date=datetime.strptime(request.GET.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.GET.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    print("clientid", clientid, "target", target, "mantant", mantant, "bank",bank, "mode", mode, "npiece", npiece, "echeance", echeance,)
    for m, mod, np, ech, bk, nt in zip(mantant, mode, npiece, echeance, bank, note):
        av=Avanceclient.objects.create(
            client_id=clientid,
            amount=m,
            #today
            date=timezone.now().date(),
            echeance=ech,
            bank=bk,
            mode=mod,
            note=nt,
            npiece=np,
            isfarah=target=='f',
            isorgh=target=='o',
            issortie=target=='s',
            bonofavance=bonofavance
        )
        caisse=Caisse.objects.filter(target=target).first()
        if mod=='espece':
            if caisse:
                caisse.total+=m
                av.targetcaisse=caisse
                av.save()
                caisse.save()
    return JsonResponse({
        'success':True
    })
    
def addavancesupplier(request):
    supplierid=request.GET.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    target=request.GET.get('target')
    mantant=json.loads(request.GET.get('mantant'))
    supplier.rest-=sum(mantant)
    supplier.save()
    bank=json.loads(request.GET.get('bank'))
    mode=json.loads(request.GET.get('mode'))
    npiece=json.loads(request.GET.get('npiece'))
    # date=datetime.strptime(request.GET.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.GET.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    for m, mod, np, ech, bk in zip(mantant, mode, npiece, echeance, bank):
        av=Avancesupplier.objects.create(
            supplier_id=supplierid,
            amount=m,
            #today
            date=timezone.now().date(),
            echeance=ech,
            bank=bk,
            mode=mod,
            npiece=np,
            isfarah=target=='f',
            isorgh=target=='o'
        )
        caisse=Caisse.objects.filter(target=target).first()
        if mod=='espece':
            if caisse:
                caisse.total-=m
                av.targetcaisse=caisse
                av.save()
                caisse.save()
    return JsonResponse({
        'success':True
    })
    

def supplierlistavances(request):
    target=request.GET.get('target')
    if target=="f":
        avances=Avancesupplier.objects.filter(isfarah=True).order_by('-id')[:50]
    elif target=="o":
        avances=Avancesupplier.objects.filter(isorgh=True).order_by('-id')[:50]
    else:
        avances=Avancesupplier.objects.filter(issortie=True).order_by('-id')[:50]
    print('avaavva', len(avances))
    ctx={
        'title':'List des avances CL BL',
        'avances':avances,
        'today':timezone.now().date(),
        'target':target,
    }
    # if avances:
    #     ctx['total']=round(Avancesupplier.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)

    return render(request, 'supplierlistavances.html', ctx)

def listreglementsupp(request):
    target=request.GET.get('target')
    isfarah=target=='f'
    reglements=PaymentSupplier.objects.filter(isfarah=isfarah).order_by('-id')
    ctx={
        'title':'List des reglements Fournisseur',
        'reglements':reglements,
        'target':target,
        'suppliers':Supplier.objects.all(),
        'today':timezone.now().date(),
        'banks':Bank.objects.all()
    }
    # if reglements:
    #     ctx['total']=round(PaymentSupplier.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)
    return render(request, 'listreglementsupp.html', ctx)


def listreglementfc(request):
    reglements=PaymentClientfc.objects.filter(date__year=thisyear).order_by('-date')[:100]
    print(round(PaymentClientfc.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'] or 0, 2))
    ctx={
        'title':'List des reglements CL fc',
        'reglements':reglements,
        'today':timezone.now().date(),


    }
    if reglements:
        ctx['total']=round(PaymentClientfc.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'] or 0, 2)


    return render(request, 'listreglementfc.html', ctx)


def getclientbons(request):
    clientid=request.GET.get('clientid')
    target=request.GET.get('target')
    moderegl=request.GET.get('moderegl')
    datefrom=request.GET.get('datefrom')
    dateend=request.GET.get('dateend')
    print('>> target', target)
    trs=''
    isfarah=target=='f'
    if target=='s':
        avoir=Avoirclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], issortie=True, inreglement=False, ispaid=False).order_by('date')
        avance=Avanceclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], issortie=True, inreglement=False).order_by('date')
        bons=Bonsortie.objects.filter(client_id=clientid, ispaid=False, date__range=[datefrom, dateend]).order_by('date')
        print('>> in reglezmznt', bons)
        total=round(Bonsortie.objects.filter(client_id=clientid).aggregate(Sum('total')).get('total__sum')or 0,  2)
        # for i in bons:
        #     trs+=f'<tr class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)"></td></tr>'
        return JsonResponse({
            'bons':render(request, 'clintbonsinreglement.html', {'bons':bons}).content.decode('utf-8'),
            'avoirs':render(request, 'avoirsbl.html', {'avoirs':avoir}).content.decode('utf-8'),
            'avances':render(request, 'avancesbl.html', {'avances':avance}).content.decode('utf-8'),
            # 'total':total,
            'totalavoirs':round(avoir.aggregate(Sum('total')).get('total__sum') or 0, 2),
            'totalavances':round(avance.aggregate(Sum('amount')).get('amount__sum') or 0, 2),
            # 'total':total,
            #'soldbl':round(Client.objects.get(pk=clientid).soldbl, 2),
            'totalbons':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2),
            
        })
    # elif target=="f":
    print('>>> mode regl', moderegl)
    if moderegl=='bl':
        bons=Bonlivraison.objects.filter(ispaid=False, client_id=clientid, date__range=[datefrom, dateend], isfarah=isfarah, iscanceled=False, isfacture=False).order_by('date')
        #[:50]
        for i in bons:
            # old code, if reglement is paid it's checked from here
            # trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
            trs+=f'<tr class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)"></td></tr>'
    else:
        bons=Facture.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=isfarah, ispaid=False).order_by('date')
        for i in bons:
            # old code, if reglement is paid it's checked from here
            # trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
            trs+=f'<tr class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no} </td><td>{i.total}</td><td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)"></td></tr>'
    #total=round(Bonlivraison.objects.filter(ispaid=False, client_id=clientid).aggregate(Sum('total')).get('total__sum')or 0,  2)
    avoir=Avoirclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=isfarah, inreglement=False, ispaid=False).order_by('date')
    avance=Avanceclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=isfarah, inreglement=False).order_by('date')

    # else:
    #     if moderegl=='bl':
    #         bons=Bonlivraison.objects.filter(ispaid=False, client_id=clientid, date__range=[datefrom, dateend], isfarah=False).order_by('date')[:50]
    #         for i in bons:
    #             # old code, if reglement is paid it's checked from here
    #             # trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    #             trs+=f'<tr class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)"></td></tr>'
    #     else:
    #         bons=Facture.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=False).order_by('date')[:50]
    #         for i in bons:
    #             # old code, if reglement is paid it's checked from here
    #             # trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    #             trs+=f'<tr class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td><input type="checkbox" value="{i.id}" name="bonstopay" total={i.total} onchange="checkreglementbox(event)"></td></tr>'

    #     avoir=Avoirclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=False, inreglement=False).order_by('date')[:50]
    #     avance=Avanceclient.objects.filter(client_id=clientid, date__range=[datefrom, dateend], isfarah=False, inreglement=False).order_by('date')[:50]
    # totalbons=bons.aggregate
        #total=round(Bonlivraison.objects.filter(client_id=clientid).aggregate(Sum('total')).get('total__sum')or 0,  2)
    print("bons", bons)
    print('start, ent', datefrom, dateend)
    return JsonResponse({
        'bons':trs,
        'avoirs':render(request, 'avoirsbl.html', {'avoirs':avoir}).content.decode('utf-8'),
        'avances':render(request, 'avancesbl.html', {'avances':avance}).content.decode('utf-8'),
        # 'total':total,
        'soldbl':round(Client.objects.get(pk=clientid).soldbl, 2),
        'totalbons':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2),
        'totalavoirs':round(avoir.aggregate(Sum('total')).get('total__sum') or 0, 2),
        'totalavances':round(avance.aggregate(Sum('amount')).get('amount__sum') or 0, 2)
    })

def filternonreglr(request):
    clientid=request.GET.get('clientid')
    bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False).order_by('date')[:50]
    total=round(Bonlivraison.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum') or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldbl':Client.objects.get(pk=clientid).soldbl
    })

def filternonreglrfc(request):
    clientid=request.GET.get('clientid')
    print(clientid)
    bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('date')[:50]
    total=round(Facture.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum')or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldfc':Client.objects.get(pk=clientid).soldfacture
    })

def getclientfactures(request):
    clientid=request.POST.get('clientid')
    bons=Facture.objects.filter(client_id=clientid).order_by('date')[:50]
    trs=''
    for i in bons:
        trs+=f'<tr  class="fcreglrow" style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td>{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs,
        'totalfactures':round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum']or 0, 2),
        'soldfactureregl':round(Client.objects.get(pk=clientid).soldfacture, 2)
    })

def reglefactures(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Facture.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')
    # livraisonstotals=0
    totalmantant=sum(mantant)
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # print(totalmantant, livraisonstotals, livraisons)
    # # update client sold
    # # case1: 5000==5000:
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)

    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
        # print('case3')
        # amount=totalmantant
        # for i in livraisons:
        #     if amount<=0:
        #         break
        #     else:
        #         # if facture has rest
        #         if i.rest>0:
        #             if amount>=i.rest:
        #                 i.rest=0
        #                 i.ispaid=True
        #                 print(f'facture {i.facture_no} has rest is paid here')
        #                 i.save()
        #                 amount-=i.rest
        #             else:
        #                 print(f'facture {i.facture_no} has rest has rest here')
        #                 i.rest=round(i.rest-amount, 2)
        #                 i.save()
        #                 break
        #         else:
        #             if amount>=i.total:
        #                 amount-=i.total
        #                 print(f'facture {i.facture_no} is paid here')
        #                 i.ispaid=True
        #                 i.save()
        #             else:
        #                 print(f'facture {i.facture_no} hasrestof {round(i.total-amount, 2)}')
        #                 i.rest=round(i.total-amount, 2)
        #                 i.save()
        #                 break
        # print('rest of amount', amount)
    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        tva=round(m-(m/1.2), 2)
        regl=PaymentClientfc.objects.create(
            client_id=clientid,
            amount=m,
            date=date,
            tva=tva,
            echance=ech,
            mode=mod,
            npiece=np,

        )
        regl.factures.set(livraisons)
        # storing factures in facturesregle
        # for i in livraisons:
        #     Facturesregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )


    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldfacture=round(float(client.soldfacture)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })


def reglebons(request):
    clientid=request.POST.get('clientid')
    date=request.POST.get('date')
    whattopay=float(request.POST.get('whattopay'))
    print('>> whatt', whattopay)
    moderegl=request.POST.get('moderegl')
    target=request.POST.get('target')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    avoirs=json.loads(request.POST.get('avoirs'))
    avances=json.loads(request.POST.get('avances'))
    mantant=json.loads(request.POST.get('mantant'))
    bank=json.loads(request.POST.get('bank'))
    source=json.loads(request.POST.get('source'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    print('bons', bons, 'avoirs', avoirs, 'avances', avances)
    # date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    if moderegl=='bl':
        livraisons=Bonlivraison.objects.filter(pk__in=bons)
    else:
        livraisons=Facture.objects.filter(pk__in=bons)
    avoirs=Avoirclient.objects.filter(pk__in=avoirs)
    avances=Avanceclient.objects.filter(pk__in=avances)
    totalmantant=sum(mantant)
    totalbons=livraisons.aggregate((Sum('total')))['total__sum'] or 0
    if totalmantant>whattopay:
        livraisons.update(ispaid=True)
        livraisons.update(statusreg='r1')
        if moderegl=='facture':
            print("make bon paid")
            for livraison in livraisons:
                print("make bon paid", livraison.bons.all())
                # Update 'ispaid' for related ManyToManyField (bons)
                livraison.bons.all().update(ispaid=True)
        diff=totalmantant-whattopay
        Avanceclient.objects.create(
            client_id=clientid,
            amount=diff,
            # date=timezone.now().date(),
            date=date,
            isfarah=target=='f',
            isorgh=target=='o'
        )
    if totalmantant==whattopay:
        if moderegl=='facture':
            print("make bon paid")
            for livraison in livraisons:
                print("make bon paid", livraison.bons.all())
                # Update 'ispaid' for related ManyToManyField (bons)
                livraison.bons.all().update(ispaid=True)
        # else:

        livraisons.update(ispaid=True)
    if totalmantant<whattopay:
        Avanceclient.objects.create(
            client_id=clientid,
            amount=totalmantant,
            # date=timezone.now().date(),
            date=date,
            isfarah=target=='f',
            isorgh=target=='o'
        )

    for m, mod, np, ech, bk, s in zip(mantant, mode, npiece, echeance, bank, source):
        regl=PaymentClientbl.objects.create(
            client_id=clientid,
            amount=m,
            #today
            # date=timezone.now().date(),
            date=date,
            echance=ech,
            bank=bk,
            mode=mod,
            npiece=np,
            source=s,
            isfarah=target=='f',
            isorgh=target=='o',
            issortie=target=='s'
        )
        if moderegl=='bl':
            regl.bons.set(livraisons)
        else:
            regl.factures.set(livraisons)
            regl.usedinfacture=True
            regl.save()
        regl.avoirs.set(avoirs)
        regl.avances.set(avances)
        avoirs.update(inreglement=True)
        avances.update(inreglement=True)
        livraisons.update(ispaid=True)
        caisse=Caisse.objects.filter(target=target).first()
        if mod=='espece':
            if caisse:
                caisse.total+=m
                regl.targetcaisse=caisse
                regl.save()
                caisse.save()

        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldbl=round(float(client.soldbl)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })


def reglebonsortie(request):
    clientid=request.POST.get('clientid')
    whattopay=float(request.POST.get('whattopay'))
    print('>> whatt', whattopay)
    moderegl=request.POST.get('moderegl')
    target=request.POST.get('target')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    avoirs=json.loads(request.POST.get('avoirs'))
    avances=json.loads(request.POST.get('avances'))
    mantant=json.loads(request.POST.get('mantant'))
    bank=json.loads(request.POST.get('bank'))
    source=json.loads(request.POST.get('source'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    print('bons', bons, 'avoirs', avoirs, 'avances', avances)
    # date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Bonsortie.objects.filter(pk__in=bons)
    
    avoirs=Avoirclient.objects.filter(pk__in=avoirs)
    avances=Avanceclient.objects.filter(pk__in=avances)
    avoirs.update(inreglement=True)
    avances.update(inreglement=True)
    totalmantant=sum(mantant)
    totalbons=livraisons.aggregate((Sum('total')))['total__sum'] or 0
    # if totalmantant>whattopay:
    #     livraisons.update(ispaid=True)
    #     livraisons.update(statusreg='r1')
    #     if moderegl=='facture':
    #         print("make bon paid")
    #         for livraison in livraisons:
    #             print("make bon paid", livraison.bons.all())
    #             # Update 'ispaid' for related ManyToManyField (bons)
    #             livraison.bons.all().update(ispaid=True)
    #     diff=totalmantant-whattopay
    #     Avanceclient.objects.create(
    #         client_id=clientid,
    #         amount=diff,
    #         date=timezone.now().date(),
    #         isfarah=target=='f',
    #         isorgh=target=='o'
    #     )
    # if totalmantant==whattopay:
        

    livraisons.update(ispaid=True)
    # if totalmantant<whattopay:
    #     Avanceclient.objects.create(
    #         client_id=clientid,
    #         amount=totalmantant,
    #         date=timezone.now().date(),
    #         isfarah=target=='f',
    #         isorgh=target=='o'
    #     )

    for m, mod, np, ech, bk, s in zip(mantant, mode, npiece, echeance, bank, source):
        regl=PaymentClientbl.objects.create(
            client_id=clientid,
            amount=m,
            #today
            date=timezone.now().date(),
            echance=ech,
            bank=bk,
            mode=mod,
            npiece=np,
            issortie=True,
            source=s
        )
        regl.bonsortie.set(livraisons)
        
        regl.avoirs.set(avoirs)
        regl.avances.set(avances)
        caisse=Caisse.objects.filter(target=target).first()
        if mod=='espece':
            if caisse:
                caisse.total+=m
                regl.targetcaisse=caisse
                regl.save()
                caisse.save()
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldbl=round(float(client.soldbl)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })



def checknpiece(request):
    npiece=request.GET.get('npiece')
    print('>>> eee',npiece, PaymentClientbl.objects.filter(npiece=npiece))
    if PaymentClientbl.objects.filter(npiece=npiece).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })


def viewreglement(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)
    reglements=Bonsregle.objects.filter(payment=reglement)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglements':reglements
    }
    return render(request, 'viewreglement.html', ctx)

def viewreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglfc':True
    }
    return render(request, 'viewreglement.html', ctx)



def situationcl(request):
    target=request.GET.get('target')
    ctx={
        'title':'Situation des clients',
        'today':timezone.now().date(),
        'target':target
    }
    return render(request, 'situationcl.html', ctx)

def situationclblfc(request):
    # when clicking on download in situationclbl/fc get the client id from function arguments
    clientname=request.GET.get('clientname')
    monthtostart=request.GET.get('monthtostart')
    monthtoend=request.GET.get('monthtoend')
    print('>>', monthtostart, monthtoend)
    clientid=request.GET.get('clientid')
    ctx={
        'monthtostart':monthtostart,
        'monthtoend':monthtoend,
        'title':'Situation des clients',
        'clientname':clientname,
        'clientid':clientid,
        'today':timezone.now().date()
    }
    return render(request, 'situationclblfc.html', ctx)

def situationsupplier(request):
    ctx={
        'title':'Situation des Fouurnisseurs',
        'suppliers':Supplier.objects.all(),
        'target':request.GET.get('target'),
    }
    return render(request, 'situationsupplier.html', ctx)

def situationclfc(request):
    ctx={
        'title':'Situation des clients (Factures)',
        'today':timezone.now().date()
    }
    return render(request, 'situationclfc.html', ctx)

def recevoirexcel(request):
    farah=request.GET.get('farah')=='1'
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna(0)
    print('>>>>>>>>>> exceel')
    trs=''
    totalbon=0
    refnotexist=[]
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #name=d.name
        qty=d.qty
        prixachat=float(d.prixachat)
        # prixachat 12,0 make it 12,00
        remise=0 if pd.isna(d.remise) else int(d.remise)
        devise=0 if pd.isna(d.devise) else d.devise
        prixnet=round(prixachat-(prixachat*float(remise)/100), 2)

        # prixnet=round(prixachat-(prixachat*float(remise)/100), 2)
        # if remise==0:
        #     total=round(prixachat*qty, 2)
        # else:
        #     total=round(prixnet*qty, 2)
        # totalbon+=total
        # make total 2 digits after point
        if remise==0:
            total=round(prixachat*qty, 2)
        else:
            total=round(prixnet*qty, 2)
        product=Produit.objects.filter(ref=ref).first()
        if product:
            trs+=f"""<tr>
            <td class="ref">{ref}</td>
            <td class="name">{product.name}</td>
            <td>{product.buyprice if product.buyprice else 0}</td>
            <td>{product.stocktotal if product.stocktotal else 0}</td>
            <td>
            <input style="width:65px;" type="number" class="devise" value="{devise:.2f}">
            </td>
            <td><input style="width:65px;" type="number" class="calculatebonachat qty" onkeyup="calculate(event)" name="qtybonachat" value="{qty}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat price" onkeyup="calculate(event)" name="pricebonachat" value="{prixachat:.2f}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat remise" onkeyup="calculate(event)" name="remise" value="{remise}"></td>
            <td class="total">{total:.2f}</td>
            <input type="hidden" name="productid" value="{product.id}">
            </tr>"""

            totalbon+=total
        else:
            print('W>>>>not exist', ref)
            refnotexist.append(ref)
        # print('product exists')
        # trs+=f"""<tr>
        #     <td class="ref">{ref}</td>
        #     <td class="name">{product.name}</td>
        #     <td>{product.buyprice if product.buyprice else 0}</td>
        #     <td>{product.stocktotal if product.stocktotal else 0}</td>
        #     <td>
        #     <input style="width:65px;" type="number" class="devise" value="{devise:.2f}">
        #     </td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat qty" onkeyup="calculate(event)" name="qtybonachat" value="{qty}"></td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat price" onkeyup="calculate(event)" name="pricebonachat" value="{prixachat:.2f}"></td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat remise" onkeyup="calculate(event)" name="remise" value="{remise}"></td>
        #     <td class="total">{total:.2f}</td>
        #     <input type="hidden" name="productid" value="{product.id}">
        # </tr>"""
    return JsonResponse({
        'trs':trs,
        'totalbon':totalbon,
        'refnotexist':refnotexist
    })


def avoirclient(request):
    target=request.GET.get('target')
    ctx={
            'title':'Avoir client',
            'target':target,
            # 'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }
    return render(request, 'avoirclient.html', ctx)

def avoirclientinpointvente(request):
    target=request.GET.get('target')
    ctx={
            'title':'Avoir client point de vente',
            'target':target,
            # 'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }
    return render(request, 'avoirclientinpointvente.html', ctx)



def addavoirclient(request):
    mantant=json.loads(request.POST.get('mantant'))
    bank=json.loads(request.POST.get('bank'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    # date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    target=request.POST.get('target')
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # mode=request.POST.get('mode')
    # isfacture=True if mode=='facture' else False
    # orderno
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    isfarah=False
    isorgh=False
    issortie=False
    year = timezone.now().strftime("%y")
    if target=="f":
        isfarah=True

        prefix = f'FR-AV{year}'
        try:
            avoirclients = Avoirclient.objects.filter(no__startswith=prefix).last()
            latest_receipt_no = int(avoirclients.no.split('/')[1])
            receipt_no = f"FR-AV{year}/{latest_receipt_no + 1}"
        except:
            receipt_no = f"FR-AV{year}/1"
    elif target=='o':
        isorgh=True
        prefix = f'AV{year}'
        try:
            avoirclients = Avoirclient.objects.filter(no__startswith=prefix).last()
            latest_receipt_no = int(avoirclients.no.split('/')[1])
            receipt_no = f"AV{year}/{latest_receipt_no + 1}"
        except:
            receipt_no = f"AV{year}/1"
    else:
        issortie=True
        prefix = f'SO-AV{year}'
        try:
            avoirclients = Avoirclient.objects.filter(no__startswith=prefix).last()
            latest_receipt_no = int(avoirclients.no.split('/')[1])
            receipt_no = f"SO-AV{year}/{latest_receipt_no + 1}"
        except:
            receipt_no = f"SO-AV{year}/1"
    print(receipt_no, clientid, repid, totalbon, datebon)

    avoir=Avoirclient.objects.create(
        no=receipt_no,
        client_id=clientid,
        representant_id=repid,
        total=totalbon,
        date=datebon,
        # avoirfacture=isfacture,
        isfarah=isfarah,
        isorgh=isorgh,
        issortie=issortie
    )

    for i in json.loads(products):
        product=Produit.objects.get(pk=i['productid'])
        # if isfacture:
        #     product.stockfacture=int(product.stockfacture)+int(i['qty'])
        
        if isfarah:
            ins = Stockin.objects.filter(
                qtyofprice__gt=0,
                product=product,
                isfarah=True
            ).exclude(
                qtyofprice=F('quantity')
            )
            product.stocktotalfarah=float(product.stocktotalfarah)+float(i['qty'])
            #prices=Stockin.objects.filter(product=product, isfarah=True)
            
        if isorgh:
            product.stocktotalorgh=float(product.stocktotalorgh)+float(i['qty'])
            
        product.save()
        Stockin.objects.create(
            isavoir=True,
            avoir=avoir,
            product=product,
            quantity=i['qty'],
            remise1=0 if i['remise']=="" else i['remise'],
            price=0 if i['price']=="" else i['price'],
            total=i['total'],
            isfarah=isfarah,
            isorgh=isorgh,
            issortie=issortie,
            date=datebon,
            qtyofprice=i['qty']
        )
    totalamount=sum([i for i in mantant if i is not None])
    if totalamount>0:
        print('totalamount', totalamount)
        if float(totalamount)==float(totalbon):
            avoir.ispaid=True
        else:
            diff=round(float(totalbon)-float(totalamount), 2)
            avoir.rest=diff
        avoir.save()
        for m, mod, np, ech, bk in zip(mantant, mode, npiece, echeance, bank):
            if m is not None:
                regl=PaymentClientbl.objects.create(
                    client_id=clientid,
                    amount=m,
                    #today
                    date=timezone.now().date(),
                    echance=ech,
                    bank=bk,
                    mode=mod,
                    npiece=np,
                    isfarah=target=='f',
                    isorgh=target=='o',
                    issortie=target=='s',
                    isavoir=True
                )
                regl.avoirs.set([avoir])
                caisse=Caisse.objects.filter(target=target).first()
                if mod=='espece':
                    if caisse:
                        caisse.total-=m
                        regl.targetcaisse=caisse
                        regl.save()
                        caisse.save()
    

    # increment it
    return JsonResponse({
        'success':True
        # 'html':render(request, 'avoirclient.html', {
        #     'title':'Bon de livraison',
        #     #'clients':Client.objects.all(),
        #     #'products':Produit.objects.all(),
        #     'commercials':Represent.objects.all(),
        #     #'receipt_no':receipt_no
        # }).content.decode('utf-8')

    })

def avoirsupplier(request):
    target=request.GET.get('target')
    year = timezone.now().strftime("%y")
    isfarah=target=='f'
    if isfarah:
        prefix = f'FR-FAV{year}'
    else:
        prefix = f'FAV{year}'
    try:
        avoirsuppliers = Avoirsupplier.objects.filter(no__startswith=prefix).last()
        print('>>>>', avoirsuppliers.no)
        latest_receipt_no = int(avoirsuppliers.no.split('/')[1])
        receipt_no = f"{prefix}{year}/{latest_receipt_no + 1}"

    except:
        receipt_no = f"{prefix}{year}/1"
    print('>>>>>>rec av supp', receipt_no)
    ctx={
            'title':'Avoir Fournisseur',
            'target':target
        }
    return render(request, 'avoirsupplier.html', ctx)



def addavoirsupp(request):
    mantant=json.loads(request.POST.get('mantant'))
    bank=json.loads(request.POST.get('bank'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    
    isfarah=request.POST.get('target')=="f"
    supplierid=request.POST.get('supplierid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    # avoirfacture=True if request.POST.get('mode')=='facture' else False
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    supplier=Supplier.objects.get(pk=supplierid)
    year = timezone.now().strftime("%y")
    if isfarah:
        prefix = f'FR-FAV{year}'
    else:
        prefix = f'FAV{year}'
    try:
        lastid = Avoirsupplier.objects.last().id
        
    except:
        lastid = 0
    receipt_no = f"{prefix}/{lastid+1}"
    print(receipt_no)
    avoir=Avoirsupplier.objects.create(
        no=receipt_no,
        supplier_id=supplierid,
        total=totalbon,
        date=datebon,
        # avoirfacture=avoirfacture,
        isfarah=isfarah
    )
    # supplier.rest-=float(totalbon)
    # supplier.save()
    with transaction.atomic():
        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            if isfarah:
                product.stocktotalfarah=float(product.stocktotalfarah)-float(i['qty'])
            else:
                product.stocktotalorgh=float(product.stocktotalorgh)-float(i['qty'])
            product.save()
            remise1=0 if i['remise1']=="" else i['remise1']
            remise2=0 if i['remise2']=="" else i['remise2']
            remise3=0 if i['remise3']=="" else i['remise3']
            remise4=0 if i['remise4']=="" else i['remise4']
            print('>>', remise1, remise2, remise3, remise4)
            Returnedsupplier.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
                isfarah=isfarah,
                remise1=0 if i['remise1']=="" else i['remise1'],
                remise2=0 if i['remise2']=="" else i['remise2'],
                remise3=0 if i['remise3']=="" else i['remise3'],
                remise4=0 if i['remise4']=="" else i['remise4']
            )
    totalamount=sum([i for i in mantant if i is not None])
    if totalamount>0:
        print('totalamount', totalamount)
        avoir.ispaid=True
        avoir.save()
        for m, mod, np, ech, bk in zip(mantant, mode, npiece, echeance, bank):
            print(m)
            if m is not None:
                regl=PaymentSupplier.objects.create(
                    supplier_id=supplierid,
                    amount=m,
                    #today
                    date=timezone.now().date(),
                    echeance=ech,
                    bank=bk,
                    mode=mod,
                    npiece=np,
                    isfarah=isfarah,
                    isavoir=True
                )
                regl.avoirs.set([avoir])

    # increment it
    return JsonResponse({
        'success':True
    })


def checkbl(request, id):
    client=Client.objects.get(pk=id)
    print(Bonlivraison.objects.filter(client=client).first())
    if Bonlivraison.objects.filter(client=client).first() is not None:
        return JsonResponse({
            'hasbl':True
        })
    else:
        return JsonResponse({
            'hasbl':False
        })

def checkfacture(request, id):
    client=Client.objects.get(pk=id)
    print(client.name)
    print(Facture.objects.filter(client=client).first())
    if Facture.objects.filter(client=client).first() is not None:
        # get products
        return JsonResponse({
            'hasfacture':True
        })
    else:
        return JsonResponse({
            'hasfacture':False
        })

def sendrelevclient(request):
    clientcode=request.GET.get('clientcode')
    client=Client.objects.get(code=clientcode)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client=client, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client=client, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client=client, date__range=[startdate, enddate], total__gt=0)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevecl.html', {
            'releve':[sorted_releve[i:i+34] for i in range(0, len(sorted_releve), 34)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8'),
        'soldfc':client.soldfacture,
    })
def relevclient(request):
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> target', target)
    if target=="f":
        print('>>> here in farah', clientid, )
        bons=Bonlivraison.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=True)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isfarah=True, ispaid=False)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True, isavoir=False, amount__gt=0)
    elif target=="s":
        bons=Bonsortie.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], issortie=True, ispaid=False)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True, isavoir=False, amount__gt=0)
    else:
        bons=Bonlivraison.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=False)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isorgh=True, ispaid=False)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True, isavoir=False, amount__gt=0)
        #bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0, isfarah=isfarah)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    # print('>> avances', avances[0].amount)
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((avance, 'avanceclient') for avance in avances),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevecl.html', {
            'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8'),
        #'soldfc':client.soldfacture,
    })


def relevclientnr(request):
    clientid=request.GET.get('clientid')
    target=request.GET.get('target')
    client=Client.objects.get(pk=clientid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> target', target)
    if target=="f":
        print('>>> here in farah', clientid, )
        bons=Bonlivraison.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=True, ispaid=False)
        
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isfarah=True, ispaid=False, inreglement=False)
        avances=[]
        reglementsbl=[]
    elif target=="s":
        bons=Bonsortie.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0, ispaid=False)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], issortie=True, ispaid=False, inreglement=False)
        avances=[]
        reglementsbl=[]
        #reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True, isavoir=False, inreglement=False)
    else:
        bons=Bonlivraison.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=False, ispaid=False)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isorgh=True, ispaid=False, inreglement=False)
        avances=[]
        reglementsbl=[]
        #bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0, isfarah=isfarah)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    # print('>> avances', avances[0].amount)
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((avance, 'avanceclient') for avance in avances),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'releveclprint.html', {
            'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
            'client':client,
            'title':"Situation client non reglé",
            'startdate':startdate,
            'enddate':enddate,

        })
    # return JsonResponse({
    #     'html':render(request, 'releveclprint.html', {
    #         'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
    #         'client':client,

    #         'startdate':startdate,
    #         'enddate':enddate,

    #     }).content.decode('utf-8'),
    #     #'soldfc':client.soldfacture,
    # })




def relevsupplier(request):
    supplierid=request.POST.get('supplierid')
    target=request.POST.get('target')
    isrelevefacture=request.POST.get('facture')=='1'
    print('>> isfacture releve', request.POST.get('facture'))
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> is facture', isrelevefacture)
    if target=="f":
        avoirs=Avoirsupplier.objects.filter(isfarah=True, supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
        reglementsbl=PaymentSupplier.objects.filter(isfarah=True, supplier_id=supplierid, date__range=[startdate, enddate])
        avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=True, date__range=[startdate, enddate])
        if isrelevefacture:
            bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=True)
            print('>>>>>>>>>>>< infacture relve farah', bons)
        else:
            bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=True)
        print('rr', supplierid)
    else:
        # orgh
        avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
        reglementsbl=PaymentSupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
        avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=False, date__range=[startdate, enddate])
        if isrelevefacture:
            bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
        else:
            bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
        print('rr', supplierid)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'factureachat') for bon in bons) if isrelevefacture else ((bon, 'Itemsbysupplier') for bon in bons),
    ((avoir, 'Avoirsupplier') for avoir in avoirs),
    ((avance, 'avancesupplier') for avance in avances),
    ((reglementbl, 'Paymentsupplier') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevesupp.html', {
            'releve':sorted_releve,
            'supplier':supplier,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })

def relevsuppliernonregle(request):
    supplierid=request.POST.get('supplierid')
    target=request.POST.get('target')
    isfarah=target=='f'
    isrelevefacture=request.POST.get('facture')=='1'
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> is facture', isrelevefacture)
    avoirs=Avoirsupplier.objects.filter(isfarah=isfarah, supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentSupplier.objects.filter(isfarah=isfarah, supplier_id=supplierid, date__range=[startdate, enddate])
    avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=isfarah, date__range=[startdate, enddate])
    if isrelevefacture:
        bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=isfarah, ispaid=False)
        print('>>>>>>>>>>>< infacture relve farah', bons)
    else:
        bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=isfarah, ispaid=False)
    #     print('rr', supplierid)
    # else:
    #     # orgh
    #     avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
    #     reglementsbl=PaymentSupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
    #     avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=False, date__range=[startdate, enddate])
    #     if isrelevefacture:
    #         bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
    #     else:
    #         bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
    #     print('rr', supplierid)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'factureachat') for bon in bons) if isrelevefacture else ((bon, 'Itemsbysupplier') for bon in bons),
    ((avoir, 'Avoirsupplier') for avoir in avoirs),
    ((avance, 'avancesupplier') for avance in avances),
    ((reglementbl, 'Paymentsupplier') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevesupp.html', {
            'releve':sorted_releve,
            'supplier':supplier,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })


def sendrelevclientfc(request):
    clientcode=request.GET.get('clientcode')
    target=request.GET.get('target')
    client=Client.objects.get(code=clientcode)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    if target=='f':
        avoirs=Avoirclient.objects.filter(client=client, date__range=[startdate, enddate])
        reglementsfc=PaymentClientbl.objects.filter(client=client, date__range=[startdate, enddate])

        bons=Facture.objects.filter(client=client, date__range=[startdate, enddate], isfarah=True)
    else:
        avoirs=Avoirclient.objects.filter(client=client, date__range=[startdate, enddate], isorgh=True)
        reglementsfc=PaymentClientbl.objects.filter(client=client, date__range=[startdate, enddate], isorgh=True)

        bons=Facture.objects.filter(client=client, date__range=[startdate, enddate], isorgh=True)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facture') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)
    for i in sorted_releve:
        print(i)

    return JsonResponse({
        'html':render(request, 'releveclfc.html', {
            'releve':[sorted_releve[i:i+26] for i in range(0, len(sorted_releve), 26)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })
def relevclientfc(request):
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    if target=="f":
        print('>>> here in farah', clientid, )
        bons=Facture.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=True)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isfarah=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True)
    elif target=="s":
        bons=Bonsortie.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], issortie=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True)
    else:
        bons=Facture.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=False)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isorgh=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True)
        #bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0, isfarah=isfarah)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    print('>> bons, target', bons, target)
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((avance, 'avanceclient') for avance in avances),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevecl.html', {
            'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8'),
        #'soldfc':client.soldfacture,
    })

def getclientrep(request, id):
    client=Client.objects.get(pk=id)
    print(client.represent_id)
    return JsonResponse({
        'id':client.represent_id,
        'name':client.represent.name
    })

def listbonachat(request):
    target=request.GET.get('target')
    thisyear=timezone.now().year
    if target=='f':
        bons=Itemsbysupplier.objects.filter(date__year=thisyear, isfarah=True, isvalid=False).order_by('-id')[:50]
    elif target=='o':
        bons=Itemsbysupplier.objects.filter(date__year=thisyear, isfarah=False, isvalid=False).order_by('-id')[:50]
    ctx={
        'title':'List des bon achat',
        'bons':bons,
        'today':timezone.now().date(),
        'target':target
    }
    if bons:
        ctx['total']=round(Itemsbysupplier.objects.all().aggregate(Sum('total'))['total__sum'], 2)
        ctx['totaltva']=round(Itemsbysupplier.objects.all().aggregate(Sum('tva'))['tva__sum'], 2)
    return render(request, 'listbonachat.html', ctx)

def listboncommnd(request):
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    orders=Order.objects.filter(date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'List des bon commnd',
        'bons':orders,
        'today':timezone.now().date()
    }
    if orders:
        ctx['total']=round(Order.objects.all().aggregate(Sum('total'))['total__sum'], 2)
    return render(request, 'listboncommnd.html', ctx)

def bonachatdetails(request, id):
    target=request.GET.get('target')
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    payments=PaymentSupplier.objects.filter(bons__in=[bon])
    #orderitems=[items[i:i+36] for i in range(0, len(items), 36)]

    ctx={
        'title':f'Bon achat {bon.nbon}',
        'bon':bon,
        'items':items,
        'payments':payments,
        'orderitems':items,
        'target':target
    }
    return render(request, 'bonachatdetails.html', ctx)

def modifierbonachat(request, id):
    target=request.GET.get('target')
    print('>>', target)
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    ctx={
        'title':'Modifier bon achat',
        'bon':bon,
        'items':items,
        'suppliers':Supplier.objects.all(),
        'target':target
    }
    return render(request, 'modifierbonachat.html', ctx)


def modifierfactureachat(request):
    target=request.GET.get('target')
    id=request.GET.get('id')
    print('>>', target)
    bon=Factureachat.objects.get(pk=id)
    items=Outfactureachat.objects.filter(facture=bon)
    print(items)
    ctx={
        'title':'Modifier Facture achat '+bon.facture_no,
        'bon':bon,
        'items':items,
        'suppliers':Supplier.objects.all(),
        'target':target
    }
    return render(request, 'modifierfactureachat.html', ctx)



def updatebonachat(request):
    id=request.POST.get('bonid')
    target=request.POST.get('target')
    print('>>> target', target)
    isfarah=target=='f'
    bon=Itemsbysupplier.objects.get(pk=id)
    # we need to see if this bon is the last bon, to edit the prices and rmises
    
    bon.date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    bon.nbon=request.POST.get('orderno')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    supplier=Supplier.objects.get(pk=request.POST.get('supplierid'))
    thissupplier=bon.supplier
    facture=Factureachat.objects.filter(bons__in=[bon]).first()
    if facture:
        print('>>> has facture', facture.total, 'updating total', round(float(facture.total)-float(bon.total)+float(totalbon), 2))
        newtotal=round(float(facture.total)-float(bon.total)+float(totalbon), 2)
        facture.total=newtotal
        facture.save()
    # if bon.supplier==supplier:
    #     print('same supplier', float(thissupplier.rest), float(bon.total), float(totalbon))
    #     thissupplier.rest=round(float(thissupplier.rest)-float(bon.total)+float(totalbon), 2)
    #     thissupplier.save()
    # else:
    #     print('not same')
    #     thissupplier.rest=round(float(thissupplier.rest)-float(bon.total), 2)
    #     thissupplier.save()
    #     print('old', thissupplier.rest)
    #     supplier.rest=round(float(supplier.rest)+float(totalbon), 2)
    #     supplier.save()
    #     print('new', supplier.rest)
    # bon.supplier.rest=float(bon.supplier.rest)-float(bon.total)
    # bon.supplier.save()
    items=Stockin.objects.filter(nbon=bon)
    # update this items
    for i in items:
        product=i.product
        print('removing from total')
        if bon.isfarah:
            product.stocktotalfarah=float(product.stocktotalfarah)-float(i.quantity)
        else:
            product.stocktotalorgh=float(product.stocktotalorgh)-float(i.quantity)
        # if bon.isfacture:
        #     print('removing from facture')
        #     product.stockfacture=int(product.stockfacture)-int(i.quantity)
        product.save()
        i.delete()

    bon.supplier=supplier
    bon.total=totalbon
    bon.nbon=request.POST.get('orderno')
    # bon.isfacture=isfacture
    bon.save()

    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            remise1=0 if i['remise1']=="" else i['remise1']
            remise2=0 if i['remise2']=="" else i['remise2']
            remise3=0 if i['remise3']=="" else i['remise3']
            remise4=0 if i['remise4']=="" else i['remise4']
            buyprice=float(i['price'])
            print('>>> buyprice', buyprice, buyprice-(buyprice*(float(remise1)/100)))
            # netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
            netwithremise1=round(buyprice-(buyprice*(float(remise1)/100)), 2)
            netprice=round(float(i['total'])/float(i['qty']), 2)
            qty=0 if i['qty']=="" else float(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            #product.originsupp=
            print('>>>>>>>adding total')
            
            # if isfacture:
            #     print('>>>>>>>adding fc')
            #     product.stockfacture=int(product.stockfacture)+qty
            #product.save()
            # create new livraison items
            # if i.get('achatitemid', None):
            #     st=Stockin.objects.filter(pk=i['achatitemid']).first()
            # # if not itemid means it's a new line
            #     print('>>> already')
            #     st.quantity=i['qty']
            #     st.price=i['price']
            #     st.remise1=0 if i['remise1']=="" else i['remise1']
            #     st.remise2=0 if i['remise2']=="" else i['remise2']
            #     st.remise3=0 if i['remise3']=="" else i['remise3']
            #     st.remise4=0 if i['remise4']=="" else i['remise4']
            #     st.net=netwithremise1
            #     st.save()
            # else:
            #     print('>>> new line')
            Stockin.objects.create(
                nbon=bon,
                supplier=supplier,
                remise1=0 if i['remise1']=="" else i['remise1'],
                remise2=0 if i['remise2']=="" else i['remise2'],
                remise3=0 if i['remise3']=="" else i['remise3'],
                remise4=0 if i['remise4']=="" else i['remise4'],
                product=product,
                date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d'),
                quantity=qty,
                price=0 if i['price']=="" else i['price'],
                total=0 if i['total']=="" else i['total'],
                qtyofprice=qty,
                isfarah=bon.isfarah,
                isorgh=bon.isorgh,
                net=netwithremise1
            )
            
            # totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
            # totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
            # print(totalprices, totalqty)
            #product.coutmoyen=round(totalprices/totalqty, 2)
            if bon.isfarah:
                # if product.stocktotalfarah>0:
                #     print('>> has stock')
                #     totalqtys=int(product.stocktotalfarah)+int(i['qty'])
                #     lastbuyprice=Stockin.objects.filter(product=product, isfarah=True).last()
                #     if lastbuyprice:
                #         actualtotal=product.stocktotalfarah*lastbuyprice.price
                #         thistotal=int(i['qty'])*buyprice
                #         totalprices=round(thistotal+actualtotal, 2)
                #         pondire=round(totalprices/totalqtys, 2)
                #         product.frcoutmoyen=pondire
                #         #product.save()
                #     else:
                #         product.frcoutmoyen=buyprice
                # else:
                #     product.frcoutmoyen=buyprice
                product.stocktotalfarah=float(product.stocktotalfarah)+qty
                islastbon = bon == Itemsbysupplier.objects.filter(isfarah=True).last()
                if islastbon:
                    product.frbuyprice=0 if i['price']=="" else i['price']
                    product.frremise1=0 if i['remise1']=="" else i['remise1']
                    product.frremise2=0 if i['remise2']=="" else i['remise2']
                    product.frremise3=0 if i['remise3']=="" else i['remise3']
                    product.frremise4=0 if i['remise4']=="" else i['remise4']
                    #product.frsellprice=0 if i['price']=="" else i['price']
                    #product.frremisesell=0 if i['remise1']=="" else i['remise1']
                    product.frnetbuyprice=netprice
                    product.froriginsupp=supplier
            else:
                # claculate cout moyen
                # if product.stocktotalorgh>0:
                #     print('>> has stock')
                #     totalqtys=int(product.stocktotalorgh)+int(i['qty'])
                #     lastbuyprice=Stockin.objects.filter(product=product, isfarah=False).last()
                #     if lastbuyprice:
                #         actualtotal=product.stocktotalorgh*lastbuyprice.price
                #         thistotal=int(i['qty'])*buyprice
                #         totalprices=round(thistotal+actualtotal, 2)
                #         pondire=round(totalprices/totalqtys, 2)
                #         product.coutmoyen=pondire
                #         #product.save()
                #     else:
                #         product.coutmoyen=buyprice
                # else:
                #     product.coutmoyen=buyprice
                product.stocktotalorgh=float(product.stocktotalorgh)+qty
                islastbon = bon == Itemsbysupplier.objects.filter(isfarah=False).last()
                if islastbon:
                    product.buyprice=0 if i['price']=="" else i['price']
                    product.remise1=0 if i['remise1']=="" else i['remise1']
                    product.remise2=0 if i['remise2']=="" else i['remise2']
                    product.remise3=0 if i['remise3']=="" else i['remise3']
                    product.remise4=0 if i['remise4']=="" else i['remise4']
                    #product.sellprice=0 if i['price']=="" else i['price']
                    #product.remisesell=0 if i['remise1']=="" else i['remise1']
                    product.netbuyprice=netprice
                    product.originsupp=supplier

            product.save()

    return JsonResponse({
        'success':True
    })

def updatefactureachat(request):
    id=request.POST.get('bonid')
    target=request.POST.get('target')
    print('>>> target', target)
    isfarah=target=='f'
    bon=Factureachat.objects.get(pk=id)
    bon.date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    bon.factue_no=request.POST.get('orderno')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    supplier=Supplier.objects.get(pk=request.POST.get('supplierid'))
    thissupplier=bon.supplier
    if bon.supplier==supplier:
        print('same supplier', float(thissupplier.rest), float(bon.total), float(totalbon))
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total)+float(totalbon), 2)
        thissupplier.save()
    else:
        print('not same')
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total), 2)
        thissupplier.save()
        print('old', thissupplier.rest)
        supplier.rest=round(float(supplier.rest)+float(totalbon), 2)
        supplier.save()
        print('new', supplier.rest)
    # bon.supplier.rest=float(bon.supplier.rest)-float(bon.total)
    # bon.supplier.save()
    items=Outfactureachat.objects.filter(facture=bon)
    # update this items
    for i in items:
        product=i.product
        print('removing from total')
        if isfarah:
            product.stocktotalfarah=float(product.stocktotalfarah)-float(i.qty)
        else:
            product.stocktotalorgh=float(product.stocktotalorgh)-float(i.qty)
        # if bon.isfacture:
        #     print('removing from facture')
        #     product.stockfacture=int(product.stockfacture)-int(i.quantity)
        product.save()
        #i.delete()

    bon.supplier=supplier
    bon.total=totalbon
    bon.facture_no=request.POST.get('orderno')
    #bon.isfacture=isfacture
    bon.save()

    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            remise1=0 if i['remise1']=="" else i['remise1']
            remise2=0 if i['remise2']=="" else i['remise2']
            remise3=0 if i['remise3']=="" else i['remise3']
            remise4=0 if i['remise4']=="" else i['remise4']
            buyprice=float(i['price'])
            print('>>> buyprice', buyprice, buyprice-(buyprice*(float(remise1)/100)))
            # netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
            netwithremise1=round(buyprice-(buyprice*(float(remise1)/100)), 2)
            netprice=round(float(i['total'])/float(i['qty']), 2)
            qty=0 if i['qty']=="" else int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            print('>>>>>>>adding total')
            if isfarah:
                product.stocktotalfarah=int(product.stocktotalfarah)+qty
            else:
                product.stocktotalorgh=int(product.stocktotalorgh)+qty
            # if isfacture:
            #     print('>>>>>>>adding fc')
            #     product.stockfacture=int(product.stockfacture)+qty
            #product.save()
            # create new livraison items
            if i.get('achatitemid', None):
                st=Outfactureachat.objects.filter(pk=i['achatitemid']).first()
            # if not itemid means it's a new line
                print('>>> already')
                st.qty=i['qty']
                st.price=i['price']
                st.remise1=0 if i['remise1']=="" else i['remise1']
                st.remise2=0 if i['remise2']=="" else i['remise2']
                st.remise3=0 if i['remise3']=="" else i['remise3']
                st.remise4=0 if i['remise4']=="" else i['remise4']
                st.save()
            else:
                print('>>> new line')
                Outfactureachat.objects.create(
                    facture=bon,
                    supplier=supplier,
                    remise1=0 if i['remise1']=="" else i['remise1'],
                    remise2=0 if i['remise2']=="" else i['remise2'],
                    remise3=0 if i['remise3']=="" else i['remise3'],
                    remise4=0 if i['remise4']=="" else i['remise4'],
                    product=product,
                    date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d'),
                    qty=qty,
                    price=0 if i['price']=="" else i['price'],
                    total=0 if i['total']=="" else i['total'],
                    isfarah=isfarah,
                )
                
            totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
            totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
            print(totalprices, totalqty)
            #product.coutmoyen=round(totalprices/totalqty, 2)
            if isfarah:
                product.frbuyprice=0 if i['price']=="" else i['price']
                product.frremise1=0 if i['remise1']=="" else i['remise1']
                product.frremise2=0 if i['remise2']=="" else i['remise2']
                product.frremise3=0 if i['remise3']=="" else i['remise3']
                product.frremise4=0 if i['remise4']=="" else i['remise4']
                #product.frsellprice=0 if i['price']=="" else i['price']
                #product.frremisesell=0 if i['remise1']=="" else i['remise1']
                product.frnetbuyprice=netprice
            else:
                product.buyprice=0 if i['price']=="" else i['price']
                product.remise1=0 if i['remise1']=="" else i['remise1']
                product.remise2=0 if i['remise2']=="" else i['remise2']
                product.remise3=0 if i['remise3']=="" else i['remise3']
                product.remise4=0 if i['remise4']=="" else i['remise4']
                #product.sellprice=0 if i['price']=="" else i['price']
                #product.remisesell=0 if i['remise1']=="" else i['remise1']
                product.netbuyprice=netprice
            product.save()

    return JsonResponse({
        'success':True
    })



def getsuppbons(request):
    supplierid=request.POST.get('supplierid')
    mode=request.POST.get('mode')
    datestart=request.POST.get('datestart')
    dateend=request.POST.get('dateend')
    isfarah=request.POST.get('target')=='f'
    print('terget', request.POST.get('target'))
    trs=''
    if mode=="bl":
        bons=Itemsbysupplier.objects.filter(isfarah=isfarah, supplier_id=supplierid, ispaid=False, date__range=[datestart, dateend], isfacture=False)
        for i in bons:
            trs+=f'<tr><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.nbon}</td><td>{i.supplier.name}</td><td>{round(i.total, 2)}</td><td><input type="checkbox" value="{i.id}" name="bonstopay" total="{round(i.total, 2)}" onchange="checkreglementbox(event)"></td></tr>'

    else:
        bons=Factureachat.objects.filter(isfarah=isfarah, supplier_id=supplierid, ispaid=False, date__range=[datestart, dateend])
        for i in bons:
            trs+=f'<tr><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{round(i.total, 2)}</td><td><input type="checkbox" value="{i.id}" name="bonstopay" total={round(i.total, 2)} onchange="checkreglementbox(event)"></td></tr>'
    print('isfarah', isfarah)
    avoirs=Avoirsupplier.objects.filter(isfarah=isfarah, supplier_id=supplierid, inreglement=False, ispaid=False, date__range=[datestart, dateend])
    print('avoirs', avoirs)
    avances=Avancesupplier.objects.filter(isfarah=isfarah, supplier_id=supplierid, inreglement=False, date__range=[datestart, dateend])

    return JsonResponse({
        'bons':trs,
        'avoirs':render(request, 'avoirsbl.html', {'avoirs':avoirs}).content.decode('utf-8'),
        'avances':render(request, 'avancesbl.html', {'avances':avances}).content.decode('utf-8'),
        # 'total':total,
        'soldbl':round(Supplier.objects.get(pk=supplierid).rest, 2),
        'totalbons':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2),
        'totalavoirs':round(avoirs.aggregate(Sum('total')).get('total__sum') or 0, 2),
        'totalavances':round(avances.aggregate(Sum('amount')).get('amount__sum') or 0, 2)
    })

def reglebonsachat(request):
    target=request.POST.get('target')
    isfarah=target=="f"
    supplierid=request.POST.get('supplierid')
    whattopay=float(request.POST.get('whattopay'))
    moderegl=request.POST.get('moderegl')
    supplier=Supplier.objects.get(pk=supplierid)
    bons=json.loads(request.POST.get('bons'))
    avoirs=json.loads(request.POST.get('avoirs'))
    avances=json.loads(request.POST.get('avances'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    bank=json.loads(request.POST.get('bank'))
    source=json.loads(request.POST.get('source'))
    date=request.POST.get('date')
    # date=timezone.now().date()
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    if moderegl=='bl':
        livraisons=Itemsbysupplier.objects.filter(pk__in=bons)
    else:
        livraisons=Factureachat.objects.filter(pk__in=bons)
    avoirs=Avoirsupplier.objects.filter(pk__in=avoirs)
    avances=Avancesupplier.objects.filter(pk__in=avances)
    print('>> ll', livraisons)
    print('>> avances', avances)
    print('>> avoirs', avoirs)

    avoirs.update(inreglement=True)
    avances.update(inreglement=True)
    totalmantant=sum(mantant)
    totalbons=livraisons.aggregate(Sum('total'))['total__sum'] or 0
    # if whattopay<totalmantant:
    #     livraisons.update(ispaid=True)
    #     if moderegl=='facture':
    #         for livraison in livraisons:
    #             # Update 'ispaid' for related ManyToManyField (bons)
    #             print('>> bons of facture', livraison.bons.all())
    #             livraison.bons.all().update(ispaid=True)
    #     print('>> rel is ,ore than pay', totalmantant, whattopay)
    #     diff=round(totalmantant-whattopay, 2)
    #     print('>>>>>>< deiff', diff)
    #     # create avance supp
    #     Avancesupplier.objects.create(
    #         supplier_id=supplierid,
    #         amount=diff,
    #         date=date,
    #         isfarah=isfarah,
    #     )
    #if whattopay==totalmantant:
    livraisons.update(ispaid=True)
    if moderegl=='facture':
        print('facture is paid', livraisons.values())
        for livraison in livraisons:
            # Update 'ispaid' for related ManyToManyField (bons)
            livraison.bons.all().update(ispaid=True)
    # if whattopay>totalmantant:
    #     Avancesupplier.objects.create(
    #         supplier_id=supplierid,
    #         amount=totalmantant,
    #         date=date,
    #         isfarah=isfarah,
    #     )
    print('>> date suppp', date)
    for m, mod, np, ech, bk, s in zip(mantant, mode, npiece, echeance, bank, source):
        regl=PaymentSupplier.objects.create(
            supplier_id=supplierid,
            amount=m,
            date=date,
            echeance=ech,
            mode=mod,
            npiece=np,
            bank=bk,
            isfarah=isfarah,
            source=s
        )
        if moderegl=='bl':
            regl.bons.set(livraisons)
        else:
            regl.factures.set(livraisons)
            regl.usedinfacture=True
            regl.save()
        caisse=Caisse.objects.filter(target=target).first()
        if mod=='espece':
            if caisse:
                caisse.total-=m
                regl.targetcaisse=caisse
                regl.save()
                caisse.save()
        regl.avoirs.set(avoirs)
        regl.avances.set(avances)
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    supplier.rest=round(float(supplier.rest)-float(totalmantant), 2)
    supplier.save()
    return JsonResponse({
        "success":True
    })


def journalachat(request):
    target=request.GET.get('target')
    # items=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[:50]
    # bons=Bonlivraison.objects.filter(date__year=thisyear)
    # totaltotal=round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    # total2=round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # print('>>>>>', totaltotal, total2)
    ctx={
        'title':'Journal Achat',
        'target':target
    }
    return render(request, 'journalachat.html', ctx)

def laodjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.all()[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalacha-row">
            <td>{i.date}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def journalachatfc(request):
    items=Stockin.objects.filter(isfacture=True, date__year=thisyear)[:50]
    print('>>>>>>>>>>', round(Stockin.objects.filter(isfacture=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),)
    ctx={
        'title':'Journal Achat Facture',
        'items':items,
        'today':timezone.now().date(),
        'total':round(Stockin.objects.filter(isfacture=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':round(Stockin.objects.filter(isfacture=True, date__year=thisyear).aggregate(Sum('quantity'))['quantity__sum'] or 0, 2),
    }
    return render(request, 'journalachatfc.html', ctx)

def loadjournalachatfc(request):
    page = int(request.GET.get('page', 1))
    print(">>>>> journal achat")

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.filter(isfacture=True)[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalachafc-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total:.2f}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 100  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.order_by('-id')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalacha-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total:.2f}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })


def journalvente(request):
    target=request.GET.get('target')
    # items=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[:50]
    # bons=Bonlivraison.objects.filter(date__year=thisyear)
    # totaltotal=round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    # total2=round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # print('>>>>>', totaltotal, total2)
    ctx={
        'title':'Journal vente',
        'target':target
    }
    return render(request, 'journalvente.html', ctx)

def yeardatajournalv(request):
    year=request.GET.get('year')
    print(year)
    items=Livraisonitem.objects.filter(isfacture=False, date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalvente-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def yeardatajournalvfc(request):
    year=request.GET.get('year')
    print(year)
    items=Outfacture.objects.filter(date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalventefc-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td class="totaljv"></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def journalventefc(request):
    items=Outfacture.objects.filter(date__year=thisyear).order_by('-date')[:50]
    ctx={
        'title':'Journal vente Facture',
        'items':items,
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0,2),
        'today':timezone.now().date()
    }
    return render(request, 'journalventefc.html', ctx)

def loadjournalvente(request):
    page = int(request.GET.get('page', 1))
    year = request.GET.get('year')
    term = request.GET.get('term')
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    per_page = 50  # Adjust as needed
    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if term != '0':

        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(bon__bon_no__iregex=term))
        if year=='0':
            # means the year is not selected, so the records of the current year
            products = Livraisonitem.objects.filter(isfacture=False).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if startdate!='0' and enddate!='0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        products=Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).order_by('-date')[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[start:end]
    else:
        products = Livraisonitem.objects.filter(isfacture=False, date__year=year).order_by('-date')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalvente-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref.upper()}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalventefc(request):
    page = int(request.GET.get('page', 1))
    term=request.GET.get('term')
    year=request.GET.get('year')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term))
        if year=='0':
            # means the year i not selected, so the records of the current year
            products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
            total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
            total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalventefc-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'total':total,
            'totalqty':totalqty
        })

    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        ctx={
            'trs':trs,
        }
        if bons:
            ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
            ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
        return JsonResponse(ctx)
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Outfacture.objects.filter(date__year=thisyear).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Outfacture.objects.filter(date__year=year).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0

    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year}>
            <td>{i.date.strftime("%d/%m/%Y")}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen}</td>
            <td class="text-danger prachatjvfc">{i.product.buyprice}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td>{i.total}</td>
            <td></td>
            <td>{i.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page,
        'total':(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
    })

# product search selects2 for bon sortie,
def searchproductbonsortie(request):
    # get url pams
    term=request.GET.get('term').lower().strip()
    products=Produit.objects.filter(Q(ref__startswith=term) |Q(farahref__startswith=term))
    print('term>>', term)
    results=[]
    for i in products:
        ref=i.farahref if term.startswith('fr-') else i.ref
        results.append({
            'id':f'{i.ref}§{i.name}§{i.buyprice}§{i.stocktotalfarah}§{i.stockfacturefarah}§{i.stocktotalorgh}§{i.stockfactureorgh}§{i.id}§{i.sellprice}§{i.remisesell}§{i.prixnet}§{i.representprice}§{term}',
            'text':f'{ref.upper()} - {i.name.upper()}',
            'stock':i.stocktotalfarah+i.stocktotalorgh,
            'stockfacture':i.stockfacturefarah,
            'image':i.image.url if i.image else "",
            'mark':i.mark.name if i.mark else "",
            # return term to use it as adistinguisher
            'term':term
        })
    return JsonResponse({'results': results})
# regular search fro products
def searchproductforbonachat(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(diametre__icontains=term)|
            Q(cars__icontains=term)
        )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects)
    # check if term in product.ref or product.name

    results=[]
    for i in products:
        results.append({
            'id':f"{i.ref}§{i.name}§{i.frbuyprice if target=='f' else i.buyprice}§{i.stocktotalfarah if target=='f' else i.stocktotalorgh}§{i.stockfacturefarah if target=='f' else i.stocktotalorgh}§{i.id}§{i.frsellprice if target=='f' else i.sellprice}§{i.frremisesell if target=='f' else i.remisesell}§{i.prixnet}§{i.representprice}§{i.qtyjeu}",
            'text':f'{i.ref.upper()} - {i.name.upper()}',
            'stock':i.stocktotalfarah if target=='f' else i.stocktotalorgh,
            'stockfacture':i.stockfacturefarah if target=='f' else i.stocktotalorgh,
            # return term to use it as adistinguisher
            'term':term,
            'target':target
        })
    return JsonResponse({'results': results})

def searchproduct(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(diametre__icontains=term)|
            Q(cars__icontains=term)
        )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects)
    # check if term in product.ref or product.name

    results=[]
    print('>> products', products)
    print('>> products', products[0].ref)
    print('>> products', products[0].name)
    print('>> products', products)
    for i in products:
        results.append({
            'id':f"{i.ref}§{i.name}§{i.buyprice}§{i.stocktotalfarah if target=='f' else i.stocktotalorgh}§{i.stockfacturefarah if target=='f' else i.stocktotalorgh}§{i.id}§{i.frsellprice if target=='f' else i.sellprice}§{i.frremisesell if target=='f' else i.remisesell}§{i.prixnet}§{i.representprice}",
            'text':f'{i.ref.upper()} - {i.name.upper()}',
            'stock':i.stocktotalfarah if target=='f' else i.stocktotalorgh,
            'stockfacture':i.stockfacturefarah if target=='f' else i.stocktotalorgh,
            # return term to use it as adistinguisher
            'term':term,
            'image':i.image.url if i.image else "",
            'mark':i.mark.name if i.mark else "",
            'target':target
        })
    return JsonResponse({'results': results})


def filterbldate(request):
    target=request.GET.get('target')
    searchtype=request.GET.get('searchtype')
    isfarah=target=='f'
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> searchtype', searchtype)
    if searchtype=="waiting":
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False, iscanceled=False).order_by('-bon_no')[:50]
        total=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False, iscanceled=False).aggregate(Sum('total')).get('total__sum') or 0
    elif searchtype=="valid":
        print('>> searchtype is valid')
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).order_by('-bon_no')
        total=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).aggregate(Sum('total')).get('total__sum') or 0
    else:
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, iscanceled=True).order_by('-bon_no')
        total=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, iscanceled=True).aggregate(Sum('total')).get('total__sum') or 0
    # trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}?target={target}')">
    #         <td>{ i.bon_no }</td>
    #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #         <td>{ i.client.name }</td>
    #         <td>{ i.client.code }</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.client.region}</td>
    #         <td>{ i.client.city}</td>
    #         <td>{ i.client.soldbl}</td>
    #         <td>{ i.salseman }</td>
    #         <td class="d-flex justify-content-between">
    #           <div>
    #           {'R0' if i.ispaid else 'N1' }

    #           </div>
    #           <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #         </td>
    #         <td class="text-danger">
    #         {'OUI' if i.isfacture else 'NON'}

    #         </td>

    #         <td>
    #           {i.commande.order_no if i.commande else '--'}
    #         </td>
    #         <td>
    #           {i.modlvrsn}
    #         </td>
    #       </tr>
    #     '''
    ctx={
        'trs':render(request, 'bllist.html', {'bons':bons, 'notloading':True}).content.decode('utf-8'),
        'total':total

    }
    return JsonResponse(ctx)

def filterbsdate(request):
    target=request.GET.get('target')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Bonsortie.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[:50]
    ctx={
        'trs':render(request, 'bslist.html', {'bons':bons}).content.decode('utf-8')
    }
    if bons:
        ctx['total']=round(Bonsortie.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)



def searchclient(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    #regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        q_objects &= (Q(name__icontains=term) |
                Q(code__icontains=term) |
                Q(region__icontains=term) |
                Q(city__icontains=term) |
                Q(address__icontains=term))
    filter_params = {'clientsortie': True} if target == 's' else {'clientfarah': True} if target == 'f' else {'clientorgh': True}
    print('target >>', target)
    # Perform the query with the combined Q object and conditional filter parameters
    clients = Client.objects.filter(q_objects, **filter_params)

    # if '+' in term:
    #     term=term.split('+')
    #     for i in term:
    #         clients=Client.objects.filter(
    #             Q(name__icontains=i) |
    #             Q(code__icontains=i) |
    #             Q(region__icontains=i) |
    #             Q(city__icontains=i)
    #         )
    # else:
    #     clients=Client.objects.filter(
    #         Q(name__icontains=term) |
    #         Q(code__icontains=term) |
    #         Q(region__icontains=term) |
    #         Q(city__icontains=term)
    #     )
    results=[]
    for i in clients:
        results.append({
            'id':i.id,
            'text':f'{i.name} - {i.city}',
            'diver':i.diver
        })
    return JsonResponse({'results': results})
def searchclientrep(request):
    rep=request.user.represent
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(name__icontains=term) |
                Q(code__icontains=term) |
                Q(region__icontains=term) |
                Q(city__icontains=term) |
                Q(address__icontains=term))
    clients=Client.objects.filter(represent=rep).filter(q_objects)
    # if '+' in term:
    #     term=term.split('+')
    #     for i in term:
    #         clients=Client.objects.filter(
    #             Q(name__icontains=i) |
    #             Q(code__icontains=i) |
    #             Q(region__icontains=i) |
    #             Q(city__icontains=i)
    #         )
    # else:
    #     clients=Client.objects.filter(
    #         Q(name__icontains=term) |
    #         Q(code__icontains=term) |
    #         Q(region__icontains=term) |
    #         Q(city__icontains=term)
    #     )
    results=[]
    for i in clients:
        results.append({
            'id':i.id,
            'text':f'{i.name} - {i.city}'
        })
    return JsonResponse({'results': results})

def searchsupplier(request):
    term=request.GET.get('term')
    print(term)
    suppliers=Supplier.objects.filter(Q(name__icontains=term) | Q(phone__icontains=term)| Q(address__icontains=term)| Q(code__icontains=term))
    print('suppliers', suppliers)
    results=[]
    for i in suppliers:
        results.append({
            'id':i.id,
            'text':i.name
        })
    return JsonResponse({'results': results})

def getclientfactureprice(request):
    id=request.POST.get('id')
    clientid=request.POST.get('clientid')
    print(id, 'rr', clientid)
    try:
        clientprice=Outfacture.objects.filter(client_id=clientid, product_id=id).last()
        price=clientprice.price
        remise=clientprice.remise
        return JsonResponse({
            'price':price,
            'remise':remise
        })
    except:
        return JsonResponse({
            'price':0,
            'remise':0
        })


def updatereglebons(request):
    reglementid=request.GET.get('reglementid')
    mantant=request.GET.get('mantant')
    mode=request.GET.get('mode')
    bank=request.GET.get('bank')
    npiece=request.GET.get('npiece')
    date=datetime.strptime(request.GET.get('date'), '%Y-%m-%d')
    echeance=request.GET.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    # newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientbl.objects.get(pk=reglementid)
    # thisclient=reglement.client
    # #print('soldbl', float(thisclient.soldbl),float(reglement.amount),float(mantant))
    # thisclient.soldbl=round(float(thisclient.soldbl)+float(reglement.amount)-float(mantant), 2)
    # thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    # thisclient.save()
    # oldbons=reglement.bons.all()
    # livraisons=Bonlivraison.objects.filter(pk__in=newbons)
    # for i in oldbons:
    #     otherregl=PaymentClientbl.objects.filter(bons=i.id).exclude(pk=reglementid)
    #     if not otherregl.exists():
    #         i.ispaid=False
    #         i.statusreg='n1'
    #         i.save()
    # #oldbons.update(ispaid=False)
    # #oldbons.update(statusreg='n1')
    # #print(oldbons)
    # #print(livraisons)
    # livraisons.update(ispaid=True)
    # livraisons.update(statusreg='r0')
    # reglement.bons.clear()

    # # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.bank=bank
    reglement.npiece=npiece
    reglement.echance=echeance

    print('>>> bank', bank)
    #reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def updateavanceclient(request):
    avanceid=request.GET.get('avanceid')
    mantant=request.GET.get('mantant')
    mode=request.GET.get('mode')
    bank=request.GET.get('bank')
    note=request.GET.get('note')
    bonofavance=request.GET.get('bonofavance')
    npiece=request.GET.get('npiece')
    date=datetime.strptime(request.GET.get('date'), '%Y-%m-%d')
    echeance=request.GET.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    # newbons=json.loads(request.POST.get('bons'))
    avance=Avanceclient.objects.get(pk=avanceid)
    
    avance.date=date
    avance.amount=mantant
    avance.mode=mode
    avance.bank=bank
    avance.npiece=npiece
    avance.bonofavance=bonofavance
    avance.note=note
    avance.echeance=echeance

    print('>>> bank', bank)
    #avance.bons.set(livraisons)
    avance.save()
    return JsonResponse({
        'success':True
    })

def updatereglesupp(request):
    reglementid=request.GET.get('reglementid')
    mantant=request.GET.get('mantant')
    mode=request.GET.get('mode')
    npiece=request.GET.get('npiece')
    date=datetime.strptime(request.GET.get('date'), '%Y-%m-%d')
    echeance=request.GET.get('echeance')
    bank=request.GET.get('bank')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    reglement=PaymentSupplier.objects.get(pk=reglementid)

    # thissupplier=reglement.supplier
    # thissupplier.rest=round(float(thissupplier.rest)+float(reglement.amount)-float(mantant), 2)
    # thissupplier.save()
    # oldbons=reglement.bons.all()
    # livraisons=Itemsbysupplier.objects.filter(pk__in=newbons)
    # oldbons.update(ispaid=False)
    # livraisons.update(ispaid=True)
    # reglement.bons.clear()

    # # # update regleemnt amount
    reglement.date=date
    reglement.bank=bank
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece
    reglement.echeance=echeance

    #reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getreglementbl(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)
    # get bons
    target=request.GET.get('target')
    bons=reglement.bons.all()
    if bons:
        otherregl=PaymentClientbl.objects.filter(bons__in=bons).exclude(pk=reglement.id)
        print('>> has bons', bons)
    # get factures
    factures=reglement.factures.all()
    if factures:
        # get other regl with the same factures
        otherregl=PaymentClientbl.objects.filter(factures__in=factures).exclude(pk=reglement.id)
        print('>> has factures', factures)
    #print('>> other regl', otherregl)
    print('>>>>', reglement.issortie)
    if reglement.issortie:
        template='updatereglebs.html'
        bons=reglement.bonsortie.all().order_by('date')
        livraisons=Bonsortie.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[:50]
    else:
        template='updatereglebl.html'
        bons=reglement.bons.all().order_by('date')
        livraisons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[:50]
    # bons without bons in reglement
    # livraisons=Bonlivraison.objects.filter(client=reglement.client)
    #we need bons to calculate total bl
    bonstocalculate=Bonlivraison.objects.filter(client=reglement.client)
    # trs=''
    # for i in livraisons:
    #     trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'
    print('>> d', bons)
    ctx={
        #'otherregl':otherregl,
        'reglement':reglement,
        'avoirs':reglement.avoirs.all().order_by('date'),
        'avances':reglement.avances.all().order_by('date'),
        'bons':bons,
        'factures':reglement.factures.all().order_by('date'),
        'today':timezone.now().date(),
        'banks':Bank.objects.filter(target=target)
    }
    
    return JsonResponse({
        'html':render(request, template, ctx).content.decode('utf-8')
    })
    return JsonResponse({

        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'avoirs':render(request, 'avoirsofregl.html', {'avoirs':reglement.avoirs.all().order_by('date')}),
        'avances':render(request, 'avancesofregl.html', {'avances':reglement.avances.all().order_by('date')}),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'soldclient':reglement.client.soldbl
    })



def getreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    # facture of this reglement
    bons=reglement.factures.all()
    print([i.id for i in bons] )
    livraisons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[:50]
    bonstocalculate=Facture.objects.filter(client=reglement.client)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        # 'livraisons':list(livraisons.values()),
        'soldclientfc':reglement.client.soldfacture,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),


    })

def updatereglefactures(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientfc.objects.get(pk=reglementid)
    thisclient=reglement.client

    # get bons
    oldbons=reglement.factures.all()
    for i in oldbons:
        # search in other reglement
        otherregl=PaymentClientfc.objects.filter(factures=i.id).exclude(pk=reglementid)

        if not otherregl.exists():
            # this means there is no other regl in this fact
            i.ispaid=False
            i.statusreg='b1'
            i.save()
    # instead of update them to not paid, above code is correct, loop over them and if fact has no other reg then make it unpaid
    # oldbons.update(ispaid=False)
    # oldbons.update(statusreg='b1')
    reglement.factures.clear()
    livraisons=Facture.objects.filter(pk__in=newbons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')

    # new code, we get all nthe

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece
    reglement.echance=echeance
    reglement.factures.set(livraisons)

    # reglement.bons.set(livraisons)
    reglement.save()
    print(">> adding amount to soldfacture ans soldtotal, I made it at the end so that if there will be any error above the sold will not be affected")
    thisclient.soldfacture=round(float(thisclient.soldfacture)+float(reglement.amount)-float(mantant), 2)
    thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    thisclient.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getlastsuppprice(request):
    id=request.POST.get('id')
    isfarah=request.POST.get('target')=='f'
    supplierid=request.POST.get('supplierid')
    print(id, 'rr', supplierid)
    price=0
    remise=0
    print('isfarah', isfarah)
    prices=Stockin.objects.filter(product_id=id, supplier_id=supplierid, isfarah=isfarah, isavoir=False)
    lastprice=prices.last()
    if lastprice:
        price=lastprice.price
        remise=lastprice.remise1
    print('supp', Supplier.objects.get(pk=supplierid))
    print('pdct', Produit.objects.get(pk=id))
    print("prices", prices, 'lastprice', lastprice)
    return JsonResponse({
        'price':price,
        'remise':remise,
        # 'facture':lastprice.facture,
        'table':render(request, 'suppprodctprices.html', {'producthistory':prices.order_by('-date')[:10]}).content.decode('utf-8')
    })


def boncommandedetails(request, id):

    order=Order.objects.get(pk=id)
    orderitems=Orderitem.objects.filter(order=order).order_by('product__name')
    reliquat=Orderitem.objects.filter(order__client_id=order.client.id, order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False).exists()

    print('>>>>>>Reliquat', reliquat)
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+40] for i in range(0, len(orderitems), 40)]
    ctx={
        'hasreliquat':reliquat,
        'title':f'Bon de Commande {order.order_no}',
        'order':order,
        'reliquat':'Reliquat' in order.note,
        'orderitems':orderitems,
    }
    return render(request, 'boncommandedetails.html', ctx)


def genererbonlivraison(request, id):

    order=Order.objects.get(pk=id)
    reliquas=Orderitem.objects.filter(order__client_id=order.client.id, order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False)
    items=Orderitem.objects.filter(order=order).order_by('name')
    # we need date of last invoice

    #year = timezone.now().strftime("%y")
    #latest_receipt = Bonlivraison.objects.filter(
    #    bon_no__startswith=f'BL{year}'
    #).order_by("-id").first()
    #if latest_receipt:
    #    latest_receipt_no = int(latest_receipt.bon_no[-5:])
    #    receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    #else:
    #    receipt_no = f"BL{year}00001"
    ctx={
        'reliquas':reliquas,
        'order':order,
        'items':items,
        #'receipt_no':receipt_no,
        #'clients':Client.objects.all(),
        'reps':Represent.objects.all(),
        'today':timezone.now().date()
    }
    return render(request, 'genererbonlivraison.html', ctx)


def createclientaccount(request):
    clientid=request.POST.get('clientid')
    client= Client.objects.get(pk=clientid)
    username=request.POST.get('username')
    password=request.POST.get('password')
    #check for username
    user=User.objects.filter(username=username).first()
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })

    # create user
    user=User.objects.create_user(username=username, password=password)
    # assign user to client
    group, created = Group.objects.get_or_create(name="clients")
    user.groups.add(group)
    user.save()
    client.user=user
    client.save()
    # req.get('http://serverip/products/createclientaccount', {
    #     'clientcode':client.code,
    #     'username':username,
    #     'password':password
    # })
    return JsonResponse({
        'success':True
    })


def createrepaccount(request):
    repid=request.POST.get('repid')
    rep= Represent.objects.get(pk=repid)
    username=request.POST.get('username')
    password=request.POST.get('password')
    #check for username
    user=User.objects.filter(username=username).first()
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })

    # create user
    user=User.objects.create_user(username=username, password=password)
    # assign user to rep
    group, created = Group.objects.get_or_create(name="salsemen")
    user.groups.add(group)
    user.save()
    rep.user=user
    rep.save()
    return JsonResponse({
        'success':True
    })

def carlogospage(request):
    ctx={
        # maintain same names
        'categories':Carlogos.objects.all(),
        'title':'Voitures logo'
    }
    return render(request, 'carlogos.html', ctx)

def createlogo(request):
    name=request.POST.get('logoname')
    # get image file
    image=request.FILES.get('logoimage')
    # create category
    Carlogos.objects.create(name=name, image=image)
    ctx={
        'categories':Carlogos.objects.all(),
        'title':'Voiture logo'
    }
    return JsonResponse({
        'html':render(request, 'carlogos.html', ctx).content.decode('utf-8')
    })

def updatelogo(request):
    image=request.FILES.get('updatelogoimage') or None
    id=request.POST.get('id')
    carlogo=Carlogos.objects.get(pk=id)
    carlogo.name=request.POST.get('updatelogoname')
    if image:
        carlogo.image=image
    carlogo.save()
    ctx={
        'categories':Carlogos.objects.all(),
        'title':'Voiture logo'
    }
    return JsonResponse({
        'html':render(request, 'carlogos.html', ctx).content.decode('utf-8')
    })


def getnotpaid(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0).order_by('-date')


    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons, 'notloading':True}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),

    })

def getnotpaidfc(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Facture.objects.filter(date__lt=three_months_ago, ispaid=False)


    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    })


def filterfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    target=request.GET.get('target')
    # using isvalid directly, when waiting ==1 means we want waiting list
    isvalid=not request.GET.get('waiting')=='1'
    isfarah=target=='f'
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    print('>> isvalid', isvalid)
    bons=Facture.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid).order_by('-facture_no')[:50]
    total=round(Facture.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid).aggregate(Sum('total')).get('total__sum') or 0, 2)
    totaltva=round(Facture.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid).aggregate(Sum('tva')).get('tva__sum') or 0, 2)
    # trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
    #         <td>{ i.facture_no }</td>
    #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.tva}</td>
    #         <td>{ i.client.name }</td>
    #         <td>{ i.client.code }</td>
    #         <td>{ i.client.region}</td>
    #         <td>{ i.client.city}</td>
    #         <td>{ i.client.soldfacture}</td>
    #         <td>{ i.salseman }</td>
    #         <td class="d-flex justify-content-between">
    #           <div>
    #           {'R0' if i.ispaid else 'N1' }

    #           </div>
    #           <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #         </td>
    #         <td class="text-danger">

    #         </td>

    #         <td>
    #           {i.bon.bon_no if i.bon else "--"}
    #         </td>
    #       </tr>
    #     '''
    ctx={
        'trs':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':total,
        'totaltva':totaltva,
    }
    return JsonResponse(ctx)
    # return JsonResponse({
    #     'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
    #     'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),
    #     'totaltva':round(bons.aggregate(Sum('tva')).get('tva__sum'), 2),

    # })

def filterachatdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    isvalid=request.GET.get('wanted')=='valid'
    isfarah=request.GET.get('target')=='f'
    print('>>', startdate, enddate, isvalid, isfarah)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Itemsbysupplier.objects.filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid)
    return JsonResponse({
        'html':render(request, 'bonachattrs.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    })

def updateclientpassword(request):
    clientid=request.POST.get('clientid')
    password=request.POST.get('password')
    try:
        user=Client.objects.get(pk=clientid).user
        user.set_password(password)
        user.save()
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })

def updatereppassword(request):
    repid=request.POST.get('repid')
    password=request.POST.get('password')
    try:
        user=Represent.objects.get(pk=repid).user
        print(user, 'user', password)
        user.set_password(password)
        user.save()
        # logout(user)

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })



def filterreglbldate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientbl.objects.filter(date__range=[startdate, enddate])
    print(bons)
    return JsonResponse({
        'html':render(request, 'reglbllist.html', {'bons':bons}).content.decode('utf-8'),
        #'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })


def filterreglfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientfc.objects.filter(date__range=[startdate, enddate])
    return JsonResponse({
        'html':render(request, 'reglfclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })



def sortupbl(request):
    bons=Bonlivraison.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownbl(request):
    bons=Bonlivraison.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortupfc(request):
    bons=Facture.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownfc(request):
    bons=Facture.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

def excelclients(request):
    target=request.POST.get('target')
    isfarah=target=='f'
    isorgh=target=='o'
    issortie=target=='s'
    print('isfarah', target=='f')
    print('isorgh', target=='o')
    print('issortie', target=='s')
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna('')
    if isfarah:
        prfx='CF-'
    elif isorgh:
        prfx='CO-'
    else:
        prfx='CP-'
    ids=1
    for d in df.itertuples():
        name=d.name
        city=d.city
        code=f'{prfx}{ids}'
        phone2="" if pd.isna(d.phone2) else str(d.phone2)
        phone="" if pd.isna(d.phone) else str(d.phone)
        ice="" if pd.isna(d.ice) else str(d.ice)
        

        print('>> there is no cli')
        client=Client.objects.create(
            code=code,
            name=name,
            city=city,
            ice=ice,
            phone=phone,
            phone2=phone2,
            clientorgh=isorgh,
            clientfarah=isfarah,
            clientsortie=issortie
        )
        print('>> created')
        ids+=1
    return JsonResponse({
        'success':True
    })

def excelsupp(request):
    
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna('')
    id=1
    for d in df.itertuples():
        code=f'FOR{id}'
        name=d.name
        city="" if pd.isna(d.city) else str(d.city)
        phone2="" if pd.isna(d.phone2) else str(d.phone2)
        phone="" if pd.isna(d.phone) else str(d.phone)
        email="" if pd.isna(d.email) else str(d.email)
        ice="" if pd.isna(d.ice) else str(d.ice)
        
        print('>> there is no cli')
        supp=Supplier.objects.create(
            name=name,
            code=code,
            city=city,
            ice=ice,
            email=email,
            phone=phone,
            phone2=phone2
        )
        print('>> created')
        id+=1
    return JsonResponse({
        'success':True
    })


def excelpdcts(request):
    #target=request.GET.get('targt')
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        farahref=f'fr-{ref}'
        name = d.name
        refeq = '' if pd.isna(d.refeq) else d.refeq
        #status = False if pd.isna(d.status) else True
        #coderef = '' if pd.isna(d.code) else d.code
        #diam = '' if pd.isna(d.diam) else d.diam
        #qty = 0 if pd.isna(d.qty) else d.qty
        #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        #devise = 0 if pd.isna(d.devise) else d.devise
        
        #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        #ctg = None if pd.isna(d.ctg) else d.ctg
        #order = '' if pd.isna(d.order) else d.order
        #img = None if pd.isna(d.img) else d.img
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        print('>> adding ', ref)
        product=Notavailable.objects.create(
            ref=ref,
            name=name,
            equiv=refeq,
        )
        # try:
        #     ref = d.ref.lower().strip()
        # except:
        #     ref=d.ref
        # #reps=json.dumps(d.rep)
        # farahref=f'fr-{ref}'
        # name = d.name
        # mark = None if pd.isna(d.mark) else d.mark
        # category = None if pd.isna(d.category) else d.category
        # refeq = '' if pd.isna(d.refeq) else d.refeq
        # #status = False if pd.isna(d.status) else True
        # #coderef = '' if pd.isna(d.code) else d.code
        # #diam = '' if pd.isna(d.diam) else d.diam
        # #qty = 0 if pd.isna(d.qty) else d.qty
        # qtyjeu = 0 if pd.isna(d.qtyjeu) else d.qtyjeu
        # unite = 0 if pd.isna(d.unite) else d.unite
        # #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        # #devise = 0 if pd.isna(d.devise) else d.devise
        
        # #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        # #ctg = None if pd.isna(d.ctg) else d.ctg
        # #order = '' if pd.isna(d.order) else d.order
        # #img = None if pd.isna(d.img) else d.img
        # #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        # product=Produit.objects.create(
        #     ref=ref,
        #     name=name,
        #     category_id=category,
        #     unite=unite,
        #     mark_id=mark,
        #     qtyjeu=qtyjeu,
        #     minstock=1,
        #     equivalent=refeq,
        #     farahref=farahref
        # )

    print('>>>', entries)
    return JsonResponse({
        'success':True
    })

def excelqtypdcts(request):
    target=request.POST.get('target')
    isfarah=target=='f'
    print('>> isfarah', isfarah, )
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        qty = 0 if pd.isna(d.qty) else d.qty
        price = 0 if pd.isna(d.price) else d.price
        try:
            print('entering', ref)
            product=Produit.objects.get(ref=ref)
            if isfarah:
                product.stocktotalfarah+=qty
                product.frstockinitial=qty
                product.frpriceinitial=price
            else:
                product.stocktotalorgh+=qty
                product.stockinitial=qty
                product.priceinitial=price
            product.save()
            entries+=1

        except Exception as e:
            print('>>',e, ref)
            with open('error.txt', 'a') as ff:
                ff.write(f'>> {e} {ref}')
            
    print('>>>', entries)
    return JsonResponse({
        'success':True
    })

def excelmarks(request):
    #target=request.GET.get('targt')
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        name=d.mark
        Mark.objects.create(name=name)
            
    print('>>>', entries)
    return JsonResponse({
        'success':True
    })

def excelcategories(request):
    #target=request.GET.get('targt')
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        name=d.category
        Category.objects.create(name=name)
            
    print('>>>', entries)
    return JsonResponse({
        'success':True
    })

def deactivateaccount(request):
    userid=request.GET.get('userid')
    user=User.objects.get(id=userid)
    user.is_active=False
    user.save()
    # req.get('http://serverip/products/deactivateaccount', {
    #     'username':user.username,
    # })
    # delete user session in django session
    UserSession.objects.filter(user=user).delete()
    # Clear the user's session
    #Session.objects.filter(session_key__in=UserSession.objects.filter(user=user).values('session_key')).delete()

    return JsonResponse({
        'success':True
    })

def activateaccount(request):
    userid=request.GET.get('userid')
    user=User.objects.get(id=userid)
    user.is_active=True
    user.save()
    # req.get('http://serverip/products/activateaccount', {
    #     'username':user.username,
    # })
    return JsonResponse({
        'success':True
    })

def stocksection(request):
    target=request.GET.get('target')
    categories=Category.objects.all()
    products=Produit.objects.all()[:50]
    ctx={'categories':categories,
        'title':'Liste des Articles',
        'products':products,
        'target':target
        # 'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        # 'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,


    }
    return render(request, 'stocksection.html', ctx)

def stock(request):
    target=request.GET.get('target')
    categories=Category.objects.all()
    products=Produit.objects.all()[:50]
    ctx={
        'categories':categories,
        'isfarah':target=='f',
        'title':'Liste des Articles',
        'products':products,
        'target':target
        # 'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        # 'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,


    }
    return render(request, 'stock.html', ctx)


def getreglementsupp(request, id):
    reglement=PaymentSupplier.objects.get(pk=id)
    # ctx={
    #     'title':'Reglement',
    #     'reglement':reglement,
    # }
    # bons=reglement.bons.all()
    # # bons without bons in reglement
    # livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier).exclude(pk__in=[bon.pk for bon in bons])
    # # livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    # #we need bons to calculate total bl
    # bonstocalculate=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    # trs=''
    # for i in livraisons:
    #     trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementssupp.exists() else ""}" class="loadblinupdatereglsupp" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.nbon}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementssupp.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    target=request.GET.get('target')
    ctx={
        'reglement':reglement,
        'avoirs':reglement.avoirs.all().order_by('date'),
        'avances':reglement.avances.all().order_by('date'),
        'bons':reglement.bons.all().order_by('date'),
        'factures':reglement.factures.all().order_by('date'),
        'today':timezone.now().date(),
        'banks':Bank.objects.filter(target=target)
    }
    return JsonResponse({
        'html':render(request, 'updatereglesupp.html', ctx).content.decode('utf-8')
    })
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echeance.strftime('%Y-%m-%d') if reglement.echeance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'supplier':reglement.supplier.name,
        'mantant':reglement.amount,
        'livraisons':list(livraisons.values()),
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'], 2),
        'soldsupplier':round(reglement.supplier.rest, 2)
    })


def etatblclients(request):

    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    print('>>>>>', first_day_of_year, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))

    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%m/%y'))  # e.g., "January 2024"
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    clients = Client.objects.filter().order_by('city').exclude(diver=True)

    serialized_data = []
    #client=Client.objects.get(pk=3758)
    for clientindex, client in enumerate(clients):
        sitdata=0
        client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}

        # Filter data for the specified date range
        bons = Bonlivraison.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )

        avoirs = Avoirclient.objects.exclude(avoirfacture=True).filter(
            client=client,
            date__range=[start_date, end_date]
        )

        regls = PaymentClientbl.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        # Initialize monthly data
        monthly_data = {month: {'bons': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}

        # Populate monthly data with bons, avoirs, and regls
        for bon in bons:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['bons'] += bon.total

        for avoir in avoirs:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirs'] += avoir.total

        for regl in regls:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['regls'] += regl.amount
        # Calculate the client situation for each month and aggregate totals
        total_bons = total_avoirs = total_regls = 0
        # this for testing specific clients
        # if client.id==3785:
        #     print('>>>>>>>>> bons', bons, 'avoirs', avoirs, 'regls', regls, )
        for monthindex, month in enumerate(months):
            total_bons += monthly_data[month]['bons']
            total_avoirs += monthly_data[month]['avoirs']
            total_regls += monthly_data[month]['regls']
            #monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thisreg=monthly_data[month]['regls']
            if thisreg:
                for b in range(monthindex):
                    thisreg-=client_data['monthly_data'][b]['situation']
                    client_data['monthly_data'][b]['situation']=0
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - thisreg, 2)
            #     print('>>>>>>>>',client_data)
            #     #thismonthsit=-3000
            #     previousmonth=months[months.index(month)-1]
            #     dpreviousmonth=months[months.index(month)-2]
            #     treviousmonth=months[months.index(month)-3]
            #     diff=monthly_data[previousmonth]['situation']+thismonthsit
            #     if diff < 0:
            #         monthly_data[month]['situation'] = diff
            #         monthly_data[previousmonth]['situation']=0
            #         for month_data in client_data['monthly_data']:
            #             if month_data['month'] == previousmonth:
            #                 if month_data['situation'] == 0.00:
            #                     break


            #     else:
            #         monthly_data[previousmonth]['situation']=diff
            #         for month_data in client_data['monthly_data']:
            #             if month_data['month'] == previousmonth:
            #                 month_data['situation'] = diff
            #                 break
            #         monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'], 2)
            # else:
            #     monthly_data[month]['situation']=thismonthsit


            client_data['monthly_data'].append({
                'month': month,
                'bons': monthly_data[month]['bons'],
                'avoirs': monthly_data[month]['avoirs'],
                'regls': monthly_data[month]['regls'],
                'situation': thismonthsit
            })

        # Calculate total situation for the client
        client_data['totalsituation'] = round(total_bons - total_avoirs - total_regls, 2)
        serialized_data.append(client_data)
        # Define start and end months for the date range


    #return render(request, 'etatblclients.html')
    return render(request, 'etatblclients.html', {'title': 'Etat bl client', 'data': serialized_data, 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str})
    # clients = Client.objects.filter(soldbl__gt=0).exclude(diver=True).order_by('city')
    # current_year = timezone.now().year
    # current_month = timezone.now().month
    # monthtostart=int(request.GET.get('monthtostart', 1))
    # monthtoend=int(request.GET.get('monthtoend', 5))
    # # Initialize serialized data
    # serialized_data = []
    # months=[i for i in range(monthtostart, monthtoend+1)]
    # for client in clients:
    #     client_data = {'client_name': client.name, 'client_city': client.city, 'client_region': client.region,  'client_represent': client.represent,'monthly_data': [], 'totalsituation': 0}

    #     # Filter data for the current year from January to the current month
    #     bons = Bonlivraison.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend],
    #         total__gt=0
    #     ).order_by('date')

    #     avoirs = Avoirclient.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend]
    #     ).order_by('date')

    #     regls = PaymentClientbl.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend]
    #     ).order_by('date')

    #     # Initialize monthly data for each month from January to current month
    #     monthly_data = {month: {'bons': 0, 'avoirs': 0, 'regls': 0} for month in range(monthtostart, monthtoend + 1)}
    #     # Populate monthly data with bons, avoirs, and regls
    #     for bon in bons:
    #         month = bon.date.month
    #         monthly_data[month]['bons'] += bon.total

    #     for avoir in avoirs:
    #         month = avoir.date.month
    #         monthly_data[month]['avoirs'] += avoir.total

    #     for regl in regls:
    #         month = regl.date.month
    #         monthly_data[month]['regls'] += regl.amount

    #     # Calculate the client situation for each month and aggregate totals
    #     total_bons = total_avoirs = total_regls = 0

    #     for month in range(monthtostart, monthtoend + 1):
    #         monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
    #         total_bons += monthly_data[month]['bons']
    #         total_avoirs += monthly_data[month]['avoirs']
    #         total_regls += monthly_data[month]['regls']
    #         client_data['monthly_data'].append({
    #             'month': month,

    #             'situation': monthly_data[month]['situation']
    #         })

    #     # Calculate total situation for the client
    #     client_data['totalsituation'] = round(total_bons - total_avoirs - total_regls, 2)
    #     serialized_data.append(client_data)
    # # Return the serialized data as a JSON response
    # return render(request, 'etatblclients.html', {'data': serialized_data, 'months':months, 'monthtostart':monthtostart, 'monthtoend':monthtoend})


def etatfcclients(request):
    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%m/%y'))  # e.g., "January 2024"
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    clients = Client.objects.order_by('city').exclude(diver=True)

    serialized_data = []
    #client=Client.objects.get(pk=3417)
    for clientindex, client in enumerate(clients):
        sitdata=0
        client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}

        # Filter data for the specified date range
        factures = Facture.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )

        avoirs = Avoirclient.objects.filter(
            client=client,
            avoirfacture=True,
            date__range=[start_date, end_date]
        )

        regls = PaymentClientfc.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        # Initialize monthly data
        monthly_data = {month: {'factures': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}

        # Populate monthly data with factures, avoirs, and regls
        for bon in factures:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['factures'] += bon.total
            sitdata+=bon.total

        for avoir in avoirs:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirs'] += avoir.total
            sitdata-=bon.total

        for regl in regls:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['regls'] += regl.amount
            sitdata-=bon.total
        # Calculate the client situation for each month and aggregate totals
        total_factures = total_avoirs = total_regls = 0

        for monthindex, month in enumerate(months):
            total_factures += monthly_data[month]['factures']
            total_avoirs += monthly_data[month]['avoirs']
            total_regls += monthly_data[month]['regls']
            #monthly_data[month]['situation'] = round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thisreg=monthly_data[month]['regls']
            if thisreg:
                for b in range(monthindex):
                    thisreg-=client_data['monthly_data'][b]['situation']
                    client_data['monthly_data'][b]['situation']=0
            thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - thisreg, 2)



            client_data['monthly_data'].append({
                'month': month,
                'factures': monthly_data[month]['factures'],
                'avoirs': monthly_data[month]['avoirs'],
                'regls': monthly_data[month]['regls'],
                'situation': thismonthsit
            })

        # Calculate total situation for the client
        client_data['totalsituation'] = round(total_factures - total_avoirs - total_regls, 2)
        serialized_data.append(client_data)
        # Define start and end months for the date range


    #return render(request, 'etatblclients.html')
    return render(request, 'etatblclients.html', {'etatfacture':True ,'title': 'Etat fc client', 'data': serialized_data, 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str})

def updatebonavoir(request):
    id=request.POST.get('bonid')
    target=request.POST.get('target')
    isfarah=target=='f'
    avoir=Avoirclient.objects.get(pk=id)
    client=Client.objects.get(pk=request.POST.get('clientid'))
    mantant=json.loads(request.POST.get('mantant'))
    bank=json.loads(request.POST.get('bank'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    # date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    # we need avoir n° cause delete avoir will delete id, id is used in avoir n°
    avoirno=avoir.no
    avoiritems=Stockin.objects.filter(avoir=avoir)
    totalbon=request.POST.get('totalbon')
    newmode=request.POST.get('mode')
    isfacture=True if newmode=='facture' else False
    print("target", target)
    thisclient=avoir.client
    # regle stock
    # # regle soldclient
    # if avoir.avoirfacture:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.stockfacture+=i.qty
    #         i.product.save()
    #         i.delete()
    # else:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.save()
    #         i.delete()

    # # delete old avoir
    # # create new avoir
    # if avoir.client==client:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total)-float(totalbon), 2)
    #     if avoir.avoirfacture:
    #         thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
    #     else:
    #         thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
    #     if isfacture:
    #         thisclient.soldfacture=round(float(thisclient.soldfacture)-float(totalbon), 2)
    #     else:
    #         thisclient.soldbl=round(float(thisclient.soldbl)-float(totalbon), 2)

    #     thisclient.save()
    # else:
    #     # not the same client
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     # add sold to old client
    #     if avoir.avoirfacture:
    #         thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
    #     else:
    #         thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
    #     thisclient.save()
    #     # add sold to new client
    #     client.soldtotal=round(float(client.soldtotal)- float(totalbon), 2)
    #     if isfacture:
    #         client.soldfacture=round(float(client.soldfacture)- float(totalbon), 2)
    #     else:
    #         client.soldbl=round(float(client.soldbl)- float(totalbon), 2)
    #     client.save()
    #     print('new', client.soldtotal)
    items=Stockin.objects.filter(avoir=avoir)
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        if target == 'f':
            product.stocktotalfarah=float(product.stocktotalfarah)-float(i.quantity)
        elif target == 'o':
            product.stocktotalorgh=float(product.stocktotalorgh)-float(i.quantity)
        # if avoir.avoirfacture:
        #     product.stockfacture=int(product.stockfacture)-int(i.qty)
        product.save()
        i.delete()
    avoir.client=client
    #avoir.representant_id=request.POST.get('repid')
    # caisse=Caisse.objects.filter(target=target).first()
    # if caisse:
    #     caisse.total+=avoir.total
    #     caisse.save()
    avoir.total=totalbon
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    avoir.date=datebon
    #avoir.no=request.POST.get('orderno')

        #client.soldbl=5
    avoir.save()
    # update this items


    print('client:', avoir.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            product=Produit.objects.get(pk=i['productid'])
            if target == 'f':
                product.stocktotalfarah=float(product.stocktotalfarah)+float(i['qty'])
            elif target == 'o':
                product.stocktotalorgh=float(product.stocktotalorgh)+float(i['qty'])
            # if isfacture:
            #     product.stockfacture=int(product.stockfacture)+int(i['qty'])
            product.save()
            Stockin.objects.create(
                avoir=avoir,
                product=product,
                quantity=i['qty'],
                remise1=i['remise'],
                price=i['price'],
                total=i['total'],
                isavoir=True,
                isfarah=target=='f',
                isorgh=target=='o',
                qtyofprice=i['qty'],
                date=datebon,
            )
    
    totalamount=sum([i for i in mantant if i is not None])
    if totalamount>0:
        print('totalamount', totalamount)
        avoir.ispaid=True
        avoir.save()
        for m, mod, np, ech, bk in zip(mantant, mode, npiece, echeance, bank):
            if m is not None:
                regl=PaymentClientbl.objects.create(
                    client=client,
                    amount=m,
                    #today
                    date=timezone.now().date(),
                    echance=ech,
                    bank=bk,
                    mode=mod,
                    npiece=np,
                    isfarah=target=='f',
                    isorgh=target=='o',
                    issortie=target=='s',
                    isavoir=True
                )
                regl.avoirs.set([avoir])
                # caisse=Caisse.objects.filter(target=target).first()
                # if mod=='espece':
                #     if caisse:
                #         caisse.total+=m
                #         caisse.save()

    return JsonResponse({
        'success':True
    })

def notifyadmin(request):
    oldnotif=Ordersnotif.objects.filter(isread=True)
    oldnotif.delete()
    newnotif=Ordersnotif.objects.filter(isread=False)
    return JsonResponse({
        "length": newnotif.count(),
    })

def disablenotif(request):
    newnotif=Ordersnotif.objects.filter(isread=False)
    newnotif.update(isread=True)
    return JsonResponse({
        'success':True
    })

def listecheance(request):
    # get payments that are cheque or effet in mode
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today).order_by('-refused').order_by('-echance')
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today).order_by('-refused').order_by('-echance')
    echeance = chain(*[
    ((rbl, 'bl') for rbl in reglbl),
    ((rfc, 'fc') for rfc in reglfc),
    ])
    totalbl=reglbl.aggregate(Sum('amount'))['amount__sum'] or 0
    totalfc=reglfc.aggregate(Sum('amount'))['amount__sum'] or 0
    total=round(totalbl+totalfc, 2)
    # Sort the items by date
    sorted_echeance = sorted(echeance, key=lambda item: item[0].refused, reverse=True)
    ctx={
        'title':'List des echeances Actuel',
        'echeances':sorted_echeance,
        'total':total
    }
    return render(request, 'listecheance.html', ctx)


def echeancetoday(request):
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance__lte=today).count()
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance__lte=today).count()
    print('>>>>>',PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), today)
    return JsonResponse({
        'length':reglbl+reglfc
    })
def tabs(request):
    return render(request, 'tabs.html')


def getconnectedusers(request):
    # more than 5 minutes means the user is not connected
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago)
    connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)
    length=connected.count()
    trs=''
    for i in connected:
        trs+=f"""
        <tr>
        <td>{i.user.username}</td>
        <td>{i.user.groups.all()[0].name}</td>
        <td>{i.activity}</td>
        <td>{(i.lasttime).strftime('%Y-%m-%d %H:%M:%S')}</td>
        </tr>
        """
    print('connected', connected)
    print('notconnected', notconnected)
    return JsonResponse({
        'length':length,
        'trs':trs
    })

# this will pay regl of clients
def payreglbl(request):
    print('>>>< js', request.GET.get('bank'))
    ids=json.loads(request.GET.get('ids'))
    bank=request.GET.get('bank')
    dateregl=request.GET.get('dateregl')
    nrecu=request.GET.get('nrecu')
    print('>>>> ids, bank, nrecu', ids, bank, nrecu, dateregl)
    regl=PaymentClientbl.objects.filter(pk__in=ids)
    print('>>>> regl', regl)
    bank=Bank.objects.get(pk=bank)
    # 7iyd total 4lbanka
    totalamount=regl.aggregate(Sum('amount'))['amount__sum']
    bank.total+=totalamount
    bank.save()
    regl.update(ispaid=True)
    regl.update(nrecu=nrecu)
    regl.update(dateregl=dateregl)
    regl.update(targetbank=bank)
    # regl.update(targebank=Bank.objects.get)
    return JsonResponse({
        'success':True
    })

def payreglsupp(request):
    print('>>>< js', request.GET.get('bank'))
    ids=json.loads(request.GET.get('ids'))
    bank=request.GET.get('bank')
    dateregl=request.GET.get('ateregl')
    nrecu=request.GET.get('nrecu')
    print('>>>> ids, bank, nrecu', ids, bank, nrecu)
    regl=PaymentSupplier.objects.filter(pk__in=ids)
    print('>>>> regl', regl)
    bank=Bank.objects.get(pk=bank)
    # 7iyd total 4lbanka
    totalamount=regl.aggregate(Sum('amount'))['amount__sum']
    bank.total-=totalamount
    bank.save()
    regl.update(ispaid=True)
    regl.update(dateregl=dateregl)
    regl.update(nrecu=nrecu)
    regl.update(targetbank=bank)
    # regl.update(targebank=Bank.objects.get)
    return JsonResponse({
        'success':True
    })

def payreglfc(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientfc.objects.get(pk=reglid)
    print(reglid, regl)
    regl.ispaid=True
    regl.save()
    return JsonResponse({
        'success':True
    })


def boncommandes(request):
    length=Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat').count(),
    return JsonResponse({
        'length':length
    })


def listeconnected(request):
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    # res=req.get('http://serverip/products/listeconnected')
    # print(json.loads(res.text)['connected'])
    # print('>>', res.text)
    notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago).order_by('-lasttime')
    connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)

    ctx={
        'title':'List utilisateurs Active',
        'connected':connected,
        'active':notconnected,
        'connectedserver':json.loads(res.text)['connected'],
        #'lastactiveserver':json.loads(res.text)['active']
    }
    return render(request, 'listconnected.html', ctx)


def promotionspage(request):
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return render(request, 'promotions.html', ctx)

def createpromotion(request):
    name=request.POST.get('promotionname')
    # get image file
    image=request.FILES.get('promotionimage')
    # create category
    pr=Promotion.objects.create(info=name, image=image)
    # req.get('http://serverip/products/createpromotion', {
    #     'name':name,
    #     # get image file
    #     'image':pr.image.url.replace('/media/', '') if pr.image else ''
    # })
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return JsonResponse({
        'html':render(request, 'promotions.html', ctx).content.decode('utf-8')
    })

def updatepromotion(request):
    image=request.FILES.get('image') or None
    id=request.POST.get('id')
    promotion=Promotion.objects.get(pk=id)
    promotion.info=request.POST.get('name')
    if image:
        promotion.image=image
    promotion.save()
    # req.get('http://serverip/products/updatepromotion', {
    #     'name':request.POST.get('name'),
    #     'id':id,
    #     # get image file
    #     'image':promotion.image.url.replace('/media/', '') if promotion.image else ''
    # })
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return JsonResponse({
        'html':render(request, 'promotions.html', ctx).content.decode('utf-8')
    })


def searchproductsforstock(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    term = request.GET.get('term').lower()

    # Remove non-alphanumeric characters and convert to lowercase

    if not '+' in term:

        # check if term in product.ref or product.name
        products=Produit.objects.filter(ref__istartswith=term)

        q_objects = Q()
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(refeq1__icontains=term) |
            Q(refeq2__icontains=term) |
            Q(refeq3__icontains=term) |
            Q(refeq4__icontains=term)
        )
            # adding other products that have equivalent
        products=products | Produit.objects.filter(q_objects)
    else:
        # Split the cleaned term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:

                # term = ''.join(char for char in term if char.isalnum())
                q_objects &= (Q(ref__icontains=term) | Q(coderef__icontains=term) | Q(name__icontains=term) | Q(category__name__icontains=term) |  Q(mark__name__icontains=term) |  Q(equivalent__icontains=term)  |  Q(refeq1__icontains=term) |  Q(refeq2__icontains=term)  |  Q(block__icontains=term) | Q(refeq3__icontains=term) | Q(refeq4__icontains=term) | Q(sellprice__icontains=term)  | Q(buyprice__icontains=term)  | Q(cars__icontains=term)  | Q(diametre__icontains=term))
            products=Produit.objects.filter(q_objects)[:50]

    return JsonResponse({
        'trs':render(request, 'stocktrs.html', {
            'products':products,
            'target':target,
            'isfarah':target=='f'
        }).content.decode('utf-8'),
        #'stocktotal':Produit.objects.filter(q_objects).aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        #'stockfacture':Produit.objects.filter(q_objects).aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
    })
# loading stock by 50, 50records at a time
def loadstock(request):
    target=request.GET.get('target')
    page = int(request.GET.get('page', 1))
    term = request.GET.get('term')
    notactive = request.GET.get('notactive')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term=='0':
        print('>>>>>>>>>>>>', term=='0')
        if notactive=='1':
            print('from notactive')
            products = Produit.objects.filter(isactive=False)[start:end]
            # trs=''
            # for i in products:
            #     trs+=f"""
            #         <tr ondblclick="ajaxpage('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
            #             style="background:{'#f3d6d694;' if not i.isactive else '' }"
            #                 data-product-id="{ i.id }" class="product-row notactive">
            #                 <td style="padding: 5px; font-weight: bold;" class="pe-2">
            #                     {i.ref.upper()}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.name}
            #                 </td>

            #                 <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
            #                     {i.buyprice if i.buyprice else 0}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
            #                     {i.sellprice if i.sellprice else 0}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;" class="text-center">
            #                     {i.remise}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;" class="text-center">
            #                     {i.prixnet}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
            #                     {i.stocktotal}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
            #                     <span class="stockfacture invisible">{i.stockfacture}</span>
            #                 </td>

            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.diametre}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;" class="text-success">
            #                     {i.block}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.coderef}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.getequivalent()[0] if i.getequivalent() else ''}
            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.getequivalent()[1] if i.getequivalent() else ''}

            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.getequivalent()[2] if i.getequivalent() else ''}

            #                 </td>
            #                 <td style="padding: 5px; font-weight: bold;">
            #                     {i.mark}
            #                 </td>
            #                  <td style="padding: 5px; font-weight: bold;"  class="text-danger">
            #                     {i.code}
            #                 </td>

            #             </tr>
            #     """

            return JsonResponse({
                'trs':render(request, 'stocktrs.html', {'products':products, 'target':target}).content.decode('utf-8'),
                'has_more': len(products) == per_page,
                'target':target
            })
        products = Produit.objects.all()[start:end]
        trs=''
        # for i in products:
        #     trs+=f'''
        #     <tr ondblclick="ajaxpage('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
        #         style="background:{'#f3d6d694;' if not i.isactive else '' }"
        #             data-product-id="{ i.id }" class="product-row">
        #             <td style="padding: 5px; font-weight: bold;" class="pe-2">
        #                 {i.ref.upper()}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;">
        #                 {i.name}
        #             </td>

        #             <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
        #                 {i.buyprice if i.buyprice else 0}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
        #                 {i.sellprice}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-center">
        #                 {i.remise}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-center">
        #                 {i.prixnet}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
        #                 {i.stocktotal}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
        #                 <span class="stockfacture invisible">{i.stockfacture}</span>
        #             </td>

        #             <td style="padding: 5px; font-weight: bold;">
        #                 {i.diametre}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-success">
        #                 {i.block}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;">
        #                 {i.coderef}
        #             </td>

        #           <td style="padding: 5px; font-weight: bold;">
        #               {i.getequivalent()[0] if i.getequivalent() else ''}
        #           </td>
        #           <td style="padding: 5px; font-weight: bold;">
        #               {i.getequivalent()[1] if i.getequivalent() else ''}

        #           </td>
        #           <td style="padding: 5px; font-weight: bold;">
        #               {i.getequivalent()[2] if i.getequivalent() else ''}

        #           </td>

        #             <td style="padding: 5px; font-weight: bold;">
        #                 {i.mark}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;"  class="text-danger">
        #                 {i.code}
        #             </td>
        #             <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
        #         </tr>
        #     '''
        return JsonResponse({
            'trs':render(request, 'stocktrs.html', {'products':products, 'target':target, 'isfarah':target=="f"}).content.decode('utf-8'),
            'has_more': len(products) == per_page,
            'target':target
        })
    else:
        print('>>>>>>>>>>>>', term=='0')


        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:

                q_objects &= (Q(ref__iregex=term) | Q(name__iregex=term) | Q(category__name__iregex=term) |  Q(mark__name__iregex=term))
        products=Produit.objects.filter(q_objects)[start:end]
        return JsonResponse({
            'trs':render(request, 'stocktrs.html', {'products':products, 'isfarah':target=='f'}).content.decode('utf-8'),
            'has_more': len(products) == per_page,
            
        })

def loadlistachat(request):
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>term', term)
        # Split the term into individual words separated by '*'

        print('>>>>>>>>> here bl')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:


            q_objects &= (Q(supplier__name__iregex=term) |
                Q(nbon__iregex=term) |
                Q(total__iregex=term)
                )
        print(startdate, enddate)
        if startdate=='0' and enddate=='0':
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__year=thisyear).order_by('-date')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-date')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Itemsbysupplier.objects.filter(date__range=[startdate, enddate]).order_by('-date')[start:end]
        total=round(Itemsbysupplier.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
    if year=="0":
        bons= Itemsbysupplier.objects.filter(date__year=thisyear).order_by('-date')[start:end]
        total=round(Itemsbysupplier.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons= Itemsbysupplier.objects.filter(date__year=year).order_by('-date')[start:end]
        total=round(Itemsbysupplier.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    for order in bons:
        trs+=f'''
        <tr class="ord achat-row" orderid="{order.id}" ondblclick="ajaxpage('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>

              <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>



            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def loadlistbl(request):
    page = int(request.GET.get('page', 1))
    target =request.GET.get('target')
    isfarah=target=='f'
    isvalid = not request.GET.get('mode')=='1'
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '':
        print('>>term', term)
        # Split the term into individual words separated by '*'

        print('>>>>>>>>> here bl')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:

            # if '-' in term:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (Q(client__name__iregex=term) |
            #         
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #     )
            # else:
            #     q_objects &= (Q(client__name__iregex=term) |
            #         
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (Q(client__name__iregex=term) |
                    
                Q(bon_no__iregex=term) |
                Q(client__region__iregex=term)|
                Q(client__city__iregex=term)|
                Q(client__code__iregex=term)|
                Q(total__iregex=term)|
                
                Q(statusreg__iregex=term)|
                Q(note__iregex=term)
                )
        print(startdate, enddate)
        if startdate=='0' and enddate=='0':
            bons=Bonlivraison.objects.filter(isfarah=isfarah, isvalid=isvalid).filter(q_objects).order_by('-bon_no')[start:end]
            
            #total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear, isfarah=isfarah).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonlivraison.objects.filter(isvalid=isvalid, date__range=[startdate, enddate], isfarah=isfarah).filter(q_objects).order_by('-bon_no')[start:end]
            
            
            # bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah).order_by('-bon_no')[start:end]
            # total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        # for i in bons:
            # trs+=f'''
            # <tr
            # style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
            # class="ord {"text-danger" if i.ispaid else ''} bl-row"
            # year={year}
            # orderid="{i.id}"
            # ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            # term="{term}">
            #     <td>{ i.bon_no }</td>
            #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
            #         <td>{ i.client.name }</td>
            #         <td>{ i.client.code }</td>
            #         <td style="color: blue;">{ i.total}</td>
            #         <td>{ i.client.region}</td>
            #         <td>{ i.client.city}</td>
            #         <td>{ "%.2f" % i.client.soldbl} </td>
            #         <td>{ i.salseman }</td>
            #         <td class="d-flex justify-content-between">
            #         <div>
            #         {'R0' if i.ispaid else 'N1' }

            #         </div>
            #         <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            #         </td>
            #         <td>
            #         {'OUI' if i.isfacture else 'NON'}
            #         </td>

            #         <td>
            #             {i.commande.order_no if i.commande else "---"}
            #         </td>
            #         <td>
            #         {i.modlvrsn}
            #         </td>
            #         <td>
            #         {i.note}
            #         </td>
            #         <td class="d-flex">
            #       <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
            #     <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
            #         </td>

            #   </tr>
            # '''
        print('>>>load bl term')
        return JsonResponse({
            'trs':render(request, 'bllist.html', {'bons':bons, 'target':target, 'isfarah':isfarah}).content.decode('utf-8'),
            'has_more': len(bons) == per_page
        })
    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Bonlivraison.objects.filter(isvalid=isvalid, date__range=[startdate, enddate], isfarah=isfarah).order_by('-bon_no')[start:end]
        
        total=round(Bonlivraison.objects.filter(date__range=[startdate, enddate], isfarah=isfarah).aggregate(Sum('total'))['total__sum'] or 0, 2)
        print('>>>load bl date f')
        return JsonResponse({
            'trs':render(request, 'bllist.html', {'bons':bons, 'target':target, 'isfarah':isfarah}).content.decode('utf-8'),
            'has_more': len(bons) == per_page
        })
    bons= Bonlivraison.objects.filter(isvalid=isvalid, isfarah=isfarah).order_by('-bon_no')[start:end]
    
    total=round(Bonlivraison.objects.filter(isfarah=isfarah).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    # for i in bons:
    #     trs+=f'''
    #     <tr style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} " class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
    #         <td>{ i.bon_no }</td>
    #             <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #             <td>{ i.client.name }</td>
    #             <td>{ i.client.code }</td>
    #             <td style="color: blue;">{ i.total}</td>
    #             <td>{ i.client.region}</td>
    #             <td>{ i.client.city}</td>
    #             <td>{ "%.2f" % i.client.soldbl}</td>
    #             <td>{ i.salseman }</td>
    #             <td class="d-flex justify-content-between">
    #             <div>
    #             {'R0' if i.ispaid else 'N1' }

    #             </div>
    #             <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #             </td>
    #             <td>
    #             {'OUI' if i.isfacture else 'NON'}

    #             </td>

    #             <td>

    #             </td>
    #             <td>
    #             {i.modlvrsn}
    #             </td>
    #             <td>
    #             {i.note}
    #             </td>
    #             <td class="d-flex">
    #               <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
    #             <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>

    #             </td>
    #       </tr>
    #     '''
    
    return JsonResponse({
        'trs':render(request, 'bllist.html', {'bons':bons, 'target':target, 'isfarah':isfarah}).content.decode('utf-8'),
        'has_more': len(bons) == per_page
    })

def loadlistbs(request):
    page = int(request.GET.get('page', 1))
    # if search type changed to valid, means the button valid clicked and wz wnt only valid bons
    searchtype =request.GET.get('searchtype')
    isgenerated=searchtype=='valid'
    all=request.GET.get('all')=='1'
    valid=request.GET.get('valid')=='1'
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>term', term)
        # Split the term into individual words separated by '*'

        print('>>>>>>>>> here bl')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:

            # if '-' in term:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (Q(client__name__iregex=term) |
            #         
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #     )
            # else:
            #     q_objects &= (Q(client__name__iregex=term) |
            #         
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (Q(client__name__iregex=term) |
                    
                Q(bon_no__iregex=term) |
                Q(client__region__iregex=term)|
                Q(client__city__iregex=term)|
                Q(client__code__iregex=term)|
                Q(total__iregex=term)|
                
                Q(statusreg__iregex=term)|
                Q(note__iregex=term)
                )
        print(startdate, enddate)
        if startdate=='0' and enddate=='0':
            bons=Bonsortie.objects.filter(q_objects).order_by('-bon_no')[start:end]
            #total=round(Bonsortie.objects.filter(q_objects).filter(date__year=thisyear).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonsortie.objects.filter(date__range=[startdate, enddate]).filter(q_objects).order_by('-bon_no')[start:end]
            # bons=Bonsortie.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no')[start:end]
            # total=round(Bonsortie.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        # for i in bons:
            # trs+=f'''
            # <tr
            # style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
            # class="ord {"text-danger" if i.ispaid else ''} bl-row"
            # year={year}
            # orderid="{i.id}"
            # ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/Bonsortiedetails/{i.id}')"
            # term="{term}">
            #     <td>{ i.bon_no }</td>
            #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
            #         <td>{ i.client.name }</td>
            #         <td>{ i.client.code }</td>
            #         <td style="color: blue;">{ i.total}</td>
            #         <td>{ i.client.region}</td>
            #         <td>{ i.client.city}</td>
            #         <td>{ "%.2f" % i.client.soldbl} </td>
            #         <td>{ i.salseman }</td>
            #         <td class="d-flex justify-content-between">
            #         <div>
            #         {'R0' if i.ispaid else 'N1' }

            #         </div>
            #         <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            #         </td>
            #         <td>
            #         {'OUI' if i.isfacture else 'NON'}
            #         </td>

            #         <td>
            #             {i.commande.order_no if i.commande else "---"}
            #         </td>
            #         <td>
            #         {i.modlvrsn}
            #         </td>
            #         <td>
            #         {i.note}
            #         </td>
            #         <td class="d-flex">
            #       <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
            #     <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
            #         </td>

            #   </tr>
            # '''
        print('>>>load bl term')
        return JsonResponse({
            'trs':render(request, 'bslist.html', {'bons':bons}).content.decode('utf-8'),
            'has_more': len(bons) == per_page
        })
    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Bonsortie.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[start:end]
        total=round(Bonsortie.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
        print('>>>load bl date f')
        return JsonResponse({
            'trs':render(request, 'bslist.html', {'bons':bons}).content.decode('utf-8'),
            'has_more': len(bons) == per_page
        })
    bons= Bonsortie.objects.order_by('-bon_no')[start:end]
    total=round(Bonsortie.objects.order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    return JsonResponse({
        'trs':render(request, 'bslist.html', {'bons':bons}).content.decode('utf-8'),
        'has_more': len(bons) == per_page
    })




def loadlistbc(request):
    # each block needs a return statement
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    print(year, startdate, enddate, term)
    per_page = 50  # Adjust as needed
    trs=''

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>>>> in term')
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term) |  Q(order_no__iregex=term) | Q(total__iregex=term))
        if startdate=='0' and enddate=='0':
            print("in  term  no filter data")
            bons=Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-order_no')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

        else:
            print('in term and filterdate')
            bons=Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

        return JsonResponse({
            'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total
        })
    if startdate != '0' and enddate != '0':
        print('>>>>> in date fil')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')+timedelta(1)
        bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        return JsonResponse({
            'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total
        })

    if year=="0":
        print('>>>>> in year list bc is null')

        bons= Order.objects.filter(date__year=thisyear).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__year=thisyear).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        print('>>>>> in year')

        bons= Order.objects.filter(date__year=year).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__year=year).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)



    return JsonResponse({
        'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
        'has_more': len(bons) == per_page,
        'total':total
    })


def searchforlistachat(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    year=request.GET.get('year')
    isfarah=target=="f"
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    isvalid=request.GET.get('wanted')=='valid'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:


        q_objects &= (Q(supplier__name__iregex=term) |
            Q(nbon__iregex=term) |
            Q(total__iregex=term)
        )
    print(term, startdate, enddate)
    if startdate=='0' and enddate=='0':
        bons=Itemsbysupplier.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=isvalid).order_by('-date')
        total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=isvalid).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid).order_by('-date')
        total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=isvalid).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
    print('>> bons', bons)
    # trs=''
    # for order in bons:
    #     trs+=f'''
    #         <tr class="ord " orderid="{order.id}" ondblclick="ajaxpage('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
    #         <td>{ order.nbon }</td>
    #         <td>{ order.date.strftime("%d/%m/%Y") }</td>
    #         <td>{ order.supplier.name }</td>
    #         <td>{ order.total}</td>
    #         <td>{ order.tva}</td>
    #         <td>{"Facture"if order.isfacture else "Bl"}</td>
    #         <td>{ round(order.supplier.rest, 2) }</td>
    #         <td class="d-flex">
    #             <div>{"R0"if order.ispaid else "N1"}</div>
    #             <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>
    #         </td>
    #         <td>
    #           <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
    #         </td>

    #       </tr>
    #         '''

    return JsonResponse({
        'trs':render(request, 'bonachattrs.html', {'bons':bons, 'target':target}).content.decode('utf-8'),
        'total':total
    })

def searchforlistbl(request):
    term=request.GET.get('term')
    searchtype=request.GET.get('searchtype')
    print(">> searchtype: ", searchtype)
    target=request.GET.get('target')
    isfarah=target=='f'
    year=request.GET.get('year')
    print('>>>', year, target, searchtype)
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:
        q_objects &= (Q(client__name__iregex=i) |
                Q(bon_no__iregex=i) |
                Q(client__region__iregex=i)|
                Q(client__city__iregex=i)|
                Q(client__code__iregex=i)|
                Q(bonsortie__bon_no__iregex=i)|
                Q(facture__facture_no__iregex=i)|
                Q(statusreg__iregex=i)|
                Q(note__iregex=i)
            )
    print(">> here 1",startdate, enddate)
    if startdate=='0' and enddate=='0':
        print('>> hhere1')
        if searchtype=='waiting':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=False).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=False).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        elif searchtype=='valid':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=True).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, isvalid=True).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, iscanceled=True).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfarah=isfarah, iscanceled=True).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        if searchtype=='waiting':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        elif searchtype=='valid':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, iscanceled=True).order_by('-bon_no')[:50]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, iscanceled=True).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    return JsonResponse({
        'trs':render(request, 'bllist.html', {'bons':bons, 'notloading':True, 'target':target}).content.decode('utf-8'),
        'total':total
    })

def searchforlistbs(request):
    term=request.GET.get('term')
    searchtype=request.GET.get('searchtype')
    # if searchtype == valid, means we want only the generated bs
    isgenerated=searchtype=='valid'
    target=request.GET.get('target')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    # we dont need this
    # if(term==''):

    #     bons=Bonsortie.objects.filter(date__year=thisyear)[:50]
    #     total=round(Bonsortie.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     trs=''
    #     for i in bons:
    #         trs+=f'''
    #         <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/Bonsortiedetails/{i.id}')">
    #             <td>{ i.bon_no }</td>
    #             <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #             <td>{ i.client.name }</td>
    #             <td>{ i.client.code }</td>
    #             <td>{ i.total}</td>
    #             <td>{ i.client.region}</td>
    #             <td>{ i.client.city}</td>
    #             <td>{ i.client.soldbl}</td>
    #             <td>{ i.salseman }</td>
    #             <td class="d-flex justify-content-between">
    #             <div>
    #             {'R0' if i.ispaid else 'N1' }

    #             </div>
    #             <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #             </td>
    #             <td class="text-danger">

    #             </td>
    #             <td class="text-danger">
    #             {'OUI' if i.isfacture else 'NON'}

    #             </td>

    #             <td>


    #             </td>
    #             <td>
    #             {i.note}
    #             </td>
    #             <td>
    #             {i.modlvrsn}
    #             </td>
    #             <td class="d-flex">
    #               <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
    #             <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
    #             </td>
    #         </tr>
    #         '''
    #     return JsonResponse({
    #         'trs':trs
    #     })

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:
        q_objects &= (Q(client__name__iregex=i) |
                Q(bon_no__iregex=i) |
                Q(client__region__iregex=i)|
                Q(client__city__iregex=i)|
                Q(client__code__iregex=i)|
                Q(total__iregex=i)|
                Q(note__iregex=i)|
                Q(car__iregex=i)
                
            )
    print(">> here 1",startdate, enddate)
    if startdate=='0' and enddate=='0':
        bons=Bonsortie.objects.filter(q_objects).order_by('-bon_no')[:50]
        # generated=isgenerated
        total=round(Bonsortie.objects.filter(q_objects).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        # generated=isgenerated
    else:
        print(">> here 1²",startdate, enddate)
        bons=Bonsortie.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no')[:50]
        # generated=isgenerated
        total=round(Bonsortie.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        # generated=isgenerated
    return JsonResponse({
        'trs':render(request, 'bslist.html', {'bons':bons, 'notloading':True}).content.decode('utf-8'),
        'total':total
    })



def searchforlistbc(request):
    term=request.GET.get('term')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    # we dont need this
    if(term==''):

        bons=Order.objects.filter(date__year=thisyear)[:50]
        total=round(Order.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" orderid="{i.id}" ondblclick="ajaxpage('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')" term={term}>
                <td>{ i.order_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>
                <td class="text-danger">
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>


                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:
        q_objects &= (Q(client__name__iregex=i) |
                Q(salseman__name__iregex=i) |
                Q(order_no__iregex=i) |
                Q(client__region__iregex=i)|
                Q(client__city__iregex=i)|
                Q(client__code__iregex=i)|
                Q(total__iregex=i)|
                Q(note__iregex=i)

            )
    print(startdate, enddate)
    if startdate=='0' and enddate=='0':
        bons=Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-order_no')[:50]
        total=round(Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no')[:50]
        total=round(Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)



    return JsonResponse({
        'trs':render(request, 'bclist.html', {'bons':bons,
        'loadmore':True,
        'startdate':startdate, 'enddate':startdate, 'term':term}).content.decode('utf-8'),
        'total':total
    })

def loadlistfc(request):
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    target =request.GET.get('target')
    isfarah=target=='f'
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    comptable =request.GET.get('comptable')
    # request.GET.get('mode')=='1' means waiting so is valid will be not
    isvalid = not request.GET.get('mode')=='1'
    print('>> isvalid, isfarah', isvalid, isfarah)
    per_page = 50  # Adjust as needed
    print('>>>>> term', term)
    start = (page - 1) * per_page
    end = page * per_page
    print('>>>>>', start, end, page)
    if term != '':

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in term.split():
            # print('>>>>>>>term ', term)
            # if '-' in term[0]:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #         )
            # else:
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (
                Q(client__name__iregex=term)|
                Q(client__city__iregex=term)|
                Q(salseman__name__iregex=term)|
                Q(facture_no__iregex=term)|
                Q(bon__bon_no__iregex=term)|
                Q(client__region__iregex=term)|
                Q(client__code__iregex=term)|
                Q(note__iregex=term)|
                Q(statusreg__iregex=term)|
                Q(total__iregex=term)
            )

        if startdate=='0' and enddate=='0':

            bons=Facture.objects.filter(q_objects).filter(isvalid=isvalid, isfarah=isfarah, date__year=year).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(isvalid=isvalid, date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(isvalid=isvalid, date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        else:
            print('>>>>>daterange ')
            bons=Facture.objects.filter(q_objects).filter(isvalid=isvalid, isfarah=isfarah, date__range=[startdate, enddate]).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(isvalid=isvalid, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(isvalid=isvalid, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        # trs=''
        # for i in bons:
        #     trs+=f'''
        #     <tr class="ord {"text-danger" if i.ispaid else ''}
        #      fc-row"
        #         style="color:{"blue" if i.bon else ""} "
        #       year={year} term={term} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
        #         <td>{ i.facture_no }</td>
        #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
        #         <td>{ i.total}</td>
        #         <td>{ i.tva}</td>
        #         <td>{ i.client.name }</td>
        #         <td>{ i.client.code }</td>
        #         <td>{ i.client.region}</td>
        #         <td>{ i.client.city}</td>
        #         <td>{ i.client.soldfacture}</td>
        #         <td>{ i.salseman }</td>
        #         <td class="d-flex justify-content-between">
        #         <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
        #         <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
        #         </td>
        #         <td >
        #             {i.note}
        #         </td>

        #         <td>
        #         {i.bon.bon_no if i.bon else "--"}
        #         </td>
        #         <td class="d-flex">
        #             <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
        #             <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
        #         </td>
        #     </tr>
        #     '''
        
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })
    if startdate != '0' and enddate != '0':
        print('>>>>>>>>>>in start end dat')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        print(startdate, enddate)
        bons=Facture.objects.filter(isvalid=isvalid, isfarah=isfarah, date__range=[startdate, enddate]).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(isvalid=isvalid, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(isvalid=isvalid, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        # trs=''
        # for i in bons:
        #     trs+=f'''
        #     <tr class="ord {"text-danger" if i.ispaid else ''}
        #      fc-row"
        #         style="color:{"blue" if i.bon else ""} "
        #       startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
        #         <td>{ i.facture_no }</td>
        #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
        #         <td>{ i.total}</td>
        #         <td>{ i.tva}</td>
        #         <td>{ i.client.name }</td>
        #         <td>{ i.client.code }</td>
        #         <td>{ i.client.region}</td>
        #         <td>{ i.client.city}</td>
        #         <td>{ i.client.soldfacture:.2f}</td>
        #         <td>{ i.salseman }</td>
        #         <td class="d-flex justify-content-between">
        #         <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
        #         <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
        #         </td>
        #         <td>
        #             {i.note}
        #         </td>

        #         <td>
        #         {i.bon.bon_no if i.bon else "--"}
        #         </td>
        #         <td class="d-flex">
        #             <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
        #             <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
        #         </td>
        #     </tr>
        #     '''
        
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })
    print('>> we are her', year)
    bons= Facture.objects.filter(isvalid=isvalid, isfarah=isfarah, date__year=year).order_by('-facture_no')[start:end]
    total=round(Facture.objects.filter(isvalid=isvalid, isfarah=isfarah, date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(Facture.objects.filter(isvalid=isvalid, isfarah=isfarah, date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    # trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
    #         <td>{ i.facture_no }</td>
    #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.tva}</td>
    #         <td>{ i.client.name }</td>
    #         <td>{ i.client.code }</td>
    #         <td>{ i.client.region}</td>
    #         <td>{ i.client.city}</td>
    #         <td>{ i.client.soldfacture:.2f}</td>
    #         <td>{ i.salseman }</td>
    #         <td class="d-flex justify-content-between">
    #         <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
    #         <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
    #         </td>
    #         <td>
    #             {i.note}
    #         </td>

    #         <td>
    #         {i.bon.bon_no if i.bon else "--"}
    #         </td>
    #         <td class="d-flex">
    #             <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
    #             <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
    #         </td>
    #     </tr>
    #     '''
    
    return JsonResponse({
        'trs':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'has_more': len(bons) == per_page,
        'total':total,
        'totaltva':totaltva,
    })


def searchforlistfc(request):
    term=request.GET.get('term')
    waiting=request.GET.get('waiting')== '1'
    target=request.GET.get('target')
    searchedterm=request.GET.get('term')
    startdate=request.GET.get('startdate') or '0'
    enddate=request.GET.get('enddate') or '0'
    print('>>', startdate, enddate)
    year=request.GET.get('year')
    isfarah=target=='f'

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        print('>>>>>> term', term)
        q_objects &= (
            Q(client__name__iregex=term)|
            Q(client__city__iregex=term)|
            Q(salseman__name__iregex=term)|
            Q(facture_no__iregex=term)|
            Q(bon__bon_no__iregex=term)|
            Q(client__region__iregex=term)|
            Q(client__code__iregex=term)|
            Q(note__iregex=term)|
            Q(statusreg__iregex=term)|
            Q(total__iregex=term)
        )

    if startdate=='0' and enddate=='0':
        print('>>>> search list fc, startdate and enddate are 0')
        if waiting:
            bons=Facture.objects.filter(q_objects).filter(isfarah=isfarah, isvalid=False).order_by('-facture_no')[:50]
            total=round(Facture.objects.filter(q_objects).filter(isfarah=isfarah, isvalid=False).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Facture.objects.filter(q_objects).filter(isfarah=isfarah, isvalid=True).order_by('-facture_no')[:50]
            total=round(Facture.objects.filter(q_objects).filter(isfarah=isfarah, isvalid=True).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    else:
        if waiting:
            bons=Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False).order_by('-facture_no')[:50]
            total=round(Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=False).aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).order_by('-facture_no')[:50]
            total=round(Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfarah=isfarah, isvalid=True).aggregate(Sum('total'))['total__sum'] or 0, 2)

    # if year=='0':
    #     bons=Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # else:
    #     bons=Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)

    return JsonResponse({
        'trs':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':total,

    })

def searchforlistfcachat(request):
    term=request.GET.get('term')
    waiting=request.GET.get('waiting')== '1'
    target=request.GET.get('target')
    searchedterm=request.GET.get('term')
    startdate=request.GET.get('startdate') or '0'
    enddate=request.GET.get('enddate') or '0'
    year=request.GET.get('year')
    isfarah=target=='f'
    isvalid=request.GET.get('wanted')=='valid'
    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        print('>>>>>> term', term)
        q_objects &= (
            Q(supplier__name__iregex=term)|
            Q(supplier__city__iregex=term)|
            Q(facture_no__iregex=term)|
            Q(bon__nbon__iregex=term)|
            Q(supplier__code__iregex=term)|
            Q(total__iregex=term)
        )
    print('>>>> search list fc, startdate and enddate are 0')
    bons=Factureachat.objects.filter(q_objects).filter(date__year=thisyear, isfarah=isfarah, isvalid=False).order_by('-facture_no')[:50]
    total=round(Factureachat.objects.filter(q_objects).filter(date__year=thisyear, isfarah=isfarah, isvalid=False).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        
    # if year=='0':
    #     bons=Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # else:
    #     bons=Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)

    return JsonResponse({
        'trs':render(request, 'fcachatlist.html', {'bons':bons}).content.decode('utf-8'),
        'total':total,

    })



def createnewclientaccount(request):
    clientid=request.POST.get('clientid')
    client= Client.objects.get(pk=clientid)
    olduser=client.user
    username=request.POST.get('username').strip()
    password=request.POST.get('password')
    #check for username
    user=User.objects.filter(username=username).first()
    print('>>>>>>>>>>>', username, client.code)
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })
    user=User.objects.create_user(username=username, password=password)
    try:
        # response=req.get('http://serverip/products/createnewclientaccount', {
        #     'username':username,
        #     'password':password,
        #     'clientcode':client.code
        # })
        # response.raise_for_status()
        cart=Cart.objects.filter(user=olduser).first()
        wich=Wich.objects.filter(user=olduser).first()
        if cart:
            cart.user=user
            cart.save()
        if wich:
            wich.user=user
            wich.save()
        olduser.delete()
        # create user
        # assign user to client
        group, created = Group.objects.get_or_create(name="clients")
        user.groups.add(group)
        user.save()
        client.user=user
        client.save()

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print('>>>>>>>', e)
        return JsonResponse({
            'success':False,
            'error':'ERROR SERVER'
        })


def createnewrepaccount(request):
    repid=request.POST.get('repid')
    rep= Represent.objects.get(pk=repid)
    olduser=rep.user
    username=request.POST.get('username')
    password=request.POST.get('password')
    user=User.objects.filter(username=username).first()
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })
    try:
        # response=req.get('http://serverip/products/createnewrepaccount', {
        #     'username':username,
        #     'password':password,
        #     'repid':repid
        # })
        # response.raise_for_status()
        user=User.objects.create_user(username=username, password=password)
        # assign user to rep
        cart=Cart.objects.filter(user=olduser).first()
        wich=Wich.objects.filter(user=olduser).first()
        if cart:
            cart.user=user
            cart.save()
        if wich:
            wich.user=user
            wich.save()
        olduser.delete()
        group, created = Group.objects.get_or_create(name="salsemen")
        user.groups.add(group)
        user.save()
        rep.user=user
        rep.save()
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print('>>>>>>>', e)
        return JsonResponse({
            'success':False,
            'error':'ERROR SERVER'
        })



def yeardatabl(request):
    year=request.GET.get('year')
    target=request.GET.get('target')
    # get all bls of that year
    bls=Bonlivraison.objects.filter(date__year=year, isfarah=target=='f').order_by('-id')[:50]
    # trs=''
    # for i in bls:
    #     trs+=f'''
    #     <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
    #         <td>{ i.bon_no }</td>
    #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #         <td>{ i.client.name }</td>
    #         <td>{ i.client.code }</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.client.region}</td>
    #         <td>{ i.client.city}</td>
    #         <td>{ i.client.soldbl}</td>
    #         <td>{ i.salseman }</td>
    #         <td class="d-flex justify-content-between">
    #           <div>
    #           {'R0' if i.ispaid else 'N1' }

    #           </div>
    #           <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #         </td>
    #         <td class="text-danger">
    #         {'OUI' if i.isfacture else 'NON'}

    #         </td>

    #         <td>
    #           {i.commande.order_no if i.commande else '--'}
    #         </td>
    #         <td>
    #           {i.modlvrsn}
    #         </td>
    #       </tr>
    #     '''
    
    return JsonResponse({
        'trs':render(request, 'bllist.html', {'bons':bls}).content.decode('utf-8'),
        'total':round(Bonlivraison.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabs(request):
    year=request.GET.get('year')
    target=request.GET.get('target')
    # get all bls of that year
    bls=Bonsortie.objects.filter(date__year=year).order_by('-id')[:50]
    print('>> bons ', bls)
    return JsonResponse({
        'trs':render(request, 'bslist.html', {'bons':bls}).content.decode('utf-8'),
        'total':round(Bonlivraison.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabachat(request):
    this_year=datetime.now().year
    year=request.GET.get('year')
    target=request.GET.get('target')
    print('year>>>', year) 
    # get all bls of that year
    bons=Itemsbysupplier.objects.filter(date__year=year, isfarah=target=='f').order_by('-id')[:50]
    print('>> bons in', year, bons)
    return JsonResponse({
        'trs':render(request, 'bonachatlist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(Itemsbysupplier.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabc(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bons=Order.objects.filter(date__year=year).order_by('-id')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} orderid="{i.code}" ondblclick="ajaxpage('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
            <td>{ i.order_no }</td>
            <td>{ i.date.strftime('%d/%m/%Y') }</td>
            <td>{ i.client.name if i.client else '' }</td>
            <td>{ i.client.code if i.client else '' }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region if i.client else '' }</td>
            <td>{ i.client.city if i.client else '' }</td>
            <td>{ i.client.soldbl if i.client else '' }</td>
            <td>{ i.salseman }</td>
            <td>
            {"Non" if i.isclientcommnd else "OUI"}
            </td>
            <td>
              {"R1" if i.isdelivered else  "R0" }
            </td>

          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(Order.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def yeardatafc(request):
    year=request.GET.get('year')
    target=request.GET.get('target')
    print(year)
    # get all bls of that year
    bls=Facture.objects.filter(date__year=year, isfarah=target=='f').order_by('-facture_no')[:50]
    # trs=''
    # for i in bls:
    #     trs+=f'''
    #     <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
    #         <td>{ i.facture_no }</td>
    #         <td>{ i.date.strftime("%d/%m/%Y")}</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.tva}</td>
    #         <td>{ i.client.name }</td>
    #         <td>{ i.client.code }</td>
    #         <td>{ i.client.region}</td>
    #         <td>{ i.client.city}</td>
    #         <td>{ i.client.soldbl}</td>
    #         <td>{ i.salseman }</td>
    #         <td class="d-flex justify-content-between">
    #           <div>
    #           {'R0' if i.ispaid else 'N1' }

    #           </div>
    #           <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

    #         </td>
    #         <td class="text-danger">

    #         </td>

    #         <td>
    #           {i.bon.bon_no if i.bon else "--"}
    #         </td>
    #       </tr>
    #     '''
    
    return JsonResponse({
        # 'trs':trs,
        'trs':render(request, 'fclist.html', {'bons':bls}).content.decode('utf-8'),
        'total':round(bls.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def loadreglbl(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    print('start', start, end)
    target=request.GET.get('target')
    if target=='f':
        reblbls = PaymentClientbl.objects.filter(isfarah=True, isavoir=False).order_by('-id')[start:end]
    elif target=='o':
        reblbls = PaymentClientbl.objects.filter(isorgh=True, isavoir=False).order_by('-id')[start:end]
    else:
        reblbls = PaymentClientbl.objects.filter(issortie=True, isavoir=False).order_by('-id')[start:end]
    
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })


def laodreglfc(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page

    reblbls = PaymentClientfc.objects.all()[start:end]
    # trs=''
    # for i in reblbls:
    #     tooltip_content = ''.join([bon.bon_no for bon in i.bons.all()])
    #     trs += f'''
    #     <tr style="{ "background: yellow;" if (i.echance and i.echance == today) else ("color: red;" if (i.echance and i.echance < today) else "") }" class="reglbl-row">
    #         <td>
    #             { i.mode } <br>
    #             <!-- <select name="updatemodereglbl{ i.id }" id=""></select> -->
    #         </td>
    #         <td>
    #             { i.npiece } <br>
    #             <input type="text" class="d-none updatenpiecereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.amount }<br>
    #             <!-- <input type="text" class="d-none updateamountreglbl{ i.id }"> -->
    #         </td>
    #         <td>
    #             { i.date }<br>
    #             <input type="date" class="d-none updatetdatereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.client.name }<br>
    #         </td>
    #         <td>
    #             { i.client.code }<br>
    #         </td>
    #         <td>
    #             { i.echance} <br>
    #         </td>
    #         <td class="d-flex justify-content-between">
    #             <button type="button" class="btn btn-secondary btn-sm" data-toggle="tooltip" data-placement="top" data-tooltip="Bon Nos: {tooltip_content}">
    #             </button>
    #             <button class="btn btn-success btn-sm" onclick="updatereglementbl({ i.id })">✐</button>
    #         </td>
    #     </tr>
    # '''

    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })



def loadclients(request):
    page = int(request.GET.get('page', 1))
    target=request.GET.get('target')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if target=='f':
        clients = Client.objects.filter(clientfarah=True).order_by('-soldtotal')[start:end]
    elif target=='o':
        clients = Client.objects.filter(clientorgh=True).order_by('-soldtotal')[start:end]
    else:
        clients = Client.objects.filter(clientsortie=True).order_by('-soldtotal')[start:end]
    return JsonResponse({
        'trs':render(request, 'clienttrs.html', {
            'clients':clients, 'target':target
        }).content.decode('utf-8'),
        'has_more': len(clients) == per_page
    })


def exportproducts(request):
    categoryid=request.GET.get('categoryid')
    isfarah=request.GET.get('target')=='f'
    if categoryid=='0':
        products=Produit.objects.all()
        filename='Produit_tous'+today.strftime('%d/%m/%y')+'.xlsx'
    else:
        category=Category.objects.get(pk=categoryid)
        products=Produit.objects.filter(category__id=categoryid)
        filename='Produit_'+category.name+today.strftime('%d%m%y')+'.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['ref', 'name', 'category', 'buyprice', 'sellprice', 'remise', 'prixnet', 'stocktotal', 'mark', 'diametre', 'block', 'equivalent', 'refeq1', 'refeq2'])

    # Write product data
    for product in products:
        if isfarah:
            stock=product.stocktotalfarah
        else:
            stock=product.stocktotalorgh
        ws.append([
            product.ref, product.name,
            product.category.name if product.category else '',  # Extract category name
            product.buyprice, product.sellprice,
            product.remise1, product.prixnet, stock, product.mark.name if product.mark else '',
            product.diametre, product.block, product.equivalent, product.refeq1, product.refeq2
        ])

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Save the workbook to the response
    wb.save(response)

    return response

def datepdct(request):
    datefrom=request.GET.get('datefrom')
    dateto=request.GET.get('dateto')
    pdctid=request.GET.get('pdctid')
    print('dates>>>>>>', datefrom, dateto)
    datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    dateto=datetime.strptime(dateto, '%Y-%m-%d')
    stockin=Stockin.objects.filter(product_id=pdctid, date__range=[datefrom, dateto]).order_by('date')
    stockout=Livraisonitem.objects.filter(product_id=pdctid, bon__date__range=[datefrom, dateto]).order_by('date')
    totalqtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    totalqtyout=stockout.aggregate(Sum('qty'))['qty__sum'] or 0
    totalcoutin=stockin.aggregate(Sum('total'))['total__sum'] or 0
    totalrevenu=stockout.aggregate(Sum('total'))['total__sum'] or 0
    trin=''
    trout=''
    for i in stockin:
        trin+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.nbon}
                </td>
                <td>
                    {i.supplier.name}
                </td>
                <td>
                    {i.quantity}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    for i in stockout:
        trout+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.bon.bon_no}
                </td>
                <td>
                    {i.bon.client.name}
                </td>
                <td>
                    {i.qty}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.remise}%
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    return JsonResponse({
        'trin':trin,
        'trout':trout,
        'totalqtyin':totalqtyin,
        'totalqtyout':totalqtyout,
        'totalcoutin':totalcoutin,
        'totalrevenu':totalrevenu
    })


def showdeactivated(request):
    products=Produit.objects.filter(isactive=False)[:100]
    trs=''
    for i in products:
        trs+=f"""
            <tr ondblclick="ajaxpage('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row notactive">
                    <td >
                        {i.ref.upper()}
                    </td>
                    <td>
                        {i.name}
                    </td>
                    <td>
                        {i.category}
                    </td>
                    <td class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td class="text-center">
                        {i.sellprice}
                    </td>
                    <td class="text-center">
                        {i.remise}
                    </td>
                    <td class="text-center">
                        {i.prixnet}
                    </td>
                    <td class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td>
                        {i.diametre}
                    </td>
                    <td class="text-success">
                        {i.block}
                    </td>
                    <td>
                        {i.coderef}
                    </td>
                    <td>

                    </td>
                    <td>

                    </td>
                    <td>

                    </td>
                    <td>
                        {i.mark}
                    </td>

                </tr>
        """

    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == 100
    })

def searchforlistclient(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    if(term==''):
        clients=Client.objects.all()[:50]
        return JsonResponse({
            'trs':render(request, 'clienttrs.html', {
            'clients':clients,
            'target':target
            }).content.decode('utf-8')
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:

            # term = ''.join(char for char in term if char.isalnum())
            q_objects &= (Q(city__icontains=term) |
                Q(name__icontains=term) |
                Q(ice__icontains=term) |
                Q(phone__icontains=term) |
                Q(region__icontains=term) |
                Q(code__icontains=term) |
                Q(represent__name__icontains=term) |
                Q(address__icontains=term))
    if target=='f':
        clients=Client.objects.filter(clientfarah=True).filter(q_objects)
    elif target=='o':
        clients=Client.objects.filter(clientorgh=True).filter(q_objects)
    else:
        clients=Client.objects.filter(clientsortie=True).filter(q_objects)
    return JsonResponse({
        'trs':render(request, 'clienttrs.html', {
        'clients':clients,
        'target':target
        }).content.decode('utf-8')
    })


def laodblregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    print('>>>>>>>>>', nr)
    if nr=='0':
        print('all', page)
        bons=Bonlivraison.objects.filter(client_id=clientid).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'
    else:
        print('non regl')
        bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })


def laodblinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientbl.objects.get(pk=reglementid)

    bons=reglement.bons.all()
    # bons without bons in reglement
    bons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def laodfcinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientfc.objects.get(pk=reglementid)

    bons=reglement.factures.all()
    # bons without bons in reglement
    bons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

# the facture load in the add reglement
def laodfcregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if nr=='0':
        bons=Facture.objects.filter(client_id=clientid).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    else:
        bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def searchclientbls(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Use list comprehension for building search terms
    search_terms = [term for term in regex_search_term.split('*') if term]

    # Use __icontains for case-insensitive searches
    q_objects = Q()
    q_objects &= (Q(salseman__name__icontains=term) | Q(bon_no__icontains=term) | Q(total__icontains=term))

    # Combine filter conditions in one line
    bons = Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]

    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })


def searchclientfcs(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= ( Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)" {"checked" if i.reglementsfc.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })

# search fc in update reglement
def searchclientfcsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= ( Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })

def searchclientblsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= ( Q(bon_no__iregex=term) | Q(total__iregex=term))
    bons=Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons, clientid)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })



def searchregl(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__icontains=term) | Q(client__code__icontains=term) | Q(mode__icontains=term) | Q(npiece__icontains=term) | Q(amount__icontains=term))
    if target=='f':
        regls=PaymentClientbl.objects.filter(isfarah=True).filter(q_objects)[:50]
    if target=='o':
        regls=PaymentClientbl.objects.filter(isorgh=True).filter(q_objects)[:50]
    if target=='s':
        regls=PaymentClientbl.objects.filter(issortie=True).filter(q_objects)[:50]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':regls}).content.decode('utf-8'),

    })

def searchreglsupp(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    year=request.GET.get('year')
    
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__icontains=term) |Q(supplier__code__icontains=term) | Q(mode__icontains=term) | Q(npiece__icontains=term) | Q(amount__icontains=term))
    if target=='f':
        regls=PaymentSupplier.objects.filter(isfarah=True).filter(q_objects)[:50]
    if target=='o':
        regls=PaymentSupplier.objects.filter(isfarah=False).filter(q_objects)[:50]
    return JsonResponse({
        'trs':render(request, 'reglsupplist.html', {'reglements':regls}).content.decode('utf-8'),

    })




def searchreglfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    q_objects = Q()
    if term=='':
        if year=='0':
            regls=PaymentClientfc.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        else:
            regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        return JsonResponse({
            'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &

    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term) | Q(client__code__iregex=term) | Q(mode__iregex=term) | Q(npiece__iregex=term) | Q(amount__iregex=term))
    if year=='0':
        regls=PaymentClientfc.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
    else:
        regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),

    })

def admindash(request):
    # if post method
    if request.method=='POST':
        # get user and password
        username=request.POST.get('username')
        password=request.POST.get('password')
        # check if user exist
        user=authenticate(username=username, password=password)
        if user:
            group=user.groups.all().first().name
            if group == 'admin':
                login(request, user)
                return redirect('main:system')
    if request.user.groups.all():
        group=request.user.groups.all().first().name
        if group == 'admin':
            return redirect('main:system')
    return render(request, 'admindash.html')

def yeardatareglfc(request):
    year=request.GET.get('year')
    regls=PaymentClientfc.objects.filter(date__year=year).order_by('-date')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date(), 'year':{year}}).content.decode('utf-8'),
    })

def filterbcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')+ timedelta(days=1)
    bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-id')[:50]
    trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="ord orderrow {'text-danger' if not i.isdelivered else ''}"  startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="ajaxpage('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
    #         <td>{ i.order_no }</td>
    #         <td>{ i.date.strftime('%d/%m/%Y') }</td>
    #         <td>{ i.client.name if i.client else '' }</td>
    #         <td>{ i.client.code if i.client else '' }</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.client.region if i.client else '' }</td>
    #         <td>{ i.client.city if i.client else '' }</td>
    #         <td>{ i.client.soldbl if i.client else '' }</td>
    #         <td>{ i.salseman }</td>
    #         <td>
    #         {'Non' if i.isclientcommnd else 'Oui'}

    #         </td>
    #         <td>
    #           {'R1' if i.isdelivered else 'R0'}
    #         </td>
    #       </tr>
    #     '''
    print('>>>>>', bons)
    ctx={
        'bons':render(request, 'bclist.html', {'bons':bons, 'loadmore':True, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate')}).content.decode('utf-8')
    }
    if bons:
        ctx['total']=round(Order.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)


def deletebonachat(request):

    bon=Itemsbysupplier.objects.get(pk=request.GET.get('id'))
    bontotal=bon.total
    bonsupplier=bon.supplier
    bonsupplier.rest=round(float(bonsupplier.rest)-float(bontotal), 2)
    bonsupplier.total=round(float(bonsupplier.total)-float(bontotal), 2)
    bonsupplier.save()
    items=Stockin.objects.filter(nbon=bon)
    for i in items:
        i.product.stocktotal=int(i.product.stocktotal)-int(i.quantity)
        if bon.isfacture:
            i.product.stockfacture=int(i.product.stockfacture)-int(i.quantity)

    items.delete()
    bon.delete()
    return JsonResponse({
        'success':True
    })



def searchforjv(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(bon__bon_no__iregex=term))
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    trs=''
    print('>>>>>>', products)
    for i in products:
        trs+=f'''
        <tr class="journalvente-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref.upper()}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def searchforjvfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term))
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="totaljvfc">{i.total}</td>
            <td></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def filterjvdate(request):
    target=request.GET.get('target')
    productid=request.GET.get('productid') or None
    startdate=request.GET.get('datefrom') or None
    enddate=request.GET.get('dateto') or None
    print('>> productid, dateto, datefrom', productid, startdate, enddate, target)
    # startdate = datetime.strptime(startdate, '%Y-%m-%d') or None
    # enddate = datetime.strptime(enddate, '%Y-%m-%d') or None
    # if target=="s":
    #     if productid==None:
    #         print('>> produuct is none')
    #         if enddate==None:
    #             bons=Sortieitem.objects.order_by('-date')
    #         else:
    #             bons=Sortieitem.objects.filter(date__range=[startdate, enddate]).order_by('-date')
    #     else:
    #         if enddate==None:
    #             bons=Sortieitem.objects.filter(product_id=productid).order_by('-date')
    #             print('>> produuct is not none', bons)
    #         else:
    #             bons=Sortieitem.objects.filter(product_id=productid, date__range=[startdate, enddate]).order_by('-date')
    #     ctx={
    #         'trs':render(request, 'journalventetrs.html', {'bons':bons, 'target':target}).content.decode('utf-8')
    #     }
        
        
    #     if bons:
    #         ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
    #         ctx['totalqty']=round(bons.aggregate(Sum('qty')).get('qty__sum'), 2)
    #     return JsonResponse(ctx)
    # else:
    #     if productid==None:
    #         print('>> produuct is none')
    #         if enddate==None:
    #             bons=Bonlivraison.objects.filter(isfarah=target=="f").order_by('-date')
    #             bons=[Livraisonitem.objects.filter(bon=i).order_by('-date') for i in bons]
    #         else:
    #             bons=Livraisonitem.objects.filter(isfarah=target=="f", date__range=[startdate, enddate]).order_by('-date')
    #             bons=[Livraisonitem.objects.filter(bon=i).order_by('-date') for i in bons]
    #     else:
    #         print('>> produuct is not none')
    #         bons=Bonlivraison.objects.filter(isfarah=target=="f", product_id=productid,  date__range=[startdate, enddate]).order_by('-date')
    #         bons=[Livraisonitem.objects.filter(bon=i).order_by('-date') for i in bons]

    if target=='f':
        if productid==None:
            print('>> produuct is none in farah')
            if enddate==None:
                bons=Livraisonitem.objects.filter(bon__isfarah=True).order_by('-date')
            else:
                bons=Livraisonitem.objects.filter(bon__isfarah=True, date__range=[startdate, enddate]).order_by('-date')
        else:
            print('>> produuct is not none  in farah')
            bons=Livraisonitem.objects.filter(product_id=productid, bon__isfarah=True, date__range=[startdate, enddate]).order_by('-date')

    elif target=='o':
        if productid==None:
            print('>> produuct is none in orgh')
            if enddate==None:
                bons=Livraisonitem.objects.filter(bon__isfarah=False).order_by('-date')
            else:
                bons=Livraisonitem.objects.filter(bon__isfarah=False, date__range=[startdate, enddate]).order_by('-date')
        else:
            print('>> produuct is not none  in orgh')
            if enddate==None:
                bons=Livraisonitem.objects.filter(bon__isfarah=False, product_id=productid).order_by('-date')
            else:
                bons=Livraisonitem.objects.filter(bon__isfarah=False,  product_id=productid, date__range=[startdate, enddate]).order_by('-date')
    else:
        if productid==None:
            print('>> produuct is none')
            if enddate==None:
                bons=Sortieitem.objects.order_by('-date')
            else:
                bons=Sortieitem.objects.filter(date__range=[startdate, enddate]).order_by('-date')
        else:
            if enddate==None:
                bons=Sortieitem.objects.filter(product_id=productid).order_by('-date')
                print('>> produuct is not none', bons)
            else:
                bons=Sortieitem.objects.filter(product_id=productid, date__range=[startdate, enddate]).order_by('-date')

        #bons=Sortieitem.objects.filter(product_id=productid, date__range=[startdate, enddate]).order_by('-date')
    #elif target=='o':
    
    # trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
    #         <td>{i.date.strftime('%d/%m/%Y')}</td>
    #         <td>{i.bon.bon_no}</td>
    #         <td>{i.product.ref}</td>
    #         <td>{i.product.name}</td>
    #         <td>{i.price}</td>
    #         <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
    #         <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
    #         <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
    #         <td class="text-danger qtyjv">{i.qty}</td>
    #         <td class="totaljv">{i.total}</td>
    #         <td>{i.bon.client.name}</td>
    #         <td>{i.bon.salseman.name}</td>
    #         <td class="text-success margejv">

    #         </td>
    #     </tr>
    #     '''
    #items = [item for queryset in bons for item in queryset]
    total=round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    totalqty=round(bons.aggregate(Sum('qty')).get('qty__sum') or 0, 2)
    ctx={
        'trs':render(request, 'journalventetrs.html', {'bons':bons, 'target':target}).content.decode('utf-8')
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totalqty']=round(bons.aggregate(Sum('qty')).get('qty__sum'), 2)
    return JsonResponse(ctx)
    
def filterjvfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="totaljvfc">{i.total}</td>
            <td></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
    return JsonResponse(ctx)

def filterjachdate(request):
    target=request.GET.get('target')
    productid=request.GET.get('productid') or None
    startdate=request.GET.get('datefrom') or None
    enddate=request.GET.get('dateto') or None
    print('>> productid, dateto, datefrom', productid, startdate, enddate, target)
    # startdate = datetime.strptime(startdate, '%Y-%m-%d') or None
    # enddate = datetime.strptime(enddate, '%Y-%m-%d') or None
    if target=='f':
        if productid==None:
            print('>> produuct is none')
            if enddate==None:
                bons=Stockin.objects.filter(isfarah=True, isavoir=False).order_by('-date')
            else:
                bons=Stockin.objects.filter(product_id=productid, isfarah=True, isavoir=False, date__range=[startdate, enddate]).order_by('-date')
        else:
            print('>> produuct is not none')
            if enddate==None:
                bons=Stockin.objects.filter(product_id=productid, isfarah=True, isavoir=False).order_by('-date')
            else:
                bons=Stockin.objects.filter(product_id=productid, isfarah=True, isavoir=False, date__range=[startdate, enddate]).order_by('-date')

    elif target=='o':
        if productid==None:
            print('>> produuct is none')
            if enddate==None:
                bons=Stockin.objects.filter(isfarah=False, isavoir=False).order_by('-date')
            else:
                bons=Stockin.objects.filter(isfarah=False, isavoir=False, date__range=[startdate, enddate]).order_by('-date')
        else:
            print('>> produuct is not none')
            if enddate==None:
                bons=Stockin.objects.filter(product_id=productid, isfarah=False, isavoir=False).order_by('-date')
            else:
                bons=Stockin.objects.filter(product_id=productid, isfarah=False, isavoir=False, date__range=[startdate, enddate]).order_by('-date')
    else:
        bons=Sortieitem.objects.filter(product_id=productid, date__range=[startdate, enddate]).order_by('-date')
    
    ctx={
        'trs':render(request, 'journalachattrs.html', {'bons':bons, 'target':target}).content.decode('utf-8')
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totalqty']=round(bons.aggregate(Sum('quantity')).get('quantity__sum'), 2)
    return JsonResponse(ctx)

def searchforjach(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    q_forhistory = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
            q_forhistory &= (Q(fournisseur__iregex=term)|Q(designation__iregex=term)|Q(ref__iregex=term)|Q(mantant__iregex=term))
    # means the year i not selected, so the records of the current year
    products1 = Stockin.objects.filter(q_objects)
    producthistory = Achathistory.objects.filter(q_forhistory)
    products = chain(*[
        ((bon, 0) for bon in products1),
        ((b, 1) for b in producthistory)
    ])
    products=sorted(products, key=lambda item: item[0].date)
    total=round(products1.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(producthistory.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)
    # if year=='0':
    #     print('thisyear')
    #     # means the year i not selected, so the records of the current year
    #     products = Stockin.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')
    #     products = Achathistory.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')
    #     total=round(Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    # else:
    #     products = Stockin.objects.filter(date__year=year).filter(q_objects).order_by('-date')
    #     products = Achathistory.objects.filter(date__year=year).filter(q_objects).order_by('-date')
    #     total=round(Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    print('his>>>', products)
    for i in products:
        print(i[0])
        # trs+=f'''
        # <tr class="jach-row" year={year} term={term}>
        #     <td>{i.date.strftime('%d/%m/%Y')}</td>
        #     <td>{i.product.ref}</td>
        #     <td>{i.product.name}</td>
        #     <td>{i.price}</td>
        #     <td>{i.supplier.name}</td>
        #     <td>{i.devise}</td>
        #     <td class="qtyjournalachat">{i.quantity}</td>
        #     <td class="totaljournalachat">{i.total}</td>
        # </tr>
        # '''
        trs+=f'''
        <tr style="background:{'#ffd79c'if i[1]==1 else ''};"">
            <td>{i[0].date.strftime('%d/%m/%Y')}</td>
            <td>{i[0].product.ref if i[1]==0 else i[0].ref}</td>
            <td>{i[0].product.name if i[1]==0 else i[0].designation}</td>
            <td>{i[0].price if i[1]==0 else i[0].prixunitaire}</td>
            <td>{i[0].supplier.name if i[1]==0 else i[0].fournisseur}</td>
            <td>{i[0].devise}</td>
            <td class="qtyjournalachat">{i[0].quantity}</td>
            <td class="totaljournalachat">{i[0].total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajach(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalacha-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def filterjachfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="jachfc-row" startdat={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('quantity'))['quantity__sum']
    return JsonResponse(ctx)

def searchforjachfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
    if year=='0':
        print('thisyear')
        # means the year i not selected, so the records of the current year
        products = Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    else:
        products = Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="jachfc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajachfc(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(facture=True, date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="jachfc-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def updaterepdata(request):
    region=request.POST.get('region')
    caneditprice=True if request.POST.get('caneditprice')=='on' else False
    slides=True if request.POST.get('slides')=='on' else False
    repid=request.POST.get('repid')
    data={
    'repid':repid,
    'caneditprice':caneditprice
    }
    # uncomment try excep if ther is server
    # try:
        # res=req.get('http://serverip/products/updaterepdata', data)
        # res.raise_for_status()
    # except:
    #     # in case connection failed
    #     return JsonResponse({
    #     'success':False,
    #     'here':'rr'
    #     })
    print(region, caneditprice, slides, repid)
    rep=Represent.objects.get(pk=repid)
    rep.region=region
    rep.caneditprice=caneditprice
    rep.slides=slides
    rep.save()

    return JsonResponse({
        'success':True,
        'here':'rr'

    })

def makebondelivered(request):
    id=request.GET.get('id')
    bon=Bonlivraison.objects.get(pk=id)
    bon.isdelivered=True
    bon.save()
    return JsonResponse({
        'success':True
    })

def getitemsforlistbl(request):
    term=request.GET.get('term')
    target=request.GET.get('target')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(diametre__icontains=term)|
            Q(cars__icontains=term)
        )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects)
    brands = [product.mark for product in products]
    categories = [product.category for product in products]
    print('>>>>', brands)
    unique_categories = set(categories)
    unique_brands = set(brands)
    brands = [{'id': mark.id, 'name': mark.name, 'image':mark.image.url if mark.image else '/media/default.png'} for mark in unique_brands]
    categories = [{'id': category.id, 'name': category.name, 'image':category.image.url if category.image else '/media/default.png'} for category in unique_categories]
    # trs=[f'''<tr class="productsbrand{i.mark.id if i.mark else ''}">
    # <td><img src={i.image.url if i.image else ''}></td>
    # <td>{i.ref.upper()}</td>
    # <td>{i.name.upper()}</td>
    # <td style="color: #ff6409;
    # font-weight: bold;">{i.stocktotalfarah if target=='f' el} </td>
    # <td style="color:blue;font-weight: bold;">{i.sellprice}</td>
    # <td>{i.remise}</td>
    # <td>{i.prixnet}</td>
    # <td>{i.diametre}</td>
    # </tr>
    # ''' for i in products]
    return JsonResponse({
        'trs':render(request, 'product_search.html', {'products':products, 'target':target}).content.decode('utf-8'),
        'brands':brands,
        'categories':categories
    })
def refspage(request):
    # res=req.get('http://serverip/products/refspage')
    # print(res)

    refs=Refstats.objects.all().order_by('-lastdate')
    return render(request, 'refspage.html', {
        'refs':refs,
        #'refserver':json.loads(res.text)['refs']
    })

def updateadmindata(request):
    from django.contrib.auth.hashers import make_password
    username=request.POST.get('adminusername')
    password=request.POST.get('adminpassword')
    # Check if the username already exists
    existing_user = User.objects.filter(username=username).exclude(id=request.user.id).first()
    print('>>>>>>>',existing_user)
    if existing_user:
        return JsonResponse({
            'success': False,
            'error': 'Username already exists.'
        })
    else:
        # Update user data
        print('>>> updating')
        user = User.objects.get(pk=request.user.id)
        user.username = username
        user.password = make_password(password)  # Ensure the password is securely hashed
        user.save()

        return JsonResponse({
            'success': True,
        })

def notavailable(request):
    ctx={
    'products':Notavailable.objects.all()
    }
    return render(request, 'notavailable.html', ctx)

def cartpage(request):
    #res=req.get('http://serverip/products/getcarts')

    ctx={
        'carts':Cart.objects.all().order_by('-total').exclude(total=0),
        #'cartsserver':list(json.loads(res.text)['carts'])
    }
    return render(request, 'cartspage.html', ctx)

def reliquatpage(request):
    #res=req.get('http://serverip/products/getwishs')
    print(list(json.loads(res.text)['carts']))
    ctx={
        'carts':Wich.objects.all().order_by('-total'),
        #'wishserver':list(json.loads(res.text)['carts'])
    }
    return render(request, 'reliquatpage.html', ctx)

def excelnotav(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        name = d.name
        mark = None if pd.isna(d.mark) else d.mark
        #refeq = '' if pd.isna(d.refeq) else d.refeq
        #status = False if pd.isna(d.status) else True
        #coderef = '' if pd.isna(d.code) else d.code
        #diam = '' if pd.isna(d.diam) else d.diam
        #qty = 0 if pd.isna(d.qty) else d.qty
        #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        #devise = 0 if pd.isna(d.devise) else d.devise
        # car = None if pd.isna(d.car) else d.car
        #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        #ctg = None if pd.isna(d.ctg) else d.ctg
        #order = '' if pd.isna(d.order) else d.order
        #img = None if pd.isna(d.img) else d.img
        #remise = 0 if pd.isna(d.remise) else d.remise
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        product=Produit.objects.filter(ref=ref).first()
        if not product:

            exis=Notavailable.objects.filter(ref=ref).first()
            if not exis:
                product=Notavailable.objects.create(
                    ref=ref,
                #     equivalent=refeq,
                #     isactive=False,
                #     coderef=coderef,
                    name=name,
                    mark_id=mark,
                #     sellprice=prixbrut,
                #     prixnet=prixnet,
                #     remise=remise,
                #     category_id=ctg,
                    image=f'/products_imags/tette.jpg',
                #     stockfacture=qty,
                #     #diametre=diam,
                #     buyprice=buyprice,
                #     devise=devise
                )
            else:
                print(ref, 'exist in hors stock')
        else:
            print(ref, 'exist in products')

    return JsonResponse({
        'success':True
    })

def deletereglbl(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    regl=PaymentClientbl.objects.get(pk=reglid)
    client=regl.client
    if password=='0000':
        client.soldbl=round(float(client.soldbl)+float(regl.amount))
        client.soldtotal=round(float(client.soldtotal)+float(regl.amount))
        client.save()
        bons=regl.bons.all()
        for i in bons:
            otherregl=PaymentClientbl.objects.filter(bons=i.id).exclude(pk=reglid)
            if not otherregl.exists():
                print(i.bon_no, 'not mentioned so it will be paid=false')
                i.ispaid=False
                i.statusreg='n1'
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })

def deletereglsupp(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    # get reglement supp
    regl=PaymentSupplier.objects.get(pk=reglid)
    #get supplier
    supp=regl.supplier
    print(reglid, password, regl.amount, supp)
    if password=='0000':
        # add amount to supplier sold
        print('adding amount to supplier sold')
        supp.rest+=regl.amount
        supp.save()
        bons=regl.bons.all()
        for i in bons:
            otherregl=PaymentSupplier.objects.filter(bons=i.id).exclude(pk=reglid)
            if not otherregl.exists():
                print('bon not exist')
                i.ispaid=False
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })



def deletereglfc(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    regl=PaymentClientfc.objects.get(pk=reglid)
    print('reglem', regl.client.name, regl.amount)
    client=regl.client
    if password=='0000':
        client.soldfacture=round(float(client.soldfacture)+float(regl.amount))
        client.soldtotal=round(float(client.soldtotal)+float(regl.amount))
        client.save()
        bons=regl.factures.all()
        for i in bons:
            # search in other reglement
            otherregl=PaymentClientfc.objects.filter(factures=i.id).exclude(pk=reglid)

            if not otherregl.exists():
                i.ispaid=False
                i.statusreg='n1'
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })
def changeclientbl(request):
    oldclient=request.GET.get('oldclient')
    orderid=request.GET.get('orderid')
    newclient=request.GET.get('newclient')
    total=request.GET.get('total')
    oldclient=Client.objects.get(pk=oldclient)
    newclient=Client.objects.get(pk=newclient)
    order=Bonlivraison.objects.get(pk=orderid)
    order.client=newclient
    order.save()
    print(oldclient.name, newclient.name, total)
    #extract total from soldbl  and soldtotal of old client
    oldclient.soldbl=float(oldclient.soldbl)-float(total)
    oldclient.soldtotal=float(oldclient.soldtotal)-float(total)
    oldclient.save()
    # add total to soldbl and soldtotal of newclient
    newclient.soldbl=float(newclient.soldbl)+float(total)
    newclient.soldtotal=float(newclient.soldtotal)+float(total)
    newclient.save()
    return JsonResponse({
        'success':True
    })

def comptable(request):
    factureid=request.GET.get('factureid')
    facture=Facture.objects.get(pk=factureid)
    facture.isaccount=not facture.isaccount
    facture.save()
    return JsonResponse({
        'success':True
    })

def deletsearchedref(request):
    refid=request.GET.get('refid')
    ref=Refstats.objects.get(pk=refid)
    ref.delete()
    return JsonResponse({
        'success':True
    })
def getcompatbilse(request):
    year=request.GET.get('year') or '2024'
    print('>>>>>>>', year, Facture.objects.filter(date__year=year, isaccount=True))
    factures=Facture.objects.filter(date__year=year, isaccount=True).order_by('-facture_no')[:50]
    total=round(Facture.objects.filter(date__year=year, isaccount=True).aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(Facture.objects.filter(date__year=year, isaccount=True).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    trs=''
    for i in factures:
        trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} " comptable="1" ondblclick="ajaxpage('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td >
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td>
                <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                </td>
            </tr>
            '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totaltva':totaltva,
    })

def listnotifications(request):
    ctx={
        'title':'List Notifications',
        'notifications':Notification.objects.all()
    }
    return render(request, 'listnotifications.html', ctx)

def addnotification(request):
    notification=request.GET.get('notification')
    try:
        #req.get('http://serverip/products/addnotification', {'notificationid':notificationid,'notification':notification})

        Notification.objects.create(notification=notification)
        return JsonResponse({
            'success':True
            })
    except:
        print('>>>>>>> error')
        return JsonResponse({
            'success':False
            })

def updatenotification(request):
    notificationid=request.GET.get('notificationid')
    notification=request.GET.get('notification')
    print(notificationid, notification)
    return JsonResponse({
        'success':True
    })
    # try:
    #     req.get('http://serverip/products/updatenotification', {'notificationid':notificationid,'notification':notification})
    #     notif=Notification.objects.get(pk=notificationid)
    #     notif.notification=notification
    #     notif.save()
    #     return JsonResponse({
    #         'success':True
    #     })
    # except:
    #     print('>>>>>>> error')
    #     return JsonResponse({
    #         'success':False
    #         })

def updatefacturenote(request):
    factureid=request.GET.get('factureid')
    note=request.GET.get('note')
    facture=Facture.objects.get(pk=factureid)
    facture.note=note
    facture.save()
    return JsonResponse({
        'success':True
        })

def updatefacturerep(request):
    factureid=request.GET.get('factureid')
    rep=request.GET.get('rep')
    facture=Facture.objects.get(pk=factureid)
    facture.salseman_id=rep
    facture.save()
    return JsonResponse({
        'success':True
        })

def updateproductstock(request):
    stock=request.GET.get('stock')
    productid=request.GET.get('productid')
    product=Produit.objects.get(pk=productid)
    diff=int(stock)-int(product.stocktotal)
    Modifierstock.objects.create(stock=diff, product=product)
    # req.get('http://serverip/products/updatepdctdata', {

    #     'id':productid,
    #     'ref':product.ref,
    #     'stocktotal':stock,
    # })
    product.stocktotal=stock
    product.save()
    return JsonResponse({
        'success':True
    })

def updateproductstockfc(request):
    stock=request.GET.get('stock')
    productid=request.GET.get('productid')
    product=Produit.objects.get(pk=productid)
    diff=int(stock)-int(product.stockfacture)
    Modifierstock.objects.create(stock=diff, product=product, stockfc=True)

    product.stockfacture=stock
    product.save()
    return JsonResponse({
        'success':True
    })

def updatetransportbl(request):
    transport=request.GET.get('transport')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.modlvrsn=transport
    bon.save()
    return JsonResponse({
        'success':True
        })
def updatenotebl(request):
    note=request.GET.get('note')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.notebon=note
    bon.save()
    return JsonResponse({
        'success':True
       } )
def updatenotebs(request):
    note=request.GET.get('note')
    blid=request.GET.get('blid')
    bon=Bonsortie.objects.get(pk=blid)
    bon.note=note
    bon.save()
    return JsonResponse({
        'success':True
       } )

def updatecarbs(request):
    car=request.GET.get('car')
    blid=request.GET.get('blid')
    bon=Bonsortie.objects.get(pk=blid)
    bon.car=car
    bon.save()
    return JsonResponse({
        'success':True
       } )

def updaterepbl(request):
    rep=request.GET.get('rep')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.salseman_id=rep
    bon.save()
    return JsonResponse({
        'success':True
        })

def stockupdated(request):
    ctx={
    'products':Modifierstock.objects.all()
    }
    return render(request, 'stockupdated.html', ctx)

def getclientcode(request):
    factureno=request.GET.get('factureno')
    facture=Facture.objects.get(facture_no__icontains=factureno)
    return JsonResponse({
        'code':facture.client.code
    })

def allowcatalog(request):
    clientcode=request.GET.get('clientcode')
    try:
        # res=req.get('http://serverip/products/allowcatalog', {
        #     'clientcode':clientcode,
        # })
        # print(res)
        # res.raise_for_status()
        print('>><W>>>>>>>>>>>>>>>>>')
        client=Client.objects.get(code=clientcode)
        client.accesscatalog=not client.accesscatalog
        client.save()
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False
        })

def filterepbons(request):
    repid=request.GET.get('repid')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    bontarget=request.GET.get('bontarget')

    if bontarget=='1':
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=True).order_by('-id')
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)



        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=True).order_by('-id')
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)


    elif bontarget=='0':
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=False).order_by('-id')
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)



        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=False).order_by('-id')
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)

    else:
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid).exclude(total=0).order_by('-id')
        # this gets only bons from tablete
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)

        systembons=bons.exclude(commande__isnull= False, commande__isclientcommnd=False)

        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid).order_by('-id')
        # this gets only bons from tablete
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)
        systemfactures=factures.exclude(bon__commande__isnull= False, bon__commande__isclientcommnd=False)
    totalbl=bons.aggregate(Sum('total'))['total__sum'] or 0
    totalfc=factures.aggregate(Sum('total'))['total__sum'] or 0
    totalblfctable=round(totalbl+totalfc, 2)
    avoirs=Avoirclient.objects.filter(date__range=[startdate, enddate], representant_id=repid)
    systemtotalblfc=round(systembons.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(systemfactures.aggregate(Sum('total'))['total__sum'] or 0, 2)
    reptotalblfc=round(repbons.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(repfactures.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalavoirs=round(avoirs.aggregate(Sum('total'))['total__sum'] or 0, 2)
    systemresultttc=systemtotalblfc-totalavoirs
    represultttc=reptotalblfc-totalavoirs
    systemresultht=round(systemtotalblfc/1.2, 2)
    #systemresultht=round(systemresultttc/1.2, 2)
    #represultht=round(represultttc/1.2, 2)
    represultht=round(represultttc/1.2, 2)
    print('>>>', bons)
    trsbons=[f"""
    <tr
    style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
    class="ord {"text-danger" if i.ispaid else ''}"
    orderid="{i.id}">
        <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td style="color: blue;">{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ "%.2f" % i.client.soldbl} </td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
            <div>
            {'R0' if i.ispaid else 'N1' }

            </div>
            <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td>
            {'OUI' if i.isfacture else 'NON'}
            </td>

            <td>
                {i.commande.order_no if i.commande else "---"}
            </td>
            <td>
            {i.modlvrsn}
            </td>
            <td>
            {i.note}
            </td>
            <td>

            </td>

      </tr>
    """ for i in bons]
    trsfactures=[f"""
    <tr class="ord {"text-danger" if i.ispaid else ''}">
        <td>{ i.facture_no }</td>
        <td>{ i.date.strftime("%d/%m/%Y")}</td>
        <td>{ i.client.name }</td>
        <td>{ i.client.code }</td>
        <td>{ i.total}</td>
        <td>{ i.client.region}</td>
        <td>{ i.client.city}</td>
        <td>{ i.client.soldfacture:.2f}</td>
        <td>{ i.salseman }</td>
        <td class="d-flex justify-content-between">
        <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

        </td>
        <td></td>
        <td>
        {i.bon.bon_no if i.bon else "--"}
        </td>

        <td class="d-flex">

        </td>
        <td>
        {i.note}
        </td>
    </tr>
    """ for i in factures]
    return JsonResponse({
        'totalbons':round(systembons.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalfactures':round(systemfactures.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalblfc':systemtotalblfc,
        'totalavoirs':totalavoirs,
        'resultttc':systemtotalblfc,
        'resultht':systemresultht,

        'reptotalblfc':reptotalblfc,
        #'represultttc':represultttc,
        'represultttc':represultttc,
        #'represultht':represultht,
        'represultht':represultht,
        'totalrepbons':round(repbons.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalrepfactures':round(repfactures.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'bons':''.join(trsbons),
        'factures':''.join(trsfactures),
        'totalblfctable':totalblfctable
    })
def updateavoirnote(request):
    avoirid=request.GET.get('avoirid')
    note=request.GET.get('note')
    print('>>>>>>>', avoirid, note)
    avoir=Avoirclient.objects.get(pk=avoirid)
    avoir.note= note
    avoir.save()
    return JsonResponse({
        'success':True
    })

def getreliquatcommande(request):
    bons=Order.objects.filter(note__icontains='Reliquat', isdelivered=False)
    print(bons)
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })
#this to indicate if the facture is given to client
def printed(request):
    factureid=request.GET.get('factureid')
    facture=Facture.objects.get(pk=factureid)
    facture.printed=True
    facture.save()
    return JsonResponse({
        'success':True
    })

def getnongenerer(request):
    bons=Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat')
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def addrepnote(request):
    note=request.GET.get('note')
    repid=request.GET.get('repid')
    repnote=Notesrepresentant.objects.filter(represent_id=repid).first()
    if repnote:
        repnote.note=note
        repnote.save()
    else:
        Notesrepresentant.objects.create(represent_id=repid, note=note)
    return JsonResponse({
        'success':True
    })

def alertreliquatcommande(request):
    orders=Orderitem.objects.filter(order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False)
    bons=[i.order for i in orders]
    bons=set(bons)
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),

    })

def makeitemsent(request):
    orderitemid=request.GET.get('orderitemid')
    password=request.GET.get('password')
    if password=="1574":
        orderitem=Orderitem.objects.get(pk=orderitemid)
        orderitem.islivraison=True
        orderitem.save()
        return JsonResponse({
            'success':True
        })
    return JsonResponse({
        'success':False,
        'error':'Mot de pass incorrect'
    })

def updatedateavsupp(request):
    avoirid=request.GET.get('avoirsupid')
    date=request.GET.get('date')
    datebon=datetime.strptime(date, '%Y-%m-%d')
    avoirsupp=Avoirsupplier.objects.get(pk=avoirid)
    avoirsupp.date=datebon
    avoirsupp.save()
    print('>>>>', datebon, avoirsupp.date, avoirsupp.total)
    return JsonResponse({
        'success':True
    })

def tsgs(request):
    import random
    most_sold_items = Livraisonitem.objects.values('product__id', 'product__name').annotate(total_sold=Sum('qty')).order_by('-total_sold')[:10]
    most_sold_products = [(item['product__name'], item['total_sold']) for item in most_sold_items]
    for product_name, total_sold in most_sold_products:
        print(f">>* {''.join([random.sample(product_name, 1)[0] for i in range(len(product_name))])} {product_name}: {total_sold} units")
    sldsbl=round(PaymentClientbl.objects.aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    sldsfc=round(PaymentClientfc.objects.aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    sldsup=round(PaymentSupplier.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)
    nt=sldsbl+sldsfc-sldsup

    # Annotate and aggregate to get the total per day
    totals_per_day = (Bonlivraison.objects
                      .annotate(day=TruncDay('date'))
                      .values('day')
                      .annotate(total=Sum('total'))
                      .order_by('-total'))

    # Get the day with the highest total
    if totals_per_day.exists():
        highest_total_day = totals_per_day.first()
        print(">>> Day with the highest total:", highest_total_day['day'], "with total:", highest_total_day['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per client
    totals_per_client = (Bonlivraison.objects
                         .values('client')
                         .annotate(total=Sum('total'))
                         .order_by('-total'))

    # Get the client with the highest total
    if totals_per_client.exists():
        highest_total_client_data = totals_per_client.first()
        highest_total_client = Client.objects.get(id=highest_total_client_data['client'])
        print(">>> Client with the highest total:", highest_total_client, "with total:", highest_total_client_data['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per product
    totals_per_product = (Livraisonitem.objects
                          .values('product')
                          .annotate(total=Sum('total'))
                          .order_by('-total'))

    # Get the product with the highest total
    if totals_per_product.exists():
        highest_total_product_data = totals_per_product.first()
        highest_total_product = Produit.objects.get(id=highest_total_product_data['product'])
        print(">>> Product with the highest total:", highest_total_product, "with total:", highest_total_product_data['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per client
    totals_per_client = (Avoirclient.objects
                         .values('client')
                         .annotate(total=Sum('total'))
                         .order_by('-total'))

    # Get the client with the highest total
    if totals_per_client.exists():
        highest_total_client_data = totals_per_client.first()
        highest_total_client = Client.objects.get(id=highest_total_client_data['client'])
        print(">>> Client with the highest avoir total:", highest_total_client, "with total:", highest_total_client_data['total'])
    else:
        print("No data available.")
    print(nt)
    print(nt/7)
    return JsonResponse({
        'rr':'rr'
    })
def bonlivraisonprint(request, id):
    isfarah=request.GET.get('target')=='f'
    print('<>> target', request.GET.get('target'))
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__name')
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+38] for i in range(0, len(orderitems), 38)]
    ctx={
        'isfarah':isfarah,
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'bonlivraisonprint.html', ctx)

def bonlivraisonprintht(request, id):
    isfarah=request.GET.get('target')=='f'
    category=request.GET.get('category')=='1'
    print('<>> target', request.GET.get('target'))
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__name')
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+38] for i in range(0, len(orderitems), 38)]
    ctx={
        'isfarah':isfarah,
        'category':category,
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'bonlivraisonprintht.html', ctx)



def printdevi(request):
    isfarah=request.GET.get('target')=='f'
    deviid=request.GET.get('deviid')
    print('<>> target', request.GET.get('target'))
    order=Devi.objects.get(pk=deviid)
    orderitems=DeviItem.objects.filter(devi=order).order_by('product__name')
    
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+38] for i in range(0, len(orderitems), 38)]
    ctx={
        'isfarah':isfarah,
        'title':f'devi {order.bon_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'printdevi.html', ctx)

def printboncommand(request):
    isfarah=request.GET.get('target')=='f'
    deviid=request.GET.get('deviid')
    print('<>> target', request.GET.get('target'))
    order=Command.objects.get(pk=deviid)
    orderitems=CommandItem.objects.filter(command=order).order_by('product__name')
    
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+38] for i in range(0, len(orderitems), 38)]
    ctx={
        'isfarah':isfarah,
        'title':f'devi {order.bon_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'printboncommand.html', ctx)



def sortieprint2(request, id):
    a6=request.GET.get('a6')=='1'
    order=Bonsortie.objects.get(pk=id)
    orderitems=Sortieitem.objects.filter(bon=order).order_by('product__name')
    orderitems=list(orderitems)
    if a6:
        orderitems=[orderitems[i:i+15] for i in range(0, len(orderitems), 15)]
    else:
        orderitems=[orderitems[i:i+25] for i in range(0, len(orderitems), 25)]
    
    ctx={
        'A6':a6,
        'title':f'Bon de Sortie {order.bon_no}',
        'order':order,
        'orderitems':orderitems
    }
    if order.remise:
        ctx['isremise']=True
        ctx['net']=round(order.total-order.remiseamount, 2)
    return render(request, 'sortieprint2.html', ctx)

def boncmndprint(request, id):
    order=Order.objects.get(pk=id)
    orderitems=Orderitem.objects.filter(order=order).order_by('product__name')


    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+45] for i in range(0, len(orderitems), 45)]
    ctx={
        'title':f'Bon commande {order.order_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'boncmndprint.html', ctx)

def factureprint(request, id):
    names=request.GET.get('names')=='1'
    isfarah=request.GET.get('target')=='f'
    order=Facture.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon__in=order.bons.all())
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+25] for i in range(0, len(orderitems), 25)]
    hasespece=PaymentClientbl.objects.filter(factures__in=[order], mode='espece').exists()
    print('>> hasespece')
    ht=order.total/1.2
    tva=ht*0.2
    if hasespece:
        #dr=round(PaymentClientbl.objects.filter(factures__in=[order], mode='espece')[0].amount*.25/100, 2)
        dr=round(order.total*.25/100, 2)
        netapy=order.total+dr
    ctx={
        'isfarah':isfarah,
        'names':names,
        'hasespece':hasespece,
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':tva,
        'ttc':order.total,
        'ht':ht,
    }
    if hasespece:
        ctx['dr']=dr
        ctx['netapy']=netapy
    else:
        ctx['netapy']=order.total
    return render(request, 'factureprint.html', ctx)

def achatprint(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    payments=PaymentSupplier.objects.filter(bons__in=[bon])
    orderitems=[items[i:i+36] for i in range(0, len(items), 36)]
    # ctx={
    #     'title':f'Bon de livraison {order.bon_no}',
    #     'order':order,
    #     'orderitems':orderitems,
    #     'reglements':reglements,
    #     'reps':Represent.objects.all()
    # }
    ctx={
        'title':f'Bon achat {bon.nbon}',
        'bon':bon,
        'items':items,
        'orderitems':orderitems,
        'payments':payments
    }
    return render(request, 'bonachatprint.html', ctx)

def refusedreglfc(request):
    reglfcid=request.GET.get('reglfcid')
    regl=PaymentClientfc.objects.get(pk=reglfcid)
    regl.refused=True
    regl.save()
    return JsonResponse({
        'success':True
    })

def refusedreglbl(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientbl.objects.get(pk=reglid)
    regl.refused=True
    regl.ispaid=False
    regl.save()
    bank=regl.targetbank
    bank.total-=regl.amount
    bank.save()
    regl.targetbank=None
    return JsonResponse({
        'success':True
    })

def refusedreglsupp(request):
    reglid=request.GET.get('reglid')
    regl=PaymentSupplier.objects.get(pk=reglid)
    regl.refused=True
    regl.ispaid=False
    bank=regl.targetbank
    bank.total+=regl.amount
    bank.save()
    regl.targetbank=None
    regl.save()
    return JsonResponse({
        'success':True
    })

def relevblprint(request):
    clientid=request.GET.get('clientid')
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    target=request.GET.get('target')
    isfarah=target=='f'
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    # print(clientid, startdate, enddate)
    # return JsonResponse({
    #     'rr':'rr'
    #     })
    client=Client.objects.get(pk=clientid)
    avoirs=Avoirclient.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)

    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'releveclprint.html', {
            'releve':[sorted_releve[i:i+36] for i in range(0, len(sorted_releve), 36)],
            'client':client,
            'title':"Situation client",
            'startdate':startdate,
            'enddate':enddate,

        })

def relevsuppprint(request):
    supplierid=request.GET.get('supplierid')
    isrelevefacture=request.GET.get('facture')=='1'
    target=request.GET.get('target')
    print('>>>', supplierid)
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    if target=="f":
        avoirs=Avoirsupplier.objects.filter(isfarah=True, supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate], ispaid=False)
        reglementsbl=PaymentSupplier.objects.filter(isfarah=True, supplier_id=supplierid, date__range=[startdate, enddate], isavoir=False)
        avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=True, date__range=[startdate, enddate])
        if isrelevefacture:
            bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=True)
            print('>>>>>>>>>>>< infacture relve farah', bons)
        else:
            bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=True)
    else:
        # orgh
        avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate], ispaid=False)
        reglementsbl=PaymentSupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isavoir=False, isfarah=False)
        avances=Avancesupplier.objects.filter(supplier_id=supplierid, isfarah=False, date__range=[startdate, enddate])
        if isrelevefacture:
            bons=Factureachat.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
        else:
            bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfarah=False)
    totaldebit = round(bons.aggregate(Sum('total'))['total__sum'], 2)
    totalcredit = round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    totalsold=round(totaldebit-totalcredit, 2)
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'relevesuppprint.html', {
            'releve':[sorted_releve[i:i+30] for i in range(0, len(sorted_releve), 30)],
            'supplier':supplier,
            'totaldebit':totaldebit,
            'totalcredit':totalcredit,
            'startdate':startdate,
            'enddate':enddate,
            'totalsold':totalsold
        })

def relevfcprint(request):
    target=request.GET.get('target')
    
    clientid=request.GET.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    if target=="f":
        print('>>> here in farah', clientid, )
        bons=Facture.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=True)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isfarah=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isfarah=True)
    elif target=="s":
        bons=Bonsortie.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], issortie=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], issortie=True)
    else:
        bons=Facture.objects.filter(client_id=clientid,  date__range=[startdate, enddate], isfarah=False)
        avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate], isorgh=True)
        avances=Avanceclient.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True)
        reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate], isorgh=True)
        #bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0, isfarah=isfarah)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    print('>> bons, target', bons, target)
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((avance, 'avanceclient') for avance in avances),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'releveclfcprint.html', {
            'releve':[sorted_releve[i:i+36] for i in range(0, len(sorted_releve), 36)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,
        })

def cancelpdctnew(request):
    pdctid=request.GET.get('pdctid')
    product=Produit.objects.get(pk=pdctid)
    product.isnew = False
    product.save()
    return JsonResponse({
        'success':True
        })

def makepdctnew(request):
    pdctid=request.GET.get('pdctid')
    product=Produit.objects.get(pk=pdctid)
    product.isnew = True
    product.save()
    return JsonResponse({
        'success':True
        })

def filterclients(request):
    rep=request.GET.get('rep')
    region=request.GET.get('region')
    print(rep, region)
    if region=="":
        clients=Client.objects.filter(represent_id=rep, soldtotal__gt=0)
    else:
        clients=Client.objects.filter(represent_id=rep, region__icontains=region, soldtotal__gt=0)
    print(clients.count())
    return JsonResponse({
        'success':True,
        'html':render(request, 'clientslist.html', {'clients':clients}).content.decode('utf-8')
        })

def pdctrepport(request):
    ctx={
    'title':'Rapport des produits',
    'categories':Category.objects.all()
    }
    return render(request, 'pdctrepport.html', ctx)

# this is used in pdct rapports, it gives the outs of the product
def getpdctins(request):

    ref=request.GET.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref).first()
    # avoirs
    products1 = Stockin.objects.filter(product=product)

    producthistory = Achathistory.objects.filter(ref=ref)
    products = chain(*[
        ((bon, 0) for bon in products1),
        ((b, 1) for b in producthistory)
    ])

    pdctins=sorted(products, key=lambda item: item[0].date)
    pdctins_serializable = [
        {
            'date': item[0].date.strftime('%d/%m/%Y'),
            'total': item[0].total,
            'supplier': item[0].supplier.name if item[1] == 0 else '--',
            'quantity': item[0].quantity,
            'price': f"{item[0].price:.2f}" if item[1] == 0 else item[0].prixunitaire,
            'devise': item[0].devise,
            'type': 'Stockin' if item[1] == 0 else 'Achathistory'
        }
        for item in pdctins
    ]
    total=round(products1.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(producthistory.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)
    return JsonResponse({
        'success':True,
        'pdctins':pdctins_serializable,
        'totalqtyin':totalqty,
        'total':total
    })

# this is used in pdct rapports, it gives the outs of the product
def getpdctouts(request):
    ref=request.GET.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref).first()
    stockout=Livraisonitem.objects.filter(product=product, isfacture=False)
    stockoutfc=Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True)
    avoirs=Stockin.objects.filter(product=product)

    totalqtyavoirs=avoirs.aggregate(Sum('qty'))['qty__sum'] or 0
    totalavoirs=avoirs.aggregate(Sum('total'))['total__sum'] or 0
    releve = chain(*[
    ((outbl, 0) for outbl in stockout),
    ((outfc, 1) for outfc in stockoutfc),
    ])
    fortable = chain(*[
    ((outbl, 0) for outbl in stockout),
    ((outfc, 1) for outfc in stockoutfc),
    ])
    pdctins=sorted(releve, key=lambda item: item[0].date)
    print('>>pdctins', pdctins)
    listfortable=sorted(fortable, key=lambda item: item[0].date, reverse=True)
    grouped_by_month = groupby(pdctins, key=lambda item: item[0].date.strftime('%m/%Y'))
    # Prepare data for frontend
    by_month = []
    for month, items in grouped_by_month:
        items=[i[0].qty for i in items]
        count = sum(items)  # Counting items in each group
        by_month.append({'month': month, 'count': count})
    pdctouts_serializable = [
        {
            'date': item[0].date.strftime('%d/%m/%Y'),
            'total': item[0].total,
            'price': f"{item[0].price:.2f}",
            'remise': item[0].remise,
            'client': item[0].bon.client.name if item[1]==0 else item[0].facture.client.name,
            'quantity': item[0].qty,
            'type': 'Livraisonitem' if item[1] == 0 else 'Outfacture',
            'no': item[0].bon.bon_no if item[1] == 0 else item[0].facture.facture_no,
        }
        for item in listfortable
    ]
    totalqty=(stockout.aggregate(Sum('qty'))['qty__sum'] or 0)+(stockoutfc.aggregate(Sum('qty'))['qty__sum'] or 0)
    total=round(stockout.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(stockoutfc.aggregate(Sum('total'))['total__sum'] or 0, 2)
    # Group by client

    client_quantities= defaultdict(int)
    client_avoirs= defaultdict(int)
    for i in avoirs:
        client_name = i.avoir.client.name
        client_avoirs[client_name] += i.qty
    for item in pdctins:
        client_name = item[0].bon.client.name if item[1] == 0 else item[0].facture.client.name
        client_id = item[0].bon.client.id if item[1] == 0 else item[0].facture.client.id
        client_quantities[client_name] += item[0].qty
        #client_quantities[client_name][1] = Stockin.objects.filter(avoir__client_id=client_id, product=product).aggregate(Sum('qty'))['qty__sum'] or 0
        #client_data[client_name]['quantity'] += item[0].qty

    print('>>>>>> client_quantities', client_avoirs)
    clients_quantities_serializable = sorted([
    {'client': client, 'quantity': quantity}
    for client, quantity in client_quantities.items()
    ], key=lambda x: x['quantity'], reverse=True)[:10]
    clients_avoirs_serializable = sorted([
    {'client': client, 'quantity': quantity}
    for client, quantity in client_avoirs.items()
    ], key=lambda x: x['quantity'], reverse=True)
    return JsonResponse({
        'pdctstock':product.stocktotal if product.stocktotal else '--',
        'pdctimg':product.image.url if product.image else '--',
        'pdctname':product.name,
        'success':True,
        'totalavoirs':totalavoirs,
        'pdctouts':pdctouts_serializable,
        'totalqtyout':totalqty,
        'totalqtyavoirs':totalqtyavoirs,
        'totalout':total,
        'outbymonth':by_month,
        'clientsqty':clients_quantities_serializable,
        'clientsavoirs':clients_avoirs_serializable
    })

def pdctscategoryrepport(request):
    # category = Category.objects.get(pk=request.POST.get('category'))
    # products = category.product.filter(category=category)[:10]
    # get ten products from the category
    products = Produit.objects.filter(category__pk=request.GET.get('categoryid')).order_by('-stocktotal')
    print(products)
    pdctdata=[]
    print('start loop')
    for i in products:
        stockout=Livraisonitem.objects.filter(product=i, isfacture=False)
        stockoutfc=Outfacture.objects.filter(product=i).exclude(facture__bon__isnull=True)
        totalqtyout=(stockout.aggregate(Sum('qty'))['qty__sum'] or 0)+(stockoutfc.aggregate(Sum('qty'))['qty__sum'] or 0)
        products1 = Stockin.objects.filter(product=i)
        producthistory = Achathistory.objects.filter(ref=i.ref)
        totalqtyin=(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)
        pdctdata.append({
        'id':i.id,
        'ref':i.ref,
        'name':i.name,
        'devise':i.devise if i.devise else 0,
        'arrivage':i.minstock if i.minstock else 0,
        'newfob':i.newfob,
        'qtycommande':i.qtycommande,
        'stocktotal':i.stocktotal,
        'totalqtyout':totalqtyout,
        'totalqtyin':totalqtyin,
        })
    print('end loop')
    return JsonResponse({
        'pdctdata':pdctdata
    })

def commandpdct(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    pdct.qtycommande=qty
    pdct.save()
    return JsonResponse({
    'success':True,
    })
# minstock is now the arrivage, here to determine the quantity in port or arriving soon
def minstockpdct(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    pdct.minstock=qty
    pdct.save()
    return JsonResponse({
    'success':True,
    })
# new fob is the new devise of the product
def newfob(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    print(request.GET.get('pdctid'), qty)
    pdct.newfob=qty
    pdct.save()
    return JsonResponse({
        'success':True,
    })



def etude(request):
    ctx={
    'title':'Etude',
    }
    return render(request, 'etude.html', ctx)

def boncomparer(request):
    bons=[f'BL24040{i}' for i in range(10, 100)]
    for i in bons:
        bon=Bonlivraison.objects.get(bon_no=i)
        items=Livraisonitem.objects.filter(bon=bon, isfacture=False)
        totals=[b.total for b in items]
        if not bon.total==round(sum(totals), 2):
            print('>>>', bon.bon_no, bon.total, round(sum(totals), 2))
            print('>>', totals)
    return JsonResponse({
    'success':True
    })

def getachatfacture(request):
    year=request.GET.get('year') or thisyear
    print('year >>', year)
    bons=Itemsbysupplier.objects.filter(date__year=year, isfacture=True).order_by('-date')
    print(bons)
    total=round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(bons.aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    trs=''
    for order in bons:
        trs+=f'''
            <tr class="" orderid="{order.id}" ondblclick="ajaxpage('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>
                <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>
            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
            '''


    return JsonResponse({
    'success':True,
    'trs':trs,
    'total':total
    })
def etatblfc(request):
    current_year = datetime.now().year
    target=request.GET.get('target')
    # Create a date object for the first day of the current year
    title = "etat client general ORGH"
    clients = Client.objects.filter(clientorgh=True)
    if target == "f":
        title = "etat client general FARAH"
        clients = Client.objects.filter(clientfarah=True)
    return render(request, 'etatblfc.html', {'title': title, 'target':target, 'clients':clients})

        # Define start and end months for the date range
        # sitdata=0
        # client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}
        #
        # # Filter data for the specified date range
        # factures = Facture.objects.filter(
        #     client=client,
        #     date__range=[start_date, end_date],
        #     total__gt=0
        # )
        #
        # avoirs = Avoirclient.objects.filter(
        #     client=client,
        #     avoirfacture=True,
        #     date__range=[start_date, end_date]
        # )
        #
        # regls = PaymentClientfc.objects.filter(
        #     client=client,
        #     date__range=[start_date, end_date]
        # )
        #
        # # Initialize monthly data
        # monthly_data = {month: {'factures': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}
        #
        # # Populate monthly data with factures, avoirs, and regls
        # for bon in factures:
        #     month = bon.date.strftime('%m/%y')
        #     monthly_data[month]['factures'] += bon.total
        #     sitdata+=bon.total
        #
        # for avoir in avoirs:
        #     month = avoir.date.strftime('%m/%y')
        #     monthly_data[month]['avoirs'] += avoir.total
        #     sitdata-=bon.total
        #
        # for regl in regls:
        #     month = regl.date.strftime('%m/%y')
        #     monthly_data[month]['regls'] += regl.amount
        #     sitdata-=bon.total
        # # Calculate the client situation for each month and aggregate totals
        # total_factures = total_avoirs = total_regls = 0
        #
        # for monthindex, month in enumerate(months):
        #     total_factures += monthly_data[month]['factures']
        #     total_avoirs += monthly_data[month]['avoirs']
        #     total_regls += monthly_data[month]['regls']
        #     #monthly_data[month]['situation'] = round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
        #     thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
        #     thisreg=monthly_data[month]['regls']
        #     if thisreg:
        #         for b in range(monthindex):
        #             thisreg-=client_data['monthly_data'][b]['situation']
        #             client_data['monthly_data'][b]['situation']=0
        #     thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - thisreg, 2)
        #
        #
        #
        #     client_data['monthly_data'].append({
        #         'month': month,
        #         'factures': monthly_data[month]['factures'],
        #         'avoirs': monthly_data[month]['avoirs'],
        #         'regls': monthly_data[month]['regls'],
        #         'situation': thismonthsit
        #     })
        #
        # # Calculate total situation for the client
        # client_data['totalsituation'] = round(total_factures - total_avoirs - total_regls, 2)
        # serialized_data.append(client_data)
def excelecheaces(request):
    # create new record to get the id


    return render(request, 'excelecheances.html')


def getnpiecedata(request):
    npiece=request.GET.get('npiece')
    data={
    'echance':'',
    'client':'',
    'codeclient':'',
    'mode':'',
    'factures':'',
    'amount':'',
    'tva':'',
    }
    try:
        reg = PaymentClientfc.objects.filter(npiece=npiece).first()
        data['echance']=reg.echance.strftime('%d/%m/%Y')
        data['client']=reg.client.name
        data['codeclient']=reg.client.code
        data['mode']=reg.mode
        data['amount']=reg.amount
        data['factures']=', '.join(list(reg.factures.values_list('facture_no', flat=True)))
        data['tva']=round(reg.amount/6, 2)
    except Exception as e:
        print('>>>>>>', e)
        try:
            reg = PaymentClientbl.objects.filter(npiece=npiece).first()
            data['echance']=reg.echance.strftime('%d/%m/%Y')
            data['client']=reg.client.name
            data['codeclient']=reg.client.code
            data['mode']=reg.mode
            data['amount']=reg.amount
            data['tva']=round(reg.amount/6, 2)
        except:
            pass
    return JsonResponse(data)

def saverowech(request):
    facturesval=request.GET.get('facturesval')
    impye=True if request.GET.get('impye')=='true' else False
    nomde=request.GET.get('nomde')
    regle=True if request.GET.get('regle')=='true' else False
    isempty=False if request.GET.get('isempty')=='False' else True
    npieceval=request.GET.get('npieceval')
    mode=request.GET.get('mode')
    tva=request.GET.get('tva')
    amount=request.GET.get('amount')
    # used to get code from ajax
    code=request.GET.get('code')
    #code=uuid.uuid4().hex
    client=request.GET.get('client')
    codeclient=request.GET.get('codeclient')
    echeance=request.GET.get('echeance')
    monthyear=request.GET.get('monthyear')
    print(f"code: {code}")
    print(f"impye: {impye}")
    print(f"nomde: {nomde}")
    print(f"regle: {regle}")
    print(f"npieceval: {npieceval}")
    print(f"mode: {mode}")
    print(f"amount: {amount}")
    print(f"client: {client}")
    print(f"codeclient: {codeclient}")
    print(f"echeance: {echeance}")
    print(f"monthyear: {monthyear}")
    print(f"isempty: {isempty}")
    # try to get this record
    try:
        data=Excelecheances.objects.get(pk=code)
        print('>>>>>> this', data)
        data.month=monthyear # Assuming 'monthyear' is in the format '09/2024'
        data.npiece=npieceval
        data.mode=mode
        data.tva=tva
        data.amount=amount
        data.echeance=echeance
        data.note=nomde
        data.client=client
        data.clientcode=codeclient
        data.factures=facturesval
        data.ispaid=regle   # Example default values
        data.isimpye=impye   # Example default values
        data.isempty=isempty
        data.save()
    # creatte if not found
    except Exception as e:
        print('>>>>>>>', e)
        data=Excelecheances.objects.create(
            month=monthyear,  # Assuming 'monthyear' is in the format '09/2024'
            npiece=npieceval,
            mode=mode,
            echeance=echeance,
            note=nomde,
            client=client,
            amount=amount,
            tva=tva,
            clientcode=codeclient,
            factures=facturesval,
            ispaid=regle,   # Example default values
            isimpye=impye,   # Example default values
            isempty=isempty,  # Example default value
        )
    return JsonResponse({
    'success':True,
    # code will be id
    'thisid':data.id,
    'nextid':Excelecheances.objects.last().id+1
    })
def getmonthecheances(request):
    month=request.GET.get('month')
    data=Excelecheances.objects.filter(month=month).order_by('code')
    total=data.aggregate(Sum('amount'))['amount__sum'] or 0
    idsource=Excelecheances.objects.filter(npiece=None).first() if Excelecheances.objects.filter(npiece=None) else Excelecheances.objects.create()

    return JsonResponse({
        'trs':render(request, 'echeancestrs.html', {'data':data, 'code':idsource.id}).content.decode('utf-8'),
        'total':total
    })
# gather muliple
def grouper(request):
    ids=json.loads(request.GET.get('ids'))
    print('>>>>>>', ids)
    echeances=Excelecheances.objects.filter(pk__in=ids)
    grandtotal=echeances.aggregate(Sum('amount'))['amount__sum']
    code=uuid.uuid4().hex
    echeances.update(code=code, grandtotal=grandtotal)

    return JsonResponse({
    'success':True
    })
def getqtyprice(request):
    target=request.GET.get('target')
    print(">>target", target)
    id=request.GET.get('id')
    isfarah=target=='f'
    if target=='s':
        term=request.GET.get('term')
        if (term.startswith('fr-')):
            histyory=Stockin.objects.filter(product_id=id, isfarah=True, isavoir=False, qtyofprice__gt=0).order_by('date')
        else:
            histyory=Stockin.objects.filter(product_id=id, isfarah=False, isavoir=False, qtyofprice__gt=0).order_by('date')
    else:
        histyory=Stockin.objects.filter(product_id=id, isfarah=isfarah, isavoir=False, qtyofprice__gt=0).order_by('date')
    return render(request, 'qtyprice.html', {'history':histyory})


def vv(request):
    p=Produit.objects.all()
    for i in p:
        ref=i.ref.lower()
        i.ref=ref
        i.farahref=f'fr-{ref}'
        i.save()
    return JsonResponse({
        'rr':True
    })

def reglsituation(request):
    bons=json.loads(request.GET.get('bons'))
    avoirs=json.loads(request.GET.get('avoirs'))
    avances=json.loads(request.GET.get('avances'))
    clientid=request.GET.get('clientid')
    target=request.GET.get('target')
    print('>> target', target)
    moderegl=request.GET.get('moderegl')
    date=request.GET.get('date')
    if moderegl=='bl':
        if target=='s':
            livraisons=Bonsortie.objects.filter(pk__in=bons)
        else:
            livraisons=Bonlivraison.objects.filter(pk__in=bons)
    else:
        livraisons=Facture.objects.filter(pk__in=bons)
    avoirs=Avoirclient.objects.filter(pk__in=avoirs)
    avances=Avanceclient.objects.filter(pk__in=avances)
    totals=round(avoirs.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(avances.aggregate(Sum('amount')).get('amount__sum') or 0, 2)
    totalbons=round(livraisons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    if not totals==totalbons:
        return JsonResponse({
            'success':False,
            'message':'Le total des avoirs et avances doit etre egale au total des bons/factures'
        })
    regl=PaymentClientbl.objects.create(
        client_id=clientid,
        amount=0,
        #today
        # date=timezone.now().date(),
        date=date,
        mode="Rsituation",
        isfarah=target=='f',
        isorgh=target=='o',
        issortie=target=='s'
    )
    if moderegl=='bl':
        if target=='s':
            regl.bonsortie.set(livraisons)
        else:
            regl.bons.set(livraisons)
    else:
        regl.factures.set(livraisons)
        regl.usedinfacture=True
        regl.save()
    regl.avoirs.set(avoirs)
    regl.avances.set(avances)
    avoirs.update(inreglement=True)
    avances.update(inreglement=True)
    livraisons.update(ispaid=True)
    return JsonResponse({
        'success':True
    })
def showavanceclient(request):
    avanceid=request.GET.get('avanceid')
    target=request.GET.get('target')
    avance=Avanceclient.objects.get(pk=avanceid)
    return render(request, 'showavanceclient.html', {'avanceid':avanceid, 'target':target, 'avance':avance})

def loadbonachat(request):
    target=request.GET.get('target')
    page = int(request.GET.get('page', 1))

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    print('>> start , end', start , end)
    wantvalid=request.GET.get('wanted')=='valid'
    if target=='f':
        bons=Itemsbysupplier.objects.filter(date__year=thisyear, isfarah=True, isvalid=wantvalid).order_by('-id')[start:end]
    elif target=='o':
        bons=Itemsbysupplier.objects.filter(date__year=thisyear, isfarah=False, isvalid=wantvalid).order_by('-id')[start:end]
    return JsonResponse({
        'html':render(request, 'bonachattrs.html', {'bons':bons}).content.decode('utf-8'),
        'has_more': len(bons) == per_page
        
    })
    
def getfacturepaidtype(request):
    isfarah=request.GET.get('target')=='f'
    mode=request.GET.get('mode')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')

    isvalid=request.GET.get('wanted')=='valid'
    filters = {
        'isfarah': isfarah,
        'isvalid': isvalid,
    }

    if startdate and enddate:
        filters['date__range'] = [startdate, enddate]
    if mode=='nonpaye':
        filters['ispaid']=False
    elif mode=='paye':
        filters['ispaid']=True
    else:
        filters['fcreglements__mode']=mode
    factures = Facture.objects.filter(**filters)
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':factures}).content.decode('utf-8'),
        'total':round(factures.aggregate(Sum('total')).get('total__sum') or 0, 2)
    })
def setinventairin(request):
    id=request.GET.get('id')
    qty=request.GET.get('qty')
    target=request.GET.get('target')
    product=Produit.objects.get(pk=id)
    if target=='f':
        product.inventaireinfarah=qty
    else:
        product.inventaireinorgh=qty
    product.save()
    return JsonResponse({
        'success':True
    })
def setinventairout(request):
    id=request.GET.get('id')
    qty=request.GET.get('qty')
    target=request.GET.get('target')
    product=Produit.objects.get(pk=id)
    if target=='f':
        product.inventaireoutfarah=qty
    else:
        product.inventaireoutorgh=qty
    product.save()
    return JsonResponse({
        'success':True
    })

def zz(request):
    reglestock = request.GET.get('reglestock')=='1'
    products=Produit.objects.all()
    data=[]
    for i in products:
        print("==>", i.ref)
        #achat + avoir
        inorgh=Stockin.objects.filter(product=i, isfarah=False, issortie=False).aggregate(Sum('quantity'))['quantity__sum'] or 0 + i.stockinitial

        sortieorgh = Sortieitem.objects.filter(product=i, isfarah=False).aggregate(Sum('qty'))['qty__sum'] or 0
        # avoir fourniss
        avsupporgh=Returnedsupplier.objects.filter(product=i, isfarah=False).aggregate(Sum('qty'))['qty__sum'] or 0

        # bon livraison
        outorgh=Livraisonitem.objects.filter(product=i, bon__isfarah=False).aggregate(Sum('qty'))['qty__sum'] or 0 + avsupporgh

        netorgh = inorgh - outorgh

        infarah=Stockin.objects.filter(product=i, isfarah=True).aggregate(Sum('quantity'))['quantity__sum'] or 0 + i.frstockinitial

        sortiefarah=Sortieitem.objects.filter(product=i, isfarah=True).aggregate(Sum('qty'))['qty__sum'] or 0

        avsuppfarah=Returnedsupplier.objects.filter(product=i, isfarah=True).aggregate(Sum('qty'))['qty__sum'] or 0

        outfarah=Livraisonitem.objects.filter(product=i, isfarah=True).aggregate(Sum('qty'))['qty__sum'] or 0 + sortiefarah + avsuppfarah

        netfarah = infarah - outfarah
        if (netorgh != i.stocktotalorgh):
            if reglestock:
                i.stocktotalorgh = netorgh
                i.save()
            data.append({
                'ref':i.ref,
                'entreeOrgh':inorgh,
                'sortiOrgh':outorgh,
                'netorgh':netorgh,
                'stockorghSYSTEM':i.stocktotalorgh,
                'ERROR':'00000'
                # 'entreeFarah':infarah,
                # 'sortiFarah':outfarah,
                # 'netfarah':netfarah,
                # 'stockfarahSYSTEM':i.stocktotalfarah,
            })
        # else:
        #     data.append({
        #         'ref':i.ref,
        #         'entreeOrgh':inorgh,
        #         'sortiOrgh':outorgh,
        #         'netorgh':netorgh,
        #         'stockorghSYSTEM':i.stocktotalorgh,
        #         # 'entreeFarah':infarah,
        #         # 'sortiFarah':outfarah,
        #         # 'netfarah':netfarah,
        #         # 'stockfarahSYSTEM':i.stocktotalfarah,
        #     })
    return JsonResponse({
        "data":data
    })

def getetatblfc(request):
    start = request.GET.get("start")
    end = request.GET.get("end")
    isfarah = request.GET.get("target")=="f"
    isorgh = request.GET.get("target")=="o"
    if isorgh:
        clients = Client.objects.filter(clientorgh=True)
    if isfarah:
        clients = Client.objects.filter(clientfarah=True)

    data = []
    totalttc, totaltva, totalht, totalavoirttc, totalavoirht, totalnetttc, totalavoirtva=0,0,0,0,0,0,0

    for i in clients:
        factures = Facture.objects.filter(client=i, date__range=[start, end])
        avoirs = Avoirclient.objects.filter(client=i, date__range=[start, end])
        # total factures
        totalfactures = round(factures.aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalttc += float(totalfactures)
        totalfacturesht=round(totalfactures/1.2, 2)
        totalht += float(totalfacturesht)
        totalfacturestva = round(totalfactures-totalfacturesht, 2)
        totaltva += float(totalfacturestva)
        totalavoirs = round(avoirs.aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalavoirttc += float(totalavoirs)
        totalavoirsht=round(totalavoirs/1.2, 2)
        totalavoirht += float(totalavoirsht)
        totalavoirstva = round(totalavoirs-totalavoirsht, 2)
        totalavoirtva += float(totalavoirstva)
        totalnet = round(totalfactures-totalavoirs, 2)
        totalnetttc += float(totalnet)
        if not totalnet == 0:
            data.append({"name":i.name, "ice": i.ice, "totalfactures": totalfactures, "totalfacturestva": totalfacturestva, "totalfacturesht": totalfacturesht, "totalavoirs": totalavoirs, "totalavoirstva": totalavoirstva, "totalavoirsht": totalavoirsht, 'totalnet':totalnet, "totalttc": round(totalttc, 2), "totaltva": round(totaltva, 2), "totalht": round(totalht, 2), "totalavoirttc": round(totalavoirttc, 2), "totalavoirht": round(totalavoirht, 2), "totalnetttc": round(totalnetttc, 2), "totalavoirtva": round(totalavoirtva, 2)})
    print(data)
    return JsonResponse({
        "success":True,
        "html":render(request, 'etatclientgeneraltrs.html', {'data':data}).content.decode("utf-8"),
        "totalttc": round(totalttc, 2),
        "totaltva": round(totaltva, 2),
        "totalht": round(totalht, 2),
        "totalavoirttc": round(totalavoirttc, 2),
        "totalavoirht": round(totalavoirht, 2),
        "totalavoirtva": round(totalavoirtva, 2),
        "totalnetttc": round(totalnetttc, 2)
    })

def etatsuppliers(request):
    target=request.GET.get('target')
    title = "etat Fournisseurs ORGH"
    return render(request, 'etatsuppliers.html', {'title': title, 'target':target})

def getetatsuppliers(request):
    start = request.GET.get("start")
    end = request.GET.get("end")
    isfarah = request.GET.get("target")=="f"
    isorgh = request.GET.get("target")=="o"
    suppliers=Supplier.objects.all()

    data = []
    totalttc, totaltva, totalht, totalavoirttc, totalavoirht, totalnetttc, totalavoirtva=0,0,0,0,0,0,0
    for i in suppliers:
        factures = Factureachat.objects.filter(supplier=i, date__range=[start, end], isfarah=isfarah)
        avoirs = Avoirsupplier.objects.filter(supplier=i, date__range=[start, end], isfarah=isfarah)
        # total factures
        totalfactures = round(factures.aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalttc += float(totalfactures)
        totalfacturesht=round(totalfactures/1.2, 2)
        totalht += float(totalfacturesht)
        totalfacturestva = round(totalfactures-totalfacturesht, 2)
        totaltva += float(totalfacturestva)
        totalavoirs = round(avoirs.aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalavoirttc += float(totalavoirs)
        totalavoirsht=round(totalavoirs/1.2, 2)
        totalavoirht += float(totalavoirsht)
        totalavoirstva = round(totalavoirs-totalavoirsht, 2)
        totalavoirtva += float(totalavoirstva)
        totalnet = round(totalfactures-totalavoirs, 2)
        totalnetttc += float(totalnet)

        if not totalnet == 0:
            data.append({"name":i.name, "ice": i.ice, "totalfactures": totalfactures, "totalfacturestva": totalfacturestva, "totalfacturesht": totalfacturesht, "totalavoirs": totalavoirs, "totalavoirstva": totalavoirstva, "totalavoirsht": totalavoirsht, 'totalnet':totalnet})
    return JsonResponse({
        "success":True,
        "html":render(request, 'etatclientgeneraltrs.html', {'data':data}).content.decode("utf-8"),
        "totalttc": round(totalttc, 2),
        "totaltva": round(totaltva, 2),
        "totalht": round(totalht, 2),
        "totalavoirttc": round(totalavoirttc, 2),
        "totalavoirht": round(totalavoirht, 2),
        "totalavoirtva": round(totalavoirtva, 2),
        "totalnetttc": round(totalnetttc, 2)
    })

def qrcodescanner(request):
    return render(request, 'qrcodescanner.html')